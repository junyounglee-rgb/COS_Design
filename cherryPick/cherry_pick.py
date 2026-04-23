"""
Cherry Pick Propagator — 핵심 로직
커밋 해시를 입력받아 선택한 브랜치에 cherry-pick + push 전파
"""

from __future__ import annotations

import difflib
import io
import queue
import subprocess
import sys
import threading
from dataclasses import dataclass, field
from pathlib import Path
from typing import Literal, Optional

import streamlit as st
import yaml
from openpyxl import load_workbook

# ─── 설정 ─────────────────────────────────────────────────────────────────────

CONFIG_PATH = Path(__file__).parent / "cherry_pick.yaml"
LOCK_FILE = Path(__file__).parent / ".cherry_pick.lock"

DEFAULT_CONFIG: dict = {
    "repo_path": r"D:\COS_Project\cos-data",
    "branches": [
        "main",
        "release_helsinki2",
        "release_helsinki3",
        "release_helsinki_vng",
    ],
}


def load_config() -> dict:
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        return {**DEFAULT_CONFIG, **data}
    return dict(DEFAULT_CONFIG)


def save_config(cfg: dict) -> None:
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False)


# ─── git 헬퍼 ─────────────────────────────────────────────────────────────────

def run_git(args: list[str], cwd: str) -> tuple[int, str, str]:
    """(returncode, stdout, stderr) 반환"""
    result = subprocess.run(
        ["git"] + args,
        cwd=cwd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return result.returncode, result.stdout.strip(), result.stderr.strip()


def current_branch(repo_path: str) -> str:
    _, out, _ = run_git(["rev-parse", "--abbrev-ref", "HEAD"], cwd=repo_path)
    return out.strip()


def validate_commit_hash(repo_path: str, commit_hash: str) -> bool:
    if not commit_hash or ".." in commit_hash:
        return False
    code, _, _ = run_git(["cat-file", "-t", commit_hash], cwd=repo_path)
    return code == 0


def get_commit_info(repo_path: str, commit_hash: str) -> Optional[dict]:
    code, out, _ = run_git(
        ["log", "-1", "--format=%H|%h|%s|%an|%ar", commit_hash],
        cwd=repo_path,
    )
    if code != 0 or not out:
        return None
    parts = out.split("|", 4)
    if len(parts) < 5:
        return None
    return {"hash": parts[0], "short": parts[1], "message": parts[2], "author": parts[3], "date": parts[4]}


def get_commit_files(repo_path: str, commit_hash: str) -> list[str]:
    code, out, _ = run_git(["show", "--name-only", "--format=", commit_hash], cwd=repo_path)
    if code != 0:
        return []
    return [line for line in out.splitlines() if line.strip()]


def get_source_branches(repo_path: str, commit_hash: str) -> list[str]:
    code, out, _ = run_git(["branch", "-r", "--contains", commit_hash], cwd=repo_path)
    if code != 0 or not out:
        return []
    branches = []
    for line in out.splitlines():
        line = line.strip()
        if not line or "HEAD" in line:
            continue
        if line.startswith("origin/"):
            line = line[len("origin/"):]
        branches.append(line)
    return branches


def get_remote_branches(repo_path: str) -> list[str]:
    code, out, _ = run_git(["branch", "-r"], cwd=repo_path)
    if code != 0 or not out:
        return []
    branches = []
    for line in out.splitlines():
        line = line.strip()
        if not line or "HEAD" in line:
            continue
        branch = line[len("origin/"):] if line.startswith("origin/") else line
        if branch.startswith("claude/"):
            continue
        branches.append(branch)
    return branches


def fetch_origin(repo_path: str) -> None:
    run_git(["fetch", "origin", "--quiet"], cwd=repo_path)


def check_needs_pull(repo_path: str, branches: list[str]) -> list[str]:
    needs_pull = []
    for branch in branches:
        code, out, _ = run_git(["rev-list", "--count", f"origin/{branch}"], cwd=repo_path)
        if code != 0:
            continue
        code2, _, _ = run_git(["rev-parse", "--verify", branch], cwd=repo_path)
        if code2 != 0:
            needs_pull.append(branch)
            continue
        code3, count_str, _ = run_git(["rev-list", "--count", f"{branch}..origin/{branch}"], cwd=repo_path)
        try:
            behind = int(count_str)
        except ValueError:
            behind = 0
        if behind > 0:
            needs_pull.append(branch)
    return needs_pull


def run_autopull(target_branches: list[str], log_queue: queue.Queue) -> None:
    autopull_dir = str(Path(__file__).parent.parent / "autopull")
    if autopull_dir not in sys.path:
        sys.path.insert(0, autopull_dir)
    from autopull import Puller, load_config as autopull_load_config  # type: ignore
    cfg = autopull_load_config()
    puller = Puller(
        repo_root=cfg["repo_root"],
        sibling_repos=cfg.get("sibling_repos", []),
        branches=target_branches,
        log_queue=log_queue,
    )
    puller.run()


# ─── Lock ─────────────────────────────────────────────────────────────────────

class LockError(Exception):
    pass

def acquire_lock() -> None:
    if LOCK_FILE.exists():
        raise LockError("이미 실행 중입니다. 잠시 후 다시 시도하세요.")
    LOCK_FILE.write_text("locked", encoding="utf-8")

def release_lock() -> None:
    if LOCK_FILE.exists():
        LOCK_FILE.unlink()


# ─── 데이터 클래스 ─────────────────────────────────────────────────────────────

@dataclass
class BranchResult:
    branch: str
    status: Literal["success", "failed", "skipped"] = "failed"
    message: str = ""
    new_commit: str = ""


# ─── CherryPickPropagator ─────────────────────────────────────────────────────

class CherryPickPropagator:
    def __init__(self, repo_path, commit_hash, targets, log_queue):
        self.repo_path = repo_path
        self.commit_hash = commit_hash
        self.targets = targets
        self.log_queue = log_queue
        self.results: list[BranchResult] = []

    def log(self, msg):
        self.log_queue.put(msg)

    def run(self):
        orig = current_branch(self.repo_path)
        try:
            acquire_lock()
            self._run_inner()
        except LockError as e:
            self.log(f"[오류] {e}")
        except Exception as e:
            self.log(f"[오류] 예상치 못한 오류: {e}")
        finally:
            release_lock()
            self.log(f"\n원래 브랜치로 복귀: {orig}")
            run_git(["checkout", orig], cwd=self.repo_path)
            self.log("✅ 모든 작업 완료!")

    def _run_inner(self):
        self.log(f"커밋: {self.commit_hash}")
        self.log(f"타겟 브랜치: {', '.join(self.targets)}\n")
        for branch in self.targets:
            self.log(f"══ {branch} {'═' * max(1, 40 - len(branch))}")
            self.results.append(self._process_branch(branch))

    def _process_branch(self, branch):
        result = BranchResult(branch=branch)
        code, _, err = run_git(["checkout", branch], cwd=self.repo_path)
        if code != 0:
            result.message = f"checkout 실패: {err[:200]}"
            self.log(f"  ❌ {result.message}")
            return result
        code, _, err = run_git(["pull", "origin", branch], cwd=self.repo_path)
        if code != 0:
            result.message = f"pull 실패: {err[:200]}"
            self.log(f"  ❌ {result.message}")
            return result
        code, out, err = run_git(["cherry-pick", "-x", self.commit_hash], cwd=self.repo_path)
        if code == 0:
            _, new_hash, _ = run_git(["rev-parse", "--short", "HEAD"], cwd=self.repo_path)
            result.new_commit = new_hash
            self.log(f"  ✅ cherry-pick 성공 [{new_hash}]")
        else:
            combined = (out + err).lower()
            if "already applied" in combined or "now empty" in combined or "empty commit" in combined:
                run_git(["cherry-pick", "--skip"], cwd=self.repo_path)
                result.status = "skipped"
                result.message = "이미 적용된 커밋"
                self.log(f"  ⏭️ 스킵 — 이미 적용됨")
                return result
            self.log(f"  ⚠️ 충돌 감지, 자동 해결 시도...")
            if self._resolve_binary_conflict():
                code2, _, err2 = run_git(["cherry-pick", "--continue", "--no-edit"], cwd=self.repo_path)
                if code2 == 0:
                    _, new_hash, _ = run_git(["rev-parse", "--short", "HEAD"], cwd=self.repo_path)
                    result.new_commit = new_hash
                    self.log(f"  ✅ 충돌 자동 해결 후 성공 [{new_hash}]")
                else:
                    run_git(["cherry-pick", "--abort"], cwd=self.repo_path)
                    result.message = f"--continue 실패: {err2[:200]}"
                    self.log(f"  ❌ {result.message}")
                    return result
            else:
                run_git(["cherry-pick", "--abort"], cwd=self.repo_path)
                result.message = f"텍스트 충돌 — 수동 해결 필요: {err[:200]}"
                self.log(f"  ❌ {result.message}")
                return result
        code, _, err = run_git(["push", "origin", branch], cwd=self.repo_path)
        if code != 0:
            result.message = f"push 실패: {err[:200]}"
            self.log(f"  ❌ push 실패 [{result.new_commit}] — 수동 push 필요")
            return result
        self.log(f"  🚀 push 완료")
        result.status = "success"
        return result

    def _resolve_binary_conflict(self):
        code, out, _ = run_git(["diff", "--name-only", "--diff-filter=U"], cwd=self.repo_path)
        if code != 0 or not out:
            return False
        conflicted = [f for f in out.splitlines() if f.strip()]
        if not conflicted:
            return False
        BINARY_EXTS = {".xlsx", ".pb", ".bytes", ".png", ".jpg", ".jpeg", ".gif", ".bmp"}
        text_conflicts = [f for f in conflicted if Path(f).suffix.lower() not in BINARY_EXTS]
        if text_conflicts:
            self.log(f"  텍스트 충돌 파일: {text_conflicts}")
            return False
        self.log(f"  바이너리 충돌 파일: {conflicted} → theirs로 자동 해결")
        for f in conflicted:
            run_git(["checkout", "--theirs", f], cwd=self.repo_path)
        run_git(["add"] + conflicted, cwd=self.repo_path)
        return True


# ─── CherryPickThread ──────────────────────────────────────────────────────────

class CherryPickThread:
    def __init__(self, propagator):
        self.propagator = propagator
        self._done = threading.Event()
        self._thread = threading.Thread(target=self._run, daemon=True)

    def _run(self):
        try:
            self.propagator.run()
        finally:
            self._done.set()

    def start(self): self._thread.start()
    def is_done(self): return self._done.is_set()
    def join(self): self._thread.join()


# ─── Excel Diff ───────────────────────────────────────────────────────────────

def run_git_binary(args, cwd):
    result = subprocess.run(["git"] + args, cwd=cwd, capture_output=True)
    return result.returncode, result.stdout


@st.cache_data(ttl=300, show_spinner=False)
def _load_xlsx_bytes(repo_path: str, ref: str, xlsx_path: str) -> Optional[bytes]:
    """git ref에서 xlsx 바이너리를 추출하여 캐싱 (5분 TTL)"""
    code, data = run_git_binary(["show", f"{ref}:{xlsx_path}"], cwd=repo_path)
    if code != 0 or not data:
        return None
    return data


def load_xlsx_from_ref(repo_path: str, ref: str, xlsx_path: str) -> Optional[object]:
    """git ref에서 xlsx를 openpyxl Workbook으로 로드. bytes는 캐싱됨."""
    data = _load_xlsx_bytes(repo_path, ref, xlsx_path)
    if not data:
        return None
    try:
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None


@st.cache_data(ttl=300, show_spinner=False)
def _cached_compare(repo_path: str, old_ref: str, new_ref: str, xlsx_path: str) -> dict[str, str]:
    """(repo, old_ref, new_ref, xlsx) 조합의 diff HTML을 캐싱 (5분 TTL)"""
    old_wb = load_xlsx_from_ref(repo_path, old_ref, xlsx_path)
    new_wb = load_xlsx_from_ref(repo_path, new_ref, xlsx_path)
    return compare_xlsx_side_by_side(old_wb, new_wb)


@dataclass
class CellChange:
    sheet: str
    row: int
    col: int
    col_header: str
    old_value: str
    new_value: str
    change_type: Literal["added", "changed", "removed"]


MAX_DIFF_CELLS = 500


def _read_sheet_data(ws) -> dict[tuple[int, int], str]:
    data: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                data[(cell.row, cell.column)] = str(cell.value)
    return data


def _get_headers(ws) -> dict[int, str]:
    headers: dict[int, str] = {}
    try:
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value is not None:
                headers[cell.column] = str(cell.value)
    except StopIteration:
        pass
    return headers


def diff_workbooks(old_wb, new_wb, old_label="old", new_label="new") -> list[CellChange]:
    changes: list[CellChange] = []
    if old_wb is None and new_wb is None:
        return changes
    old_sheets = set(old_wb.sheetnames) if old_wb else set()
    new_sheets = set(new_wb.sheetnames) if new_wb else set()
    for sheet_name in sorted(old_sheets | new_sheets):
        old_ws = old_wb[sheet_name] if old_wb and sheet_name in old_sheets else None
        new_ws = new_wb[sheet_name] if new_wb and sheet_name in new_sheets else None
        old_data = _read_sheet_data(old_ws) if old_ws else {}
        new_data = _read_sheet_data(new_ws) if new_ws else {}
        headers = _get_headers(new_ws) if new_ws else (_get_headers(old_ws) if old_ws else {})
        for row, col in sorted(set(old_data) | set(new_data)):
            old_val = old_data.get((row, col), "")
            new_val = new_data.get((row, col), "")
            if old_val == new_val:
                continue
            ct = "added" if not old_val else "removed" if not new_val else "changed"
            changes.append(CellChange(sheet_name, row, col, headers.get(col, f"col{col}"), old_val, new_val, ct))
            if len(changes) >= MAX_DIFF_CELLS:
                return changes
    return changes


def _align_rows(old_data, new_data, header_row=1):
    """
    LCS 기반 행 정렬. 첫 2컬럼 값 튜플을 키로 사용.
    반환: (old_row|None, new_row|None) 쌍 목록
    """
    old_rows = sorted(set(r for r, c in old_data if r != header_row))
    new_rows = sorted(set(r for r, c in new_data if r != header_row))

    def row_key(data, row):
        cols = sorted(c for r, c in data if r == row)
        return tuple(data.get((row, c), "") for c in cols[:2])

    sm = difflib.SequenceMatcher(None, [row_key(old_data, r) for r in old_rows], [row_key(new_data, r) for r in new_rows], autojunk=False)
    aligned = []
    for op, i1, i2, j1, j2 in sm.get_opcodes():
        if op == "equal":
            for i, j in zip(range(i1, i2), range(j1, j2)):
                aligned.append((old_rows[i], new_rows[j]))
        elif op == "replace":
            for k in range(max(i2-i1, j2-j1)):
                aligned.append((old_rows[i1+k] if i1+k < i2 else None, new_rows[j1+k] if j1+k < j2 else None))
        elif op == "delete":
            for i in range(i1, i2): aligned.append((old_rows[i], None))
        elif op == "insert":
            for j in range(j1, j2): aligned.append((None, new_rows[j]))
    return aligned


def compare_xlsx_side_by_side(old_wb, new_wb, max_rows=200) -> dict[str, str]:
    """
    두 워크북을 Beyond Compare 스타일 HTML 테이블로 비교.
    LCS 정렬로 삽입/삭제 행을 올바르게 처리. 변경 행만 표시.
    반환: {sheet_name: html_string}
    """
    if old_wb is None and new_wb is None:
        return {}
    old_sheets = set(old_wb.sheetnames) if old_wb else set()
    new_sheets = set(new_wb.sheetnames) if new_wb else set()
    results = {}
    for sheet_name in sorted(old_sheets | new_sheets):
        old_ws = old_wb[sheet_name] if old_wb and sheet_name in old_sheets else None
        new_ws = new_wb[sheet_name] if new_wb and sheet_name in new_sheets else None
        old_data = _read_sheet_data(old_ws) if old_ws else {}
        new_data = _read_sheet_data(new_ws) if new_ws else {}
        headers = _get_headers(new_ws) if new_ws else (_get_headers(old_ws) if old_ws else {})
        aligned = _align_rows(old_data, new_data)
        changed_pairs = []
        for old_r, new_r in aligned:
            if old_r is None or new_r is None:
                changed_pairs.append((old_r, new_r))
                continue
            old_cols = {c: v for (r, c), v in old_data.items() if r == old_r}
            new_cols = {c: v for (r, c), v in new_data.items() if r == new_r}
            if any(old_cols.get(c, "") != new_cols.get(c, "") for c in set(old_cols) | set(new_cols)):
                changed_pairs.append((old_r, new_r))
        if not changed_pairs:
            continue
        changed_pairs = changed_pairs[:max_rows]
        relevant_cols: set[int] = set(headers.keys())
        for old_r, new_r in changed_pairs:
            if old_r: relevant_cols.update(c for r, c in old_data if r == old_r)
            if new_r: relevant_cols.update(c for r, c in new_data if r == new_r)
        cols = sorted(relevant_cols)
        if not cols:
            continue
        results[sheet_name] = _build_comparison_html(old_data, new_data, headers, cols, changed_pairs)
    return results


_compare_counter = 0


def _build_comparison_html(old_data, new_data, headers, cols, aligned_pairs) -> str:
    """좌우 비교 HTML — 2개 테이블 flex + JS 행 높이/스크롤 동기화"""
    global _compare_counter
    _compare_counter += 1
    uid = f"bc{_compare_counter}"

    def _esc(v): return v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    def _trunc(v, n=40): return v[:n] + "…" if len(v) > n else v
    def _get(data, row, col):
        if row is None: return ""
        return _esc(_trunc(data.get((row, col), "")))

    col_headers = [_esc(_trunc(headers.get(c, f"col{c}"), 20)) for c in cols]
    header_row_html = "".join(f"<th>{h}</th>" for h in ["행"] + col_headers)

    left_rows, right_rows = [], []
    for old_r, new_r in aligned_pairs:
        row_added = old_r is None
        row_removed = new_r is None
        left_label  = str(old_r) if old_r else ""
        right_label = str(new_r) if new_r else ""
        left_cells  = [f"<td class='rn'>{left_label}</td>"]
        right_cells = [f"<td class='rn'>{right_label}</td>"]
        for c in cols:
            old_val = _get(old_data, old_r, c)
            new_val = _get(new_data, new_r, c)
            if row_added:
                left_cells.append("<td class='bg-del'>&nbsp;</td>")
                right_cells.append(f"<td class='bg-add'>{new_val}</td>")
            elif row_removed:
                left_cells.append(f"<td class='bg-del'>{old_val}</td>")
                right_cells.append("<td class='bg-del'>&nbsp;</td>")
            else:
                if old_val == new_val: cls = ""
                elif not old_val: cls = " class='bg-add'"
                elif not new_val: cls = " class='bg-del'"
                else: cls = " class='bg-chg'"
                left_cells.append(f"<td{cls}>{old_val}</td>")
                right_cells.append(f"<td{cls}>{new_val}</td>")
        left_rows.append("<tr>" + "".join(left_cells) + "</tr>")
        right_rows.append("<tr>" + "".join(right_cells) + "</tr>")

    def _make_div(side, rows):
        return (
            f"<div id='{uid}_{side}' style='overflow:auto;max-height:500px'>"
            f"<table style='border-collapse:collapse;width:max-content'>"
            f"<thead><tr>{header_row_html}</tr></thead>"
            f"<tbody>{''.join(rows)}</tbody></table></div>"
        )

    js = f"""<script>
(function(){{
  var L,R;
  function syncHeights(){{
    if(!L||!R) return;
    var lr=L.querySelectorAll('tbody tr'),rr=R.querySelectorAll('tbody tr'),n=Math.min(lr.length,rr.length);
    for(var i=0;i<n;i++){{lr[i].style.height='auto';rr[i].style.height='auto';}}
    for(var i=0;i<n;i++){{var h=Math.max(lr[i].offsetHeight,rr[i].offsetHeight);lr[i].style.height=h+'px';rr[i].style.height=h+'px';}}
  }}
  function syncSize(){{
    if(!L||!R) return;
    var lt=L.querySelector('table'),rt=R.querySelector('table');
    if(!lt||!rt) return;
    var h=Math.max(lt.offsetHeight,rt.offsetHeight);lt.style.minHeight=h+'px';rt.style.minHeight=h+'px';
    var w=Math.max(lt.offsetWidth,rt.offsetWidth);lt.style.minWidth=w+'px';rt.style.minWidth=w+'px';
  }}
  function init(){{
    L=document.getElementById('{uid}_left');R=document.getElementById('{uid}_right');
    if(!L||!R) return;
    syncHeights();syncSize();
    setTimeout(function(){{syncHeights();syncSize();}},500);
    var act=null,tmr=null;
    function rel(){{act=null;tmr=null;}}
    L.addEventListener('scroll',function(){{if(act&&act!==L)return;act=L;R.scrollLeft=L.scrollLeft;R.scrollTop=L.scrollTop;clearTimeout(tmr);tmr=setTimeout(rel,120);}});
    R.addEventListener('scroll',function(){{if(act&&act!==R)return;act=R;L.scrollLeft=R.scrollLeft;L.scrollTop=R.scrollTop;clearTimeout(tmr);tmr=setTimeout(rel,120);}});
  }}
  requestAnimationFrame(function(){{requestAnimationFrame(function(){{init();}});}});
}})();
</script>"""

    css = f"""<style>
#{uid}_left table,#{uid}_right table{{border-collapse:collapse;width:max-content}}
#{uid}_left th,#{uid}_right th{{background:#f0f0f0;border:1px solid #ddd;padding:4px 6px;font-size:12px;white-space:nowrap;position:sticky;top:0;z-index:1}}
#{uid}_left td,#{uid}_right td{{border:1px solid #eee;padding:3px 6px;font-size:12px;white-space:nowrap}}
#{uid}_left .rn,#{uid}_right .rn{{font-weight:600;background:#fafafa}}
#{uid}_left .bg-add,#{uid}_right .bg-add{{background:#d4edda}}
#{uid}_left .bg-del,#{uid}_right .bg-del{{background:#f8d7da}}
#{uid}_left .bg-chg,#{uid}_right .bg-chg{{background:#fff3cd}}
#{uid}_left tr:hover td,#{uid}_right tr:hover td{{background:#f5f5ff!important}}
</style>"""

    return (
        css
        + "<div style='display:flex;gap:8px'>"
        + f"<div style='flex:1;min-width:0'><div style='font-weight:700;margin-bottom:4px;font-size:13px'>◀ 이전 (왼쪽)</div>{_make_div('left', left_rows)}</div>"
        + f"<div style='flex:1;min-width:0'><div style='font-weight:700;margin-bottom:4px;font-size:13px'>▶ 이후 (오른쪽)</div>{_make_div('right', right_rows)}</div>"
        + "</div>" + js
    )
