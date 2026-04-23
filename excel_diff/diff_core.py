"""
Excel Diff Core — cherry_pick.py에서 diff 관련 함수만 독립 추출.
외부 의존성: openpyxl, Python 표준 라이브러리(difflib, subprocess, io)
"""

from __future__ import annotations

import difflib
import io
import subprocess
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl import Workbook


# ─── git 헬퍼 ─────────────────────────────────────────────────────────────────

def run_git_binary(args: list[str], cwd: str) -> tuple[int, bytes]:
    """git 명령을 바이너리 모드로 실행. (returncode, stdout_bytes) 반환"""
    result = subprocess.run(
        ["git"] + args,
        cwd=cwd,
        capture_output=True,
    )
    return result.returncode, result.stdout


def load_xlsx_from_git(repo_path: str, ref: str, xlsx_path: str) -> Optional[Workbook]:
    """
    git show {ref}:{xlsx_path} 로 xlsx 파일을 읽어 openpyxl Workbook 반환.
    실패 시 None 반환.

    Args:
        repo_path: 로컬 git 리포지토리 경로
        ref: git ref (예: "origin/main", "HEAD", "abc1234")
        xlsx_path: 리포 내 상대 경로 (슬래시 구분, 예: "GameData/cookie.xlsx")
    """
    code, data = run_git_binary(["show", f"{ref}:{xlsx_path}"], cwd=repo_path)
    if code != 0 or not data:
        return None
    try:
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None


def load_xlsx_from_path(local_path: str) -> Optional[Workbook]:
    """
    로컬 파일 경로에서 openpyxl Workbook 반환.
    실패 시 None 반환.
    """
    try:
        with open(local_path, "rb") as f:
            data = f.read()
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None


# ─── diff 내부 헬퍼 ────────────────────────────────────────────────────────────

def _read_sheet_data(ws) -> dict[tuple[int, int], str]:
    """시트의 모든 셀을 {(row, col): value_str} 딕셔너리로 반환"""
    data: dict[tuple[int, int], str] = {}
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is not None:
                data[(cell.row, cell.column)] = str(val)
    return data


def _index_by_row(data: dict[tuple[int, int], str]) -> dict[int, dict[int, str]]:
    """
    Flat dict `{(row, col): value}` → `{row: {col: value}}` 로 1회 재인덱싱.
    이후 행 단위 조회가 O(M) 스캔 대신 O(1) dict lookup 으로 가능.
    """
    by_row: dict[int, dict[int, str]] = {}
    for (r, c), v in data.items():
        sub = by_row.get(r)
        if sub is None:
            by_row[r] = {c: v}
        else:
            sub[c] = v
    return by_row


def _detect_header_row(sheet_name: str) -> int:
    """
    시트명에 '#'이 포함되면 1행이 헤더 (현재 동작 유지).
    그 외(일반 게임 데이터 시트)는 3행이 헤더 — 1·2행은 주석/경로 메타 행.
    """
    return 1 if "#" in sheet_name else 3


def _get_headers(ws, header_row: int = 1) -> dict[int, str]:
    """지정된 행에서 컬럼 헤더 추출"""
    headers: dict[int, str] = {}
    try:
        for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row)):
            if cell.value is not None:
                headers[cell.column] = str(cell.value)
    except StopIteration:
        pass
    return headers


def _align_rows(
    old_data: dict[tuple[int, int], str],
    new_data: dict[tuple[int, int], str],
    header_row: int = 1,
    old_by_row: dict[int, dict[int, str]] | None = None,
    new_by_row: dict[int, dict[int, str]] | None = None,
) -> list[tuple[int | None, int | None]]:
    """
    LCS 기반 행 정렬. 첫 2컬럼 값 튜플을 키로 사용.
    반환: (old_row|None, new_row|None) 쌍 목록.
    None = 해당 쪽에 없는 행 (삽입 또는 삭제).

    old_by_row / new_by_row 를 미리 넘기면 재인덱싱을 건너뛴다
    (성능 최적화: 동일 데이터에 대해 중복 인덱싱 방지).
    """
    # 행 단위 인덱스 (없으면 생성)
    if old_by_row is None:
        old_by_row = _index_by_row(old_data)
    if new_by_row is None:
        new_by_row = _index_by_row(new_data)

    # header_row 이하(메타 행 포함) 전부 제외
    old_rows = sorted(r for r in old_by_row if r > header_row)
    new_rows = sorted(r for r in new_by_row if r > header_row)

    def row_key(by_row: dict[int, dict[int, str]], row: int) -> tuple:
        cols_map = by_row.get(row, {})
        key_cols = sorted(cols_map.keys())[:2]
        return tuple(cols_map.get(c, "") for c in key_cols)

    old_keys = [row_key(old_by_row, r) for r in old_rows]
    new_keys = [row_key(new_by_row, r) for r in new_rows]

    sm = difflib.SequenceMatcher(None, old_keys, new_keys, autojunk=False)
    aligned: list[tuple[int | None, int | None]] = []

    for op, i1, i2, j1, j2 in sm.get_opcodes():
        if op == "equal":
            for i, j in zip(range(i1, i2), range(j1, j2)):
                aligned.append((old_rows[i], new_rows[j]))
        elif op == "replace":
            for k in range(max(i2 - i1, j2 - j1)):
                old_r = old_rows[i1 + k] if i1 + k < i2 else None
                new_r = new_rows[j1 + k] if j1 + k < j2 else None
                aligned.append((old_r, new_r))
        elif op == "delete":
            for i in range(i1, i2):
                aligned.append((old_rows[i], None))
        elif op == "insert":
            for j in range(j1, j2):
                aligned.append((None, new_rows[j]))

    return aligned


_compare_counter = 0


def _build_comparison_html(
    old_data: dict, new_data: dict, headers: dict,
    cols: list[int],
    aligned_pairs: list[tuple[int | None, int | None]],
    left_label: str = "◀ 서버 (origin)",
    right_label: str = "▶ 내 로컬 파일",
) -> str:
    """좌우 비교 HTML — 2개 테이블 flex + JS 행 높이 동기화"""
    global _compare_counter
    _compare_counter += 1
    uid = f"bc{_compare_counter}"

    def _esc(v: str) -> str:
        return v.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _trunc(v: str, n: int = 40) -> str:
        return v[:n] + "…" if len(v) > n else v

    def _raw(data: dict, row: int | None, col: int) -> str:
        if row is None:
            return ""
        return _trunc(data.get((row, col), ""))

    def _get(data: dict, row: int | None, col: int) -> str:
        return _esc(_raw(data, row, col))

    def _inline_diff(old: str, new: str) -> tuple[str, str]:
        """문자 단위 인라인 diff. (old_html, new_html) 반환"""
        sm = difflib.SequenceMatcher(None, old, new, autojunk=False)
        old_parts, new_parts = [], []
        for op, i1, i2, j1, j2 in sm.get_opcodes():
            if op == "equal":
                old_parts.append(_esc(old[i1:i2]))
                new_parts.append(_esc(new[j1:j2]))
            elif op == "replace":
                old_parts.append(f"<span class='di-del'>{_esc(old[i1:i2])}</span>")
                new_parts.append(f"<span class='di-ins'>{_esc(new[j1:j2])}</span>")
            elif op == "delete":
                old_parts.append(f"<span class='di-del'>{_esc(old[i1:i2])}</span>")
            elif op == "insert":
                new_parts.append(f"<span class='di-ins'>{_esc(new[j1:j2])}</span>")
        return "".join(old_parts), "".join(new_parts)

    col_headers = [_esc(_trunc(headers.get(c, f"col{c}"), 20)) for c in cols]
    header_row_html = "".join(f"<th>{h}</th>" for h in ["행"] + col_headers)

    left_rows = []
    right_rows = []
    for old_r, new_r in aligned_pairs:
        row_added = old_r is None
        row_removed = new_r is None

        # 빈칸 쪽(상대방)에만 행 번호 숨김
        left_rn  = str(old_r) if old_r else ""
        right_rn = str(new_r) if new_r else ""

        left_cells = [f"<td class='rn'>{left_rn}</td>"]
        right_cells = [f"<td class='rn'>{right_rn}</td>"]

        for c in cols:
            old_val = _get(old_data, old_r, c)
            new_val = _get(new_data, new_r, c)

            if row_added:
                left_cells.append("<td class='bg-del'>&nbsp;</td>")
                right_cells.append(f"<td class='bg-add'>{new_val}</td>")
            elif row_removed:
                left_cells.append(f"<td class='bg-del'>{old_val}</td>")
                right_cells.append("<td class='bg-del'>&nbsp;</td>")
            else:
                if old_val == new_val:
                    left_cells.append(f"<td>{old_val}</td>")
                    right_cells.append(f"<td>{new_val}</td>")
                elif not old_val:
                    left_cells.append("<td class='bg-add'></td>")
                    right_cells.append(f"<td class='bg-add'>{new_val}</td>")
                elif not new_val:
                    left_cells.append(f"<td class='bg-del'>{old_val}</td>")
                    right_cells.append("<td class='bg-del'></td>")
                else:
                    # 문자 단위 인라인 diff
                    old_html, new_html = _inline_diff(_raw(old_data, old_r, c), _raw(new_data, new_r, c))
                    left_cells.append(f"<td class='bg-chg'>{old_html}</td>")
                    right_cells.append(f"<td class='bg-chg'>{new_html}</td>")

        left_rows.append("<tr>" + "".join(left_cells) + "</tr>")
        right_rows.append("<tr>" + "".join(right_cells) + "</tr>")

    left_html = (
        f"<div id='{uid}_left' class='{uid}_scroll' style='overflow:auto;max-height:500px'>"
        "<table style='border-collapse:collapse;width:max-content'>"
        f"<thead><tr>{header_row_html}</tr></thead>"
        "<tbody>" + "".join(left_rows) + "</tbody></table></div>"
    )
    right_html = (
        f"<div id='{uid}_right' class='{uid}_scroll' style='overflow:auto;max-height:500px'>"
        "<table style='border-collapse:collapse;width:max-content'>"
        f"<thead><tr>{header_row_html}</tr></thead>"
        "<tbody>" + "".join(right_rows) + "</tbody></table></div>"
    )

    # JS: 행 높이 동기화 + syncTableSize + 스크롤 동기화
    js = f"""<script>
(function(){{
  var L,R;
  function syncHeights(){{
    if(!L||!R) return;
    var lRows=L.querySelectorAll('tbody tr');
    var rRows=R.querySelectorAll('tbody tr');
    var n=Math.min(lRows.length,rRows.length);
    for(var i=0;i<n;i++){{
      lRows[i].style.height='auto';
      rRows[i].style.height='auto';
    }}
    for(var i=0;i<n;i++){{
      var h=Math.max(lRows[i].offsetHeight,rRows[i].offsetHeight);
      lRows[i].style.height=h+'px';
      rRows[i].style.height=h+'px';
    }}
  }}
  function syncTableSize(){{
    if(!L||!R) return;
    var lt=L.querySelector('table');
    var rt=R.querySelector('table');
    if(!lt||!rt) return;
    var h=Math.max(lt.offsetHeight,rt.offsetHeight);
    lt.style.minHeight=h+'px';
    rt.style.minHeight=h+'px';
    var w=Math.max(lt.offsetWidth,rt.offsetWidth);
    lt.style.minWidth=w+'px';
    rt.style.minWidth=w+'px';
  }}
  function init(){{
    L=document.getElementById('{uid}_left');
    R=document.getElementById('{uid}_right');
    if(!L||!R) return;
    syncHeights();
    syncTableSize();
    setTimeout(function(){{ syncHeights(); syncTableSize(); }},500);
    // 스크롤 동기화 (active 추적 방식)
    var active=null,timer=null;
    function release(){{active=null; timer=null;}}
    L.addEventListener('scroll',function(){{
      if(active&&active!==L) return;
      active=L;
      R.scrollLeft=L.scrollLeft; R.scrollTop=L.scrollTop;
      clearTimeout(timer); timer=setTimeout(release,120);
    }});
    R.addEventListener('scroll',function(){{
      if(active&&active!==R) return;
      active=R;
      L.scrollLeft=R.scrollLeft; L.scrollTop=R.scrollTop;
      clearTimeout(timer); timer=setTimeout(release,120);
    }});
  }}
  requestAnimationFrame(function(){{
    requestAnimationFrame(function(){{
      init();
    }});
  }});
}})();
</script>"""

    css = f"""<style>
#{uid}_left table, #{uid}_right table {{border-collapse:collapse;width:max-content}}
#{uid}_left th, #{uid}_right th {{background:#f0f0f0;border:1px solid #ddd;padding:4px 6px;font-size:12px;white-space:nowrap;position:sticky;top:0;z-index:1}}
#{uid}_left td, #{uid}_right td {{border:1px solid #eee;padding:3px 6px;font-size:12px;white-space:nowrap}}
#{uid}_left .rn, #{uid}_right .rn {{font-weight:600;background:#fafafa}}
#{uid}_left .bg-add, #{uid}_right .bg-add {{background:#d4edda}}
#{uid}_left .bg-del, #{uid}_right .bg-del {{background:#f8d7da}}
#{uid}_left .bg-chg, #{uid}_right .bg-chg {{background:#fff3cd}}
#{uid}_left .di-del {{color:#cc0000;font-weight:600}}
#{uid}_right .di-ins {{color:#006600;font-weight:600}}
#{uid}_left tr:hover td, #{uid}_right tr:hover td {{background:#f5f5ff!important}}
</style>"""

    return (
        css
        + "<div style='display:flex;gap:8px'>"
        + f"<div style='flex:1;min-width:0'><div style='font-weight:700;margin-bottom:4px;font-size:13px'>{left_label}</div>{left_html}</div>"
        + f"<div style='flex:1;min-width:0'><div style='font-weight:700;margin-bottom:4px;font-size:13px'>{right_label}</div>{right_html}</div>"
        + "</div>"
        + js
    )


# ─── 메인 diff 함수 ────────────────────────────────────────────────────────────

MAX_DIFF_ROWS = 200


def compare_xlsx_side_by_side(
    old_wb,
    new_wb,
    max_rows: int = MAX_DIFF_ROWS,
    left_label: str = "◀ 서버 (origin)",
    right_label: str = "▶ 내 로컬 파일",
) -> dict[str, str]:
    """
    두 워크북을 Beyond Compare 스타일 HTML 테이블로 비교.
    LCS 정렬로 삽입/삭제 행을 올바르게 처리. 변경 행만 표시.

    Returns:
        {sheet_name: html_string} — 변경 없는 시트는 포함되지 않음
    """
    results: dict[str, str] = {}

    if old_wb is None and new_wb is None:
        return results

    old_sheets = set(old_wb.sheetnames) if old_wb else set()
    new_sheets = set(new_wb.sheetnames) if new_wb else set()
    all_sheets = sorted(old_sheets | new_sheets)

    for sheet_name in all_sheets:
        old_ws = old_wb[sheet_name] if old_wb and sheet_name in old_sheets else None
        new_ws = new_wb[sheet_name] if new_wb and sheet_name in new_sheets else None

        old_data = _read_sheet_data(old_ws) if old_ws else {}
        new_data = _read_sheet_data(new_ws) if new_ws else {}

        # 행 단위 인덱스를 시트 루프 최상단에서 1회만 생성 → 이후 O(1) 조회
        old_by_row = _index_by_row(old_data)
        new_by_row = _index_by_row(new_data)

        header_row = _detect_header_row(sheet_name)
        headers = (
            _get_headers(new_ws, header_row) if new_ws
            else (_get_headers(old_ws, header_row) if old_ws else {})
        )

        aligned = _align_rows(
            old_data, new_data, header_row=header_row,
            old_by_row=old_by_row, new_by_row=new_by_row,
        )

        changed_pairs: list[tuple[int | None, int | None]] = []
        for old_r, new_r in aligned:
            if old_r is None or new_r is None:
                changed_pairs.append((old_r, new_r))
                continue
            old_cols = old_by_row.get(old_r, {})
            new_cols = new_by_row.get(new_r, {})
            all_c    = set(old_cols) | set(new_cols)
            if any(old_cols.get(c, "") != new_cols.get(c, "") for c in all_c):
                changed_pairs.append((old_r, new_r))

        if not changed_pairs:
            continue

        changed_pairs = changed_pairs[:max_rows]

        relevant_cols: set[int] = set(headers.keys())
        for old_r, new_r in changed_pairs:
            if old_r:
                relevant_cols.update(old_by_row.get(old_r, {}).keys())
            if new_r:
                relevant_cols.update(new_by_row.get(new_r, {}).keys())
        cols = sorted(relevant_cols)
        if not cols:
            continue

        html = _build_comparison_html(
            old_data, new_data, headers, cols, changed_pairs,
            left_label=left_label, right_label=right_label,
        )
        results[sheet_name] = html

    return results
