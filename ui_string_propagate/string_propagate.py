"""
UI String 다중 브랜치 자동 전파 - 핵심 로직
"""

from __future__ import annotations

import io
import os
import shutil
import subprocess
import threading
import time
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import yaml
from openpyxl import load_workbook

# ──────────────────────────────────────────────────────────────
# 설정 로드
# ──────────────────────────────────────────────────────────────

CONFIG_PATH = Path(__file__).parent / "propagate_branches.yaml"
TMP_BASE = Path(__file__).parent / "_tmp"
LOCK_FILE = Path(__file__).parent / ".ui_string_propagate.lock"
ROLLBACK_LOG_PATH = Path(__file__).parent / "last_propagation.json"
XLSX_REL = "excel/ui_strings.xlsx"

DEFAULT_CONFIG = {
    "branches": [],
    "repo_root": r"D:\COS_Project\cos-data",
    "datasheet": r"D:\COS_Project\cos-data\datasheet.exe",
    "sibling_repos": [],
    "dry_run": False,
}


def load_config() -> dict:
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}
        cfg = {**DEFAULT_CONFIG, **data}
    else:
        cfg = dict(DEFAULT_CONFIG)
    return cfg


def save_config(cfg: dict) -> None:
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        yaml.dump(cfg, f, allow_unicode=True, default_flow_style=False)


# ──────────────────────────────────────────────────────────────
# git 헬퍼
# ──────────────────────────────────────────────────────────────

def run_git(args: list[str], cwd: str) -> tuple[int, str, str]:
    result = subprocess.run(
        ["git"] + args,
        cwd=cwd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return result.returncode, result.stdout.strip(), result.stderr.strip()


def current_branch(repo_root: str) -> str:
    _, out, _ = run_git(["branch", "--show-current"], cwd=repo_root)
    return out.strip()


def check_branch_exists(repo_root: str, branch: str) -> bool:
    code, _, _ = run_git(
        ["ls-remote", "--exit-code", "--heads", "origin", branch],
        cwd=repo_root,
    )
    return code == 0


# ──────────────────────────────────────────────────────────────
# xlsx 파싱 (ui_strings 전용)
# ──────────────────────────────────────────────────────────────

def find_header_row(ws) -> tuple[int, int, int]:
    """
    ^id 컬럼을 가진 행을 찾아 (header_row_1based, id_col_0based, text_col_0based) 반환.
    COS 구조: Row1=설명, Row2=export경로, Row3=헤더
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        for col_idx, cell in enumerate(row):
            if cell is not None and str(cell).startswith("^"):
                id_col = col_idx
                text_col = None
                for c2, val in enumerate(row):
                    if val == "text":
                        text_col = c2
                        break
                if text_col is None:
                    # ^id 옆 다음 컬럼을 text로 가정
                    text_col = id_col + 1
                return row_idx, id_col, text_col
    # fallback: row3, col0, col1
    return 3, 0, 1


def load_strings_from_wb(wb) -> dict[str, str]:
    ws = wb.active if "ui_strings" not in wb.sheetnames else wb["ui_strings"]
    header_row, id_col, text_col = find_header_row(ws)
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        key = row[id_col] if id_col < len(row) else None
        val = row[text_col] if text_col < len(row) else None
        if key and not str(key).startswith("#"):
            result[str(key)] = str(val) if val is not None else ""
    return result


def load_strings_from_file(xlsx_path: str | Path) -> dict[str, str]:
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    return load_strings_from_wb(wb)


def load_strings_from_head(repo_root: str, ref: str = "main", rel_path: str = XLSX_REL) -> dict[str, str] | None:
    """특정 git ref 기준 xlsx 로드. 기본값은 main 브랜치."""
    result = subprocess.run(
        ["git", "show", f"{ref}:{rel_path}"],
        cwd=repo_root,
        capture_output=True,
    )
    if result.returncode != 0:
        return None
    try:
        wb = load_workbook(io.BytesIO(result.stdout), read_only=True, data_only=True)
        return load_strings_from_wb(wb)
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────
# 변경 사항 계산
# ──────────────────────────────────────────────────────────────

@dataclass
class StringDiff:
    added: dict[str, str] = field(default_factory=dict)       # key → new_val
    changed: dict[str, tuple[str, str]] = field(default_factory=dict)  # key → (old, new)
    removed: dict[str, str] = field(default_factory=dict)     # key → old_val

    def is_empty(self) -> bool:
        return not (self.added or self.changed or self.removed)


def compute_diff(repo_root: str, xlsx_path: str | Path) -> StringDiff:
    # 항상 main 브랜치 기준으로 비교 (현재 체크아웃 브랜치와 무관)
    old = load_strings_from_head(repo_root, ref="main")
    new = load_strings_from_file(xlsx_path)

    diff = StringDiff()
    if old is None:
        diff.added = dict(new)
        return diff

    for k, v in new.items():
        if k not in old:
            diff.added[k] = v
        elif old[k] != v:
            diff.changed[k] = (old[k], v)

    for k, v in old.items():
        if k not in new:
            diff.removed[k] = v

    return diff


# ──────────────────────────────────────────────────────────────
# 잠금 파일
# ──────────────────────────────────────────────────────────────

class LockError(Exception):
    pass


def acquire_lock() -> None:
    if LOCK_FILE.exists():
        try:
            content = LOCK_FILE.read_text(encoding="utf-8")
            raise LockError(f"이미 실행 중입니다. ({content})")
        except LockError:
            raise
        except Exception:
            raise LockError("잠금 파일이 존재합니다. 이전 실행이 비정상 종료됐을 수 있습니다.")
    LOCK_FILE.write_text(
        f"PID={os.getpid()} TIME={time.strftime('%Y-%m-%d %H:%M:%S')}",
        encoding="utf-8",
    )


def release_lock() -> None:
    if LOCK_FILE.exists():
        LOCK_FILE.unlink(missing_ok=True)


# ──────────────────────────────────────────────────────────────
# 브랜치 전파 결과
# ──────────────────────────────────────────────────────────────

@dataclass
class BranchResult:
    branch: str
    success: bool = False
    skipped: bool = False
    skip_reason: str = ""
    error_stage: str = ""
    error_detail: str = ""
    commit_hash: str = ""


# ──────────────────────────────────────────────────────────────
# 전파 실행
# ──────────────────────────────────────────────────────────────

class Propagator:
    """
    브랜치 전파 실행기.
    log_callback(str)을 통해 실시간 로그를 전달한다.
    """

    def __init__(
        self,
        repo_root: str,
        datasheet_exe: str,
        xlsx_path: str,
        branches: list[str],
        commit_msg: str,
        sibling_repos: list[str] | None = None,
        dry_run: bool = False,
        log_callback: Callable[[str], None] | None = None,
    ):
        self.repo_root = Path(repo_root)
        self.datasheet_exe = Path(datasheet_exe)
        self.xlsx_path = Path(xlsx_path)
        # main을 항상 맨 앞으로 정렬 (규칙: main 필수 + 우선 처리)
        self.branches = sorted(branches, key=lambda b: (0 if b == "main" else 1, branches.index(b)))
        self.commit_msg = commit_msg
        self.sibling_repos = [Path(p) for p in (sibling_repos or [])]
        self.dry_run = dry_run
        self.log = log_callback or (lambda msg: None)
        self.results: list[BranchResult] = []
        self._session_id = str(uuid.uuid4())[:8]

    # ── 내부 헬퍼 ──────────────────────────────────────────────

    def _git(self, args: list[str], cwd: str | Path | None = None) -> tuple[int, str, str]:
        return run_git(args, cwd=str(cwd or self.repo_root))

    def _log(self, msg: str) -> None:
        self.log(msg)

    # ── 사전 검증 ─────────────────────────────────────────────

    def validate(self) -> list[str]:
        """실행 전 검증. 오류 목록 반환 (빈 리스트면 OK)."""
        errors = []

        if not self.repo_root.exists():
            errors.append(f"저장소 경로가 없습니다: {self.repo_root}")

        if not self.datasheet_exe.exists():
            errors.append(f"datasheet.exe 없음: {self.datasheet_exe}")

        if not self.xlsx_path.exists():
            errors.append(f"xlsx 파일 없음: {self.xlsx_path}")

        if not self.branches:
            errors.append("대상 브랜치를 하나 이상 선택해주세요.")

        if "main" not in self.branches:
            errors.append("main 브랜치는 필수 포함입니다.")

        if not self.commit_msg.strip():
            errors.append("커밋 메시지를 입력해주세요.")

        # sibling 저장소 경로 존재 확인
        for sr in self.sibling_repos:
            if not sr.exists():
                errors.append(f"sibling 저장소 경로가 없습니다: {sr}")

        if errors:
            return errors

        # 원격 브랜치 존재 확인 (cos-data)
        for br in self.branches:
            if not check_branch_exists(str(self.repo_root), br):
                errors.append(f"원격에 브랜치 없음: origin/{br} (cos-data)")

        # sibling 저장소 브랜치 존재 확인
        for sr in self.sibling_repos:
            for br in self.branches:
                if not check_branch_exists(str(sr), br):
                    errors.append(f"원격에 브랜치 없음: origin/{br} ({sr.name})")

        return errors

    # ── 메인 실행 ─────────────────────────────────────────────

    def run(self) -> list[BranchResult]:
        """동기 실행. threading.Thread에서 호출하면 비동기처럼 동작."""
        try:
            acquire_lock()
            self.results = self._run_inner()
        except LockError as e:
            self._log(f"[오류] {e}")
            self.results = [
                BranchResult(b, error_stage="잠금", error_detail=str(e))
                for b in self.branches
            ]
        finally:
            release_lock()

        return self.results

    def _run_inner(self) -> list[BranchResult]:
        results: list[BranchResult] = []

        src_branch = current_branch(str(self.repo_root))
        self._log(f"소스 브랜치: {src_branch}")

        # 현재 저장소에 미커밋 변경사항 확인
        code, status_out, _ = self._git(["status", "--porcelain"])
        if status_out.strip():
            self._log(f"[경고] 저장소에 미커밋 변경사항이 있습니다:\n{status_out.strip()}")

        for branch in self.branches:
            res = self._process_branch(branch)
            results.append(res)

        # 원래 브랜치 복원
        self._log(f"\n원래 브랜치 복원: {src_branch}")
        self._git(["checkout", src_branch])

        # 성공한 브랜치만 롤백 로그 저장
        save_rollback_log(self.commit_msg, results)

        return results

    def _process_branch(self, branch: str) -> BranchResult:
        res = BranchResult(branch=branch)
        self._log(f"\n[{branch}] 시작")

        try:
            # sibling 저장소 먼저 최신화 (datasheet 실행 전제조건)
            for sr in self.sibling_repos:
                self._log(f"[{branch}] {sr.name} checkout 중...")
                code, _, err = run_git(["checkout", branch], cwd=str(sr))
                if code != 0:
                    res.error_stage = f"{sr.name} checkout"
                    res.error_detail = err
                    self._log(f"[{branch}] 실패 - {sr.name} checkout\n{err}")
                    return res

                self._log(f"[{branch}] {sr.name} pull 중...")
                code, _, err = run_git(["pull", "origin", branch], cwd=str(sr))
                if code != 0:
                    res.error_stage = f"{sr.name} pull"
                    res.error_detail = err
                    self._log(f"[{branch}] 실패 - {sr.name} pull\n{err}")
                    return res

            # cos-data 브랜치 체크아웃
            self._log(f"[{branch}] checkout 중...")
            code, _, err = self._git(["checkout", branch])
            if code != 0:
                res.error_stage = "git checkout"
                res.error_detail = err
                self._log(f"[{branch}] 실패 - checkout\n{err}")
                return res

            # pull
            self._log(f"[{branch}] pull 중...")
            code, _, err = self._git(["pull", "origin", branch])
            if code != 0:
                res.error_stage = "git pull"
                res.error_detail = err
                self._log(f"[{branch}] 실패 - pull\n{err}")
                return res

            # xlsx 복사
            self._log(f"[{branch}] xlsx 복사 중...")
            dst_xlsx = self.repo_root / "excel" / "ui_strings.xlsx"
            shutil.copy2(str(self.xlsx_path), str(dst_xlsx))

            # datasheet 실행
            self._log(f"[{branch}] datasheet.exe 실행 중...")
            if not self.dry_run:
                ds_result = subprocess.run(
                    [str(self.datasheet_exe)],
                    cwd=str(self.repo_root),
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                )
                if ds_result.returncode != 0:
                    res.error_stage = "datasheet.exe"
                    res.error_detail = ds_result.stderr or ds_result.stdout
                    self._log(f"[{branch}] 실패 - datasheet\n{res.error_detail}")
                    return res
                self._log(f"[{branch}] datasheet 완료")
            else:
                self._log(f"[{branch}] [dry-run] datasheet 건너뜀")

            # 변경 파일 확인
            code, diff_out, _ = self._git(["diff", "--name-only"])
            changed_files = [f.strip() for f in diff_out.splitlines() if f.strip()]
            # 미추적 파일도 포함 (새로 생성된 pb 파일 등)
            code2, ls_out, _ = self._git(["ls-files", "--others", "--exclude-standard"])
            untracked = [f.strip() for f in ls_out.splitlines() if f.strip()]
            all_changed = changed_files + untracked
            self._log(f"[{branch}] 변경 파일 {len(all_changed)}개: {', '.join(all_changed[:5])}")

            if not all_changed and not self.dry_run:
                res.skipped = True
                res.skip_reason = "변경된 파일 없음"
                self._log(f"[{branch}] 스킵 - 변경 없음")
                return res

            if self.dry_run:
                res.success = True
                res.skipped = False
                self._log(f"[{branch}] [dry-run] 커밋/푸시 건너뜀 - 완료")
                return res

            # git add + commit + push
            self._log(f"[{branch}] git add...")
            self._git(["add", "-A"])

            self._log(f"[{branch}] git commit...")
            code, _, err = self._git(["commit", "-m", self.commit_msg])
            if code != 0:
                res.error_stage = "git commit"
                res.error_detail = err
                self._log(f"[{branch}] 실패 - commit\n{err}")
                return res

            # 커밋 해시 저장 (롤백용)
            _, hash_out, _ = self._git(["rev-parse", "HEAD"])
            res.commit_hash = hash_out.strip()

            self._log(f"[{branch}] git push...")
            code, _, err = self._git(["push", "origin", branch])
            if code != 0:
                res.error_stage = "git push"
                res.error_detail = err
                self._log(f"[{branch}] 실패 - push\n{err}")
                return res

            res.success = True
            self._log(f"[{branch}] 완료")

        except Exception as e:
            res.error_stage = "예외"
            res.error_detail = str(e)
            self._log(f"[{branch}] 예외 발생: {e}")

        return res


# ──────────────────────────────────────────────────────────────
# threading 래퍼 (Streamlit용)
# ──────────────────────────────────────────────────────────────

# ──────────────────────────────────────────────────────────────
# 롤백 로그
# ──────────────────────────────────────────────────────────────

def save_rollback_log(commit_msg: str, results: list[BranchResult]) -> None:
    import json
    entries = [
        {"branch": r.branch, "commit_hash": r.commit_hash}
        for r in results
        if r.success and r.commit_hash
    ]
    if not entries:
        return
    data = {
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "commit_msg": commit_msg,
        "entries": entries,
    }
    with open(ROLLBACK_LOG_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_rollback_log() -> dict | None:
    import json
    if not ROLLBACK_LOG_PATH.exists():
        return None
    try:
        with open(ROLLBACK_LOG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────
# 롤백 실행
# ──────────────────────────────────────────────────────────────

class Rollbacker:
    """마지막 전파 커밋을 각 브랜치에서 revert."""

    def __init__(
        self,
        repo_root: str,
        log: dict,
        log_callback: Callable[[str], None] | None = None,
    ):
        self.repo_root = Path(repo_root)
        self.log_data = log
        self.log = log_callback or (lambda msg: None)
        self.results: list[BranchResult] = []
        self._session_id = str(uuid.uuid4())[:8]

    def _git(self, args: list[str], cwd=None) -> tuple[int, str, str]:
        return run_git(args, cwd=str(cwd or self.repo_root))

    def run(self) -> list[BranchResult]:
        TMP_BASE.mkdir(parents=True, exist_ok=True)
        work_dir = TMP_BASE / self._session_id
        results = []

        try:
            acquire_lock()
            work_dir.mkdir(parents=True, exist_ok=True)
            for entry in self.log_data.get("entries", []):
                res = self._revert_branch(entry["branch"], entry["commit_hash"], work_dir)
                results.append(res)
        except LockError as e:
            self.log(f"[오류] {e}")
        finally:
            release_lock()
            if work_dir.exists():
                shutil.rmtree(work_dir, ignore_errors=True)

        self.results = results
        return results

    def _revert_branch(self, branch: str, commit_hash: str, work_dir: Path) -> BranchResult:
        res = BranchResult(branch=branch)
        wt_path = work_dir / branch.replace("/", "_")
        self.log(f"\n[{branch}] 롤백 시작 (revert {commit_hash[:8]})")

        try:
            code, _, err = self._git(["worktree", "add", str(wt_path), branch])
            if code != 0:
                res.error_stage = "worktree add"
                res.error_detail = err
                self.log(f"[{branch}] 실패 - worktree add\n{err}")
                return res

            code, _, err = self._git(["pull", "origin", branch], cwd=wt_path)
            if code != 0:
                res.error_stage = "git pull"
                res.error_detail = err
                self.log(f"[{branch}] 실패 - pull\n{err}")
                return res

            self.log(f"[{branch}] git revert...")
            code, _, err = self._git(
                ["revert", "--no-edit", commit_hash], cwd=wt_path
            )
            if code != 0:
                res.error_stage = "git revert"
                res.error_detail = err
                self.log(f"[{branch}] 실패 - revert\n{err}")
                return res

            self.log(f"[{branch}] git push...")
            code, _, err = self._git(["push", "origin", branch], cwd=wt_path)
            if code != 0:
                res.error_stage = "git push"
                res.error_detail = err
                self.log(f"[{branch}] 실패 - push\n{err}")
                return res

            res.success = True
            self.log(f"[{branch}] 롤백 완료")

        except Exception as e:
            res.error_stage = "예외"
            res.error_detail = str(e)
            self.log(f"[{branch}] 예외: {e}")

        finally:
            if wt_path.exists():
                self._git(["worktree", "remove", "--force", str(wt_path)])

        return res


class RollbackerThread:
    def __init__(self, rollbacker: Rollbacker):
        self._rollbacker = rollbacker
        self._thread: threading.Thread | None = None
        self._done = threading.Event()
        self.results: list[BranchResult] = []

    def start(self) -> None:
        self._done.clear()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def _run(self) -> None:
        self.results = self._rollbacker.run()
        self._done.set()

    def is_done(self) -> bool:
        return self._done.is_set()


class PropagatorThread:
    """Streamlit 비동기 실행을 위한 스레드 래퍼."""

    def __init__(self, propagator: Propagator):
        self._propagator = propagator
        self._thread: threading.Thread | None = None
        self._done = threading.Event()
        self.results: list[BranchResult] = []

    def start(self) -> None:
        self._done.clear()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def _run(self) -> None:
        self.results = self._propagator.run()
        self._done.set()

    def is_done(self) -> bool:
        return self._done.is_set()

    def join(self) -> None:
        if self._thread:
            self._thread.join()
