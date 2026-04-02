"""
gen-product-template.py
상품(product) 입력용 Excel 템플릿 생성

키워드/아이템/서브카테고리/뱃지·태그 정보를 동적으로 로드하여 XLOOKUP 참조 탭 구성.

실행:
  python make_outgame/product/gen-product-template.py

출력:
  make_outgame/product/product-input-template.xlsx
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime, timezone, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT   = SCRIPT_DIR.parent.parent   # cos-data 루트
OUT_PATH    = SCRIPT_DIR / "product-input-template.xlsx"

KEYWORDS_PATH = REPO_ROOT / "excel" / "keywords.xlsx"
ITEMS_PATH    = REPO_ROOT / "excel" / "items.xlsx"
PRODUCT_PATH  = REPO_ROOT / "excel" / "product.xlsx"

KST = timezone(timedelta(hours=9))

# ── 색상 ────────────────────────────────────────
CLR_REQUIRED   = "FFD7D7"   # 연분홍 — 필수
CLR_OPTIONAL   = "DDEEFF"   # 연파랑 — 선택
CLR_FIXED      = "E8E8E8"   # 회색   — 자동입력
CLR_HEADER     = "2F4F8F"   # 진파랑 — 헤더
CLR_HEADER_TXT = "FFFFFF"
CLR_EXAMPLE    = "F0FFF0"   # 연초록 — 예시행
CLR_SECTION    = "FFF3CD"   # 연노랑 — 섹션구분
CLR_XLOOKUP    = "FFFACD"   # 연황색 — XLOOKUP 자동표시
CLR_REF_HDR    = "4A6741"   # 어두운초록 — 참조탭 헤더
CLR_REF_HDR_TXT= "FFFFFF"

MAX_DATA_ROW = 203   # 4 + 최대 입력행 수

# ── 스타일 헬퍼 ──────────────────────────────────

def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def header_font(color=CLR_HEADER_TXT, size=10) -> Font:
    return Font(bold=True, color=color, size=size)

def thin_border() -> Border:
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, col: int, text: str, width: int,
                 required: bool = True, fixed: bool = False,
                 xlookup: bool = False) -> None:
    """row1=헤더, row2=필수/선택/자동 태그"""
    cell = ws.cell(row=1, column=col, value=text)
    cell.fill = fill(CLR_HEADER)
    cell.font = header_font()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(col)].width = width

    tag_cell = ws.cell(row=2, column=col)
    if xlookup:
        tag_cell.value = "자동"
        tag_cell.fill = fill(CLR_XLOOKUP)
    elif fixed:
        tag_cell.value = "자동"
        tag_cell.fill = fill(CLR_FIXED)
    elif required:
        tag_cell.value = "필수"
        tag_cell.fill = fill(CLR_REQUIRED)
    else:
        tag_cell.value = "선택"
        tag_cell.fill = fill(CLR_OPTIONAL)
    tag_cell.font = Font(size=8, color="555555")
    tag_cell.alignment = Alignment(horizontal="center")

def add_dropdown(ws, col: int, formula: str,
                 start_row: int = 4, end_row: int = MAX_DATA_ROW) -> None:
    col_letter = get_column_letter(col)
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    dv.sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.add_data_validation(dv)

def freeze_and_filter(ws, freeze_cell: str = "A4") -> None:
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions

def style_data_rows(ws, headers: list[tuple], start_row: int = 4) -> None:
    """headers: [(text, width, required, fixed, xlookup), ...]"""
    for row in range(start_row, MAX_DATA_ROW + 1):
        for c, hdr in enumerate(headers, start=1):
            fixed   = hdr[3] if len(hdr) > 3 else False
            xlookup = hdr[4] if len(hdr) > 4 else False
            req     = hdr[2] if len(hdr) > 2 else True
            cell = ws.cell(row=row, column=c)
            cell.border = thin_border()
            if xlookup:
                cell.fill = fill(CLR_XLOOKUP)
            elif fixed:
                cell.fill = fill(CLR_FIXED)
            elif req:
                cell.fill = fill(CLR_REQUIRED)
            else:
                cell.fill = fill(CLR_OPTIONAL)

def write_xlookup_kst(ws, row: int, src_col: int, dst_col: int) -> None:
    """src_col의 키워드 key → dst_col에 KST 시간 XLOOKUP"""
    src = get_column_letter(src_col)
    ws.cell(row=row, column=dst_col).value = (
        f"=IF({src}{row}=\"\",\"\","
        f"XLOOKUP({src}{row},'📖 키워드'!$A:$A,'📖 키워드'!$C:$C,\"미등록\"))"
    )

def write_xlookup_item_name(ws, row: int, src_col: int, dst_col: int) -> None:
    """src_col의 아이템 ID → dst_col에 아이템명 XLOOKUP"""
    src = get_column_letter(src_col)
    ws.cell(row=row, column=dst_col).value = (
        f"=IF({src}{row}=\"\",\"\","
        f"XLOOKUP({src}{row},'📖 아이템'!$A:$A,'📖 아이템'!$B:$B,\"ID 없음\"))"
    )

def populate_xlookup_formulas(ws, col_map: dict[str, int],
                               start_row: int = 4) -> None:
    """
    col_map 예시:
      {"kst": [(src_col, dst_col), ...], "item": [(src_col, dst_col), ...]}
    start_row ~ MAX_DATA_ROW 까지 수식 삽입
    """
    for row in range(start_row, MAX_DATA_ROW + 1):
        for src, dst in col_map.get("kst", []):
            write_xlookup_kst(ws, row, src, dst)
        for src, dst in col_map.get("item", []):
            write_xlookup_item_name(ws, row, src, dst)

def example_fill(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = fill(CLR_EXAMPLE)


# ══════════════════════════════════════════════════════
# 데이터 로더
# ══════════════════════════════════════════════════════

def load_keywords() -> list[tuple[str, int, str]]:
    """keywords.xlsx → [(id, unix_value, kst_str), ...]"""
    wb = openpyxl.load_workbook(KEYWORDS_PATH, data_only=True)
    ws = wb["timestamp"]
    result = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        kw_id = row[0]
        unix_val = row[1]
        dt_val = row[2]   # datetime or None
        if kw_id is None:
            continue
        kst_str = ""
        if isinstance(dt_val, datetime):
            # keywords.xlsx의 datetime은 UTC+0으로 저장, KST는 +9h
            kst_dt = dt_val.replace(tzinfo=timezone.utc).astimezone(KST)
            kst_str = kst_dt.strftime("%Y-%m-%d %H:%M KST")
        elif unix_val is not None:
            try:
                kst_dt = datetime.fromtimestamp(int(unix_val), tz=KST)
                kst_str = kst_dt.strftime("%Y-%m-%d %H:%M KST")
            except Exception:
                pass
        result.append((str(kw_id), int(unix_val) if unix_val else 0, kst_str))
    return result


def load_builds() -> list[str]:
    """keywords.xlsx build 시트 → ['$$NONE', '$$DEV', '$$LAUNCH_0', ...]"""
    wb = openpyxl.load_workbook(KEYWORDS_PATH, data_only=True)
    ws = wb["build"]
    result = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        bid = row[0]
        if bid is None:
            continue
        result.append(f"$${bid}")
    return result


def load_items() -> list[tuple[int, str, str]]:
    """items.xlsx → [(key, name, category_name), ...]"""
    wb = openpyxl.load_workbook(ITEMS_PATH, data_only=True)
    ws = wb["items"]
    result = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        key = row[0]
        name = row[5]         # col6: name
        cat_name = row[7]     # col8: category_name
        if key is None:
            continue
        result.append((int(key), str(name or ""), str(cat_name or "")))
    return result


def load_sub_categories() -> list[tuple[int, str, int]]:
    """product.xlsx → product_sub_categories [(id, category_name, ref_main_category_id), ...]"""
    wb = openpyxl.load_workbook(PRODUCT_PATH, data_only=True)
    ws = wb["product_sub_categories"]
    result = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        sid = row[0]
        if sid is None:
            continue
        ref_main = row[1]
        cat_name = row[3]
        result.append((int(sid), str(cat_name or ""), int(ref_main) if ref_main else 0))
    return result


def load_badge_tag_popup() -> tuple[list, list, list]:
    """badge_infos, tag_infos, popup_deco_info 로드"""
    wb = openpyxl.load_workbook(PRODUCT_PATH, data_only=True)

    badges = []
    ws = wb["badge_infos"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        badges.append((int(row[0]), str(row[1] or ""), str(row[2] or ""), str(row[3] or "")))

    tags = []
    ws = wb["tag_infos"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        tags.append((int(row[0]), str(row[1] or ""), str(row[2] or "")))

    popups = []
    ws = wb["popup_deco_info"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        popups.append((int(row[0]), str(row[1] or ""), str(row[2] or ""), str(row[3] or "")))

    return badges, tags, popups


# ══════════════════════════════════════════════════════
# 참조 탭 1: 📖 안내
# ══════════════════════════════════════════════════════

def build_guide_sheet(ws) -> None:
    ws.title = "📖 안내"
    ws.column_dimensions["A"].width = 90
    ws.sheet_view.showGridLines = False

    lines = [
        ("COS 상품(product) 입력 템플릿", True, 14),
        ("", False, 10),
        ("▶ 색상 의미", True, 11),
        ("  분홍 (필수) — 반드시 입력해야 하는 항목", False, 10),
        ("  파랑 (선택) — 없으면 기본값 자동 적용 또는 생략 가능", False, 10),
        ("  회색 (자동) — 입력하지 않아도 됨 (스크립트가 자동 처리)", False, 10),
        ("  노랑 (자동) — XLOOKUP 자동 표시 칸 (입력 불필요)", False, 10),
        ("  초록 (예시) — 3행은 예시. 실제 입력은 4행부터", False, 10),
        ("", False, 10),
        ("▶ 탭 설명", True, 11),
        ("  📖 키워드   — 판매 기간용 키워드 목록 (gen 실행 시 자동 최신화)", False, 10),
        ("  📖 아이템   — 아이템 key·이름 목록 (gen 실행 시 자동 최신화)", False, 10),
        ("  📖 서브카테고리 — product sub_category id·이름 목록", False, 10),
        ("  📖 뱃지/태그   — badge_infos, tag_infos, popup_deco_info 참조용", False, 10),
        ("  일반상품   — product_infos + general_store_products_shop", False, 10),
        ("  조건부상품 — product_infos + conditional_products", False, 10),
        ("  출석패키지 — product_infos + attendance_product + attendance_product_rewards", False, 10),
        ("  월정액     — product_infos + daily_reward_product", False, 10),
        ("  프리미엄패스 — product_infos + premium_pass_product", False, 10),
        ("  스텝상품   — product_infos + step_products", False, 10),
        ("", False, 10),
        ("▶ 사용 방법", True, 11),
        ("  1. gen 실행으로 최신 keywords/items 로드:", False, 10),
        ("     python make_outgame/product/gen-product-template.py", False, 10),
        ("  2. 해당 탭 4행부터 입력 (노랑 셀은 자동 계산됨)", False, 10),
        ("  3. 판매 기간은 📖 키워드 탭 참조하여 키워드 키 입력", False, 10),
        ("  4. 파일 저장 후 apply 실행:", False, 10),
        ("     python make_outgame/product/apply-product-input.py all product-input-template.xlsx", False, 10),
        ("     python make_outgame/product/apply-product-input.py 일반 product-input-template.xlsx", False, 10),
        ("     python make_outgame/product/apply-product-input.py 조건부 product-input-template.xlsx", False, 10),
        ("     python make_outgame/product/apply-product-input.py 출석 product-input-template.xlsx", False, 10),
        ("     python make_outgame/product/apply-product-input.py 월정액 product-input-template.xlsx", False, 10),
        ("     python make_outgame/product/apply-product-input.py 패스 product-input-template.xlsx", False, 10),
        ("     python make_outgame/product/apply-product-input.py 스텝 product-input-template.xlsx", False, 10),
        ("  5. datasheet.exe 실행하여 protobuf 변환", False, 10),
        ("", False, 10),
        ("▶ 주의사항", True, 11),
        ("  - 상품ID 비우면 product_infos 전체 최대 key+1 자동 생성", False, 10),
        ("  - 제작완료=TRUE 행은 재처리되지 않음", False, 10),
        ("  - 판매 기간 키워드: $$COMMON_STARTTAMP, $$H4_END 등 (📖 키워드 탭 참조)", False, 10),
        ("  - 출석패키지: 빈 일차(아이템ID 없음) 자동 스킵", False, 10),
    ]
    for r, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=size)
        if bold and text:
            cell.fill = fill(CLR_SECTION)


# ══════════════════════════════════════════════════════
# 참조 탭 2: 📖 키워드
# ══════════════════════════════════════════════════════

def build_keywords_ref_sheet(ws, keywords: list, builds: list) -> None:
    ws.title = "📖 키워드"
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["E"].width = 30

    # 헤더 — 타임스탬프 키워드 (A~C)
    for col, (text, w) in enumerate([
        ("키워드명", 45), ("raw값(Unix)", 16), ("KST 변환 시간", 22)
    ], start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = fill(CLR_REF_HDR)
        cell.font = Font(bold=True, color=CLR_REF_HDR_TXT, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    ws.cell(row=1, column=4, value="※ gen 실행 시 자동 최신화 / 입력탭의 판매기간 드롭다운 소스")
    ws.cell(row=1, column=4).font = Font(size=8, color="888888")

    for r, (kw_id, unix_val, kst_str) in enumerate(keywords, start=2):
        ws.cell(row=r, column=1, value=f"$${kw_id}").border = thin_border()
        ws.cell(row=r, column=2, value=unix_val).border = thin_border()
        ws.cell(row=r, column=3, value=kst_str).border = thin_border()

    # 헤더 — $filter 빌드 키워드 (E열)
    hdr_e = ws.cell(row=1, column=5, value="$filter (빌드)")
    hdr_e.fill = fill(CLR_REF_HDR)
    hdr_e.font = Font(bold=True, color=CLR_REF_HDR_TXT, size=10)
    hdr_e.alignment = Alignment(horizontal="center")
    hdr_e.border = thin_border()

    for r, bid in enumerate(builds, start=2):
        ws.cell(row=r, column=5, value=bid).border = thin_border()

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════
# 참조 탭 3: 📖 아이템
# ══════════════════════════════════════════════════════

def build_items_ref_sheet(ws, items: list) -> None:
    ws.title = "📖 아이템"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20

    for col, text in enumerate(["아이템 key", "이름", "카테고리"], start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = fill(CLR_REF_HDR)
        cell.font = Font(bold=True, color=CLR_REF_HDR_TXT, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    ws.cell(row=1, column=4, value="※ gen 실행 시 자동 최신화 / 아이템ID 입력 옆 칸에 이름 자동 표시")
    ws.cell(row=1, column=4).font = Font(size=8, color="888888")

    for r, (key, name, cat) in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=key).border = thin_border()
        ws.cell(row=r, column=2, value=name).border = thin_border()
        ws.cell(row=r, column=3, value=cat).border = thin_border()

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════
# 참조 탭 4: 📖 서브카테고리
# ══════════════════════════════════════════════════════

def build_subcategory_ref_sheet(ws, sub_cats: list) -> None:
    ws.title = "📖 서브카테고리"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14

    for col, text in enumerate(["서브카테고리 ID", "이름", "메인카테고리 ID"], start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = fill(CLR_REF_HDR)
        cell.font = Font(bold=True, color=CLR_REF_HDR_TXT, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()

    for r, (sid, cat_name, main_id) in enumerate(sub_cats, start=2):
        ws.cell(row=r, column=1, value=sid).border = thin_border()
        ws.cell(row=r, column=2, value=cat_name).border = thin_border()
        ws.cell(row=r, column=3, value=main_id).border = thin_border()

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════
# 참조 탭 5: 📖 뱃지/태그
# ══════════════════════════════════════════════════════

def build_badge_tag_ref_sheet(ws, badges: list, tags: list, popups: list) -> None:
    ws.title = "📖 뱃지태그"

    def ref_header(col, text, width):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = fill(CLR_REF_HDR)
        cell.font = Font(bold=True, color=CLR_REF_HDR_TXT, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = width

    # Badge (col A-D)
    for c, (text, w) in enumerate([("Badge ID", 10), ("resource_key", 28), ("text", 30), ("rate_text", 12)], start=1):
        ref_header(c, text, w)
    for r, (bid, rkey, text, rate) in enumerate(badges, start=2):
        ws.cell(row=r, column=1, value=bid).border = thin_border()
        ws.cell(row=r, column=2, value=rkey).border = thin_border()
        ws.cell(row=r, column=3, value=text).border = thin_border()
        ws.cell(row=r, column=4, value=rate).border = thin_border()

    # Tag (col F-H)
    for c, (text, w) in enumerate([("Tag ID", 10), ("product_tag_type", 22), ("text", 24)], start=6):
        ref_header(c, text, w)
    for r, (tid, ttype, text) in enumerate(tags, start=2):
        ws.cell(row=r, column=6, value=tid).border = thin_border()
        ws.cell(row=r, column=7, value=ttype).border = thin_border()
        ws.cell(row=r, column=8, value=text).border = thin_border()

    # PopupDeco (col J-M)
    for c, (text, w) in enumerate([("PopupDeco ID", 12), ("deco_type", 18), ("deco_img", 30), ("deco_text", 24)], start=10):
        ref_header(c, text, w)
    for r, (pid, dtype, img, dtext) in enumerate(popups, start=2):
        ws.cell(row=r, column=10, value=pid).border = thin_border()
        ws.cell(row=r, column=11, value=dtype).border = thin_border()
        ws.cell(row=r, column=12, value=img).border = thin_border()
        ws.cell(row=r, column=13, value=dtext).border = thin_border()

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ══════════════════════════════════════════════════════
# 공통 헬퍼: product_infos + product_base 헤더 블록
# ══════════════════════════════════════════════════════
# 열 배치 (공통 기본 블록 — 36 열):
# C1:  key
# C2:  $filter
# C3:  ref_sub_category_id
# C4:  slot_type
# C5:  slot_thumb_img
# C6:  order
# C7:  product_popup_resource_key
# C8:  popup_thumb_img
# C9:  badge_info_id
# C10: tag_info_id
# C11: slot_fx_info_id
# C12: popup_fx_info_id
# C13: show_review
# C14: show_goods
# C15: description_type
# C16: product_name
# C17: product_sub_name
# C18: description
# C19: product_popup_description
# C20: start_timestamp (keyword dropdown)
# C21: [start KST XLOOKUP — auto]
# C22: end_timestamp (keyword dropdown)
# C23: [end KST XLOOKUP — auto]
# C24: purchase_count_limit
# C25: available
# C26: is_show_when_sold_out
# C27: price_type
# C28: p1 (pid or item_id)
# C29: p2 (pid_sea or item_qty)
# C30: p3 (pid_vn)
# C31: p4 (pid_web)
# C32: p5 (store_type)
# C33: p6 (product_price)
# C34: refresh_type
# C35: period_second
# C36: offset_timestamp

BASE_HEADERS = [
    # (헤더명, 폭, required, fixed, xlookup)
    ("상품ID\n(비우면 자동)",       10, True,  False, False),  # C1
    ("$filter\n(선택)",             18, False, False, False),  # C2
    ("서브카테고리ID",               14, True,  False, False),  # C3
    ("슬롯크기\n(slot_type)",       22, True,  False, False),  # C4
    ("슬롯 썸네일",                  22, False, False, False),  # C5
    ("정렬순서\n(order)",            8, False, False, False),  # C6
    ("팝업 리소스키",                22, False, False, False),  # C7
    ("팝업 썸네일",                  22, False, False, False),  # C8
    ("뱃지ID",                       8, False, False, False),  # C9
    ("태그ID",                        8, False, False, False),  # C10
    ("슬롯 FX ID",                   10, False, False, False),  # C11
    ("팝업 FX ID",                   10, False, False, False),  # C12
    ("검수모드\n노출",                8, False, False, False),  # C13
    ("구성품\n노출",                  8, False, False, False),  # C14
    ("설명타입",                     20, False, False, False),  # C15
    ("상품명",                       26, True,  False, False),  # C16
    ("부제\n(sub_name)",             20, False, False, False),  # C17
    ("설명",                         30, False, False, False),  # C18
    ("팝업 설명",                    30, False, False, False),  # C19
    ("판매시작\n키워드",             26, True,  False, False),  # C20
    ("[판매시작 KST]",               22, False, False, True),   # C21 xlookup
    ("판매종료\n키워드",             26, True,  False, False),  # C22
    ("[판매종료 KST]",               22, False, False, True),   # C23 xlookup
    ("구매횟수제한\n(-1=무제한)",    12, False, False, False),  # C24
    ("활성화\n(1=활성)",              8, False, False, False),  # C25
    ("매진시노출\n(0=숨김)",          8, False, False, False),  # C26
    ("결제방식",                     14, True,  False, False),  # C27
    ("p1\n재화ID/PID",               28, True,  False, False),  # C28
    ("p2\n재화qty/SEA_PID",          28, False, False, False),  # C29
    ("p3\nVN_PID",                   28, False, False, False),  # C30
    ("p4\n웹샵_PID",                 28, False, False, False),  # C31
    ("p5\n스토어타입",               22, False, False, False),  # C32
    ("p6\n실결제가격",               12, False, False, False),  # C33
    ("리프레시 방식",                14, False, False, False),  # C34
    ("리프레시 주기\n(초)",          12, False, False, False),  # C35
    ("리프레시 기준\n(timestamp)",   12, False, False, False),  # C36
]

BASE_NO_REFRESH_HEADERS = BASE_HEADERS[:33]  # refresh 3컬럼 제외 (C34-36 없음)

# 아이템 슬롯 헤더 (최대 5개, XLOOKUP 포함)
def item_headers(n: int = 5) -> list:
    hdrs = []
    for i in range(n):
        req = (i == 0)
        hdrs += [
            (f"지급 아이템{i} ID",    10, req,  False, False),
            (f"[아이템{i} 이름]",     22, False, False, True),   # xlookup
            (f"지급 아이템{i} 수량",  10, req,  False, False),
        ]
    return hdrs


def write_base_headers(ws, offset: int = 0) -> None:
    for i, hdr in enumerate(BASE_HEADERS, start=1):
        write_header(ws, offset + i, hdr[0], hdr[1], hdr[2], hdr[3], hdr[4] if len(hdr) > 4 else False)


def write_base_no_refresh_headers(ws, offset: int = 0) -> None:
    for i, hdr in enumerate(BASE_NO_REFRESH_HEADERS, start=1):
        write_header(ws, offset + i, hdr[0], hdr[1], hdr[2], hdr[3], hdr[4] if len(hdr) > 4 else False)


def add_base_dropdowns(ws, offset: int = 0, kw_count: int = 0,
                       build_count: int = 0) -> None:
    """기본 블록 드롭다운 설정"""
    SLOT_LIST = (
        '"PRODUCT_SLOT_TYPE_SMALL,PRODUCT_SLOT_TYPE_MEDIUM,PRODUCT_SLOT_TYPE_LARGE,'
        'PRODUCT_SLOT_TYPE_LARGE_MONTHLY_ITEM,PRODUCT_SLOT_TYPE_LARGE_MONTHLY_SOULCHIP,'
        'PRODUCT_SLOT_TYPE_CREATOR_SPONSORSHIP"'
    )
    DESC_LIST = (
        '"PRODUCT_DESCRIPTION_TYPE_UNSPECIFIED,'
        'PRODUCT_DESCRIPTION_TYPE_IMPOSSIBLE,'
        'PRODUCT_DESCRIPTION_TYPE_IN7DAYS"'
    )
    PRICE_LIST  = '"price_item,payment,free"'
    STORE_LIST  = '"STORE_TYPE_ALL,STORE_TYPE_WEB"'
    REFRESH_LIST= '"empty,period_config,monthly_reset"'

    # $filter 드롭다운: 📖 키워드 탭 E2:E{build_count+1}
    if build_count:
        filter_ref = f"'📖 키워드'!$E$2:$E${build_count + 1}"
        add_dropdown(ws, offset + 2, filter_ref)   # C2: $filter

    # keyword 드롭다운: 📖 키워드 탭 A2:A{kw_count+1}
    kw_ref = f"'📖 키워드'!$A$2:$A${kw_count + 1}" if kw_count else PRICE_LIST
    add_dropdown(ws, offset + 4,  SLOT_LIST)
    add_dropdown(ws, offset + 15, DESC_LIST)
    add_dropdown(ws, offset + 20, kw_ref)   # start_ts keyword
    add_dropdown(ws, offset + 22, kw_ref)   # end_ts keyword
    add_dropdown(ws, offset + 27, PRICE_LIST)
    add_dropdown(ws, offset + 32, STORE_LIST)
    if len(BASE_HEADERS) > 33:
        add_dropdown(ws, offset + 34, REFRESH_LIST)


# ══════════════════════════════════════════════════════
# 입력 탭 1: 일반상품
# ══════════════════════════════════════════════════════
# 열 레이아웃:
# C1-C36: BASE_HEADERS (product_infos + product_base + refresh)
# C37: item0_id  C38: [item0_name]  C39: item0_qty
# C40: item1_id  C41: [item1_name]  C42: item1_qty
# C43: item2_id  C44: [item2_name]  C45: item2_qty
# C46: item3_id  C47: [item3_name]  C48: item3_qty
# C49: item4_id  C50: [item4_name]  C51: item4_qty
# C52: 제작완료

def build_general_sheet(ws, kw_count: int = 0, build_count: int = 0) -> None:
    ws.title = "일반상품"
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 16

    hdrs = BASE_HEADERS + item_headers(5) + [
        ("제작완료", 8, False, True, False),
    ]
    for i, hdr in enumerate(hdrs, start=1):
        write_header(ws, i, hdr[0], hdr[1], hdr[2], hdr[3], hdr[4] if len(hdr) > 4 else False)

    add_base_dropdowns(ws, kw_count=kw_count, build_count=build_count)

    # 예시행
    ex = [
        "", "", 38, "PRODUCT_SLOT_TYPE_LARGE", "Monthly_Example_Slot_L", 10,
        "UI_StoreGoodsPopup_M", "", 1700, "", 3, 201, 0, 0,
        "PRODUCT_DESCRIPTION_TYPE_IMPOSSIBLE",
        "예시 상품명", "", "예시 설명", "",
        "$$COMMON_STARTTAMP01", "", "$$H4_END", "", -1, 1, 0,
        "payment", "dev.cos.pkg.example.tier10",
        "com.vng.sea.cos.example.tier10", "com.vng.vn.cos.example.tier10", "",
        "STORE_TYPE_ALL", 9.99,
        "empty", "", "",
        10, "", 1000, 1, "", 100,
    ]
    example_fill(ws, 3, len(hdrs))
    for c, v in enumerate(ex, start=1):
        ws.cell(row=3, column=c, value=v).border = thin_border()

    style_data_rows(ws, hdrs, start_row=4)

    # XLOOKUP 수식 삽입
    item_xlookup_pairs = [(37 + i * 3, 38 + i * 3) for i in range(5)]
    populate_xlookup_formulas(ws, {
        "kst": [(20, 21), (22, 23)],
        "item": item_xlookup_pairs,
    })

    ws.cell(row=1, column=len(hdrs) + 2,
            value="※ 분홍=필수 | 파랑=선택 | 노랑=자동(XLOOKUP) | 초록=예시(3행)")
    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════
# 입력 탭 2: 조건부상품
# ══════════════════════════════════════════════════════
# C1-C36: BASE_HEADERS
# C37-C51: item_headers(5)
# C52: valid_time
# C53: trigger_cooltime
# C54: ref_unlock_condition_id
# C55: 제작완료

def build_conditional_sheet(ws, kw_count: int = 0, build_count: int = 0) -> None:
    ws.title = "조건부상품"
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 16

    extra = [
        ("트리거 유효시간\n(초, valid_time)",       14, True,  False, False),  # C52
        ("쿨타임\n(초, trigger_cooltime)",            12, False, False, False),  # C53
        ("조건 ID\n(ref_unlock_condition_id)",        16, True,  False, False),  # C54
        ("제작완료",                                    8, False, True,  False),  # C55
    ]
    hdrs = BASE_HEADERS + item_headers(5) + extra
    for i, hdr in enumerate(hdrs, start=1):
        write_header(ws, i, hdr[0], hdr[1], hdr[2], hdr[3], hdr[4] if len(hdr) > 4 else False)

    add_base_dropdowns(ws, kw_count=kw_count, build_count=build_count)

    ex = [
        "", "", 10, "PRODUCT_SLOT_TYPE_MEDIUM", "", 101,
        "UI_StoreGoodsPopup_S", "", "", "", "", "", 0, 0, "",
        "예시 조건부 상품", "파격 혜택!", "조건 달성 축하 패키지", "",
        "$$COMMON_STARTTAMP", "", "$$INDEFINITE_TIMESTAMP", "", 1, 1, 0,
        "payment", "dev.cos.pkg.conditional.example.tier7",
        "com.vng.sea.cos.conditional.example.tier7",
        "com.vng.vn.cos.conditional.example.tier7", "",
        "STORE_TYPE_ALL", 6.99,
        "empty", "", "",
        10, "", 500, 1, "", 50,
        "", "", "", "", "",
        2419200, 0, 300000005,
    ]
    example_fill(ws, 3, len(hdrs))
    for c, v in enumerate(ex, start=1):
        ws.cell(row=3, column=c, value=v).border = thin_border()

    style_data_rows(ws, hdrs, start_row=4)

    item_xlookup_pairs = [(37 + i * 3, 38 + i * 3) for i in range(5)]
    populate_xlookup_formulas(ws, {
        "kst": [(20, 21), (22, 23)],
        "item": item_xlookup_pairs,
    })

    ws.cell(row=1, column=len(hdrs) + 2,
            value="※ valid_time/trigger_cooltime 단위: 초 | 2419200=28일")
    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════
# 입력 탭 3: 출석패키지
# ══════════════════════════════════════════════════════
# C1-C33: BASE_NO_REFRESH_HEADERS (refresh 없음)
# C34: reward_until_days
# C35: mail_template_id
# C36-(C36+30*6-1): 1일~30일 × (id0,qty0,id1,qty1,id2,qty2) = 180칸
# C216: 제작완료
# ※ attendance_product 시트에는 items 컬럼 없음 (즉시지급 없음)

ATTENDANCE_DAYS = 30
ATTENDANCE_ITEMS_PER_DAY = 3

def build_attendance_sheet(ws, kw_count: int = 0, build_count: int = 0) -> None:
    ws.title = "출석패키지"
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 16

    # 일차별 보상 (30일 × 3아이템 × (id+qty))
    # attendance_product 시트에는 items 컬럼 없음 → 즉시지급 없음
    daily_items = []
    for d in range(1, ATTENDANCE_DAYS + 1):
        for i in range(ATTENDANCE_ITEMS_PER_DAY):
            req = (i == 0 and d == 1)
            daily_items += [
                (f"{d}일_아이템{i}_ID",   8, req,  False, False),
                (f"{d}일_아이템{i}_qty",  8, False, False, False),
            ]

    attendance_extra = [
        ("보상수령\n기간(일)\n(reward_until_days)", 12, True, False, False),   # C34
        ("우편 템플릿 ID\n(mail_template_id)",      14, False, False, False),  # C35
    ]

    hdrs = BASE_NO_REFRESH_HEADERS + attendance_extra + daily_items + [
        ("제작완료", 8, False, True, False),
    ]
    for i, hdr in enumerate(hdrs, start=1):
        write_header(ws, i, hdr[0], hdr[1], hdr[2], hdr[3], hdr[4] if len(hdr) > 4 else False)

    # 드롭다운 (refresh 없음)
    SLOT_LIST = (
        '"PRODUCT_SLOT_TYPE_SMALL,PRODUCT_SLOT_TYPE_MEDIUM,PRODUCT_SLOT_TYPE_LARGE,'
        'PRODUCT_SLOT_TYPE_LARGE_MONTHLY_ITEM,PRODUCT_SLOT_TYPE_LARGE_MONTHLY_SOULCHIP"'
    )
    PRICE_LIST  = '"price_item,payment,free"'
    STORE_LIST  = '"STORE_TYPE_ALL,STORE_TYPE_WEB"'
    DESC_LIST   = '"PRODUCT_DESCRIPTION_TYPE_UNSPECIFIED,PRODUCT_DESCRIPTION_TYPE_IMPOSSIBLE,PRODUCT_DESCRIPTION_TYPE_IN7DAYS"'
    kw_ref = f"'📖 키워드'!$A$2:$A${kw_count + 1}" if kw_count else '""'
    if build_count:
        add_dropdown(ws, 2, f"'📖 키워드'!$E$2:$E${build_count + 1}")
    add_dropdown(ws, 4,  SLOT_LIST)
    add_dropdown(ws, 15, DESC_LIST)
    add_dropdown(ws, 20, kw_ref)
    add_dropdown(ws, 22, kw_ref)
    add_dropdown(ws, 27, PRICE_LIST)
    add_dropdown(ws, 32, STORE_LIST)

    # 예시행
    ex = [
        "", "", 37, "PRODUCT_SLOT_TYPE_LARGE", "Attendance_Example_Slot_L", 110,
        "UI_StoreGoodsPopup_Attendance", "", 1700, "", 3, 201, 0, 0,
        "PRODUCT_DESCRIPTION_TYPE_IMPOSSIBLE",
        "예시 출석 패키지", "", "매일 보상!", "",
        "$$COMMON_STARTTAMP01", "", "$$H4_END", "", 1, 1, 0,
        "payment", "dev.cos.pkg.attendance.example.tier3",
        "com.vng.sea.cos.attendance.example.tier3",
        "com.vng.vn.cos.attendance.example.tier3", "",
        "STORE_TYPE_ALL", 2.99,
        # attendance extra
        7, 8,
        # 1일 보상
        13001, 60, 1, 4000, "", "",
    ]
    example_fill(ws, 3, len(hdrs))
    for c, v in enumerate(ex, start=1):
        ws.cell(row=3, column=c, value=v).border = thin_border()

    style_data_rows(ws, hdrs, start_row=4)

    populate_xlookup_formulas(ws, {
        "kst": [(20, 21), (22, 23)],
    })

    daily_start_col = len(BASE_NO_REFRESH_HEADERS) + len(attendance_extra) + 1
    ws.cell(row=1, column=len(hdrs) + 2,
            value=(f"※ C{daily_start_col}~: 1일~{ATTENDANCE_DAYS}일 보상 (아이템{ATTENDANCE_ITEMS_PER_DAY}개/일) | attendance_product 시트에는 즉시지급 items 없음"))
    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════
# 입력 탭 4: 월정액
# ══════════════════════════════════════════════════════
# C1-C33: BASE_NO_REFRESH_HEADERS
# C34: period_days
# C35: repurchase_available_days
# C36: mail_template_id
# C37: mail_bonus_template_id
# C38-C46: 즉시지급 item_headers(3)
# C47: daily_reward_type (daily_item / ref_soul_chip_info_id)
# C48: daily_item_id (if daily_item)
# C49: daily_item_qty (if daily_item) / soul_chip_info_id (if ref)
# C50: bonus_item_day
# C51: bonus_item_id
# C52: bonus_item_qty
# C53: 제작완료

def build_monthly_sheet(ws, kw_count: int = 0, build_count: int = 0) -> None:
    ws.title = "월정액"
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 16

    instant_items = []
    for i in range(3):
        req = (i == 0)
        instant_items += [
            (f"즉시지급 아이템{i} ID",    10, req,  False, False),
            (f"[아이템{i} 이름]",          22, False, False, True),
            (f"즉시지급 아이템{i} 수량",   10, req,  False, False),
        ]

    monthly_extra = [
        ("지급 기간\n(period_days)",          10, True,  False, False),  # C34
        ("재구매 가능\n잔여일\n(repurchase)", 12, True,  False, False),  # C35
        ("우편 템플릿 ID\n(mail)",            12, False, False, False),  # C36
        ("보너스 우편\n템플릿 ID",             12, False, False, False),  # C37
    ]
    daily_reward_hdrs = [
        ("일일보상 타입\n(daily_item /\nref_soul_chip_info_id)", 24, True,  False, False),  # C47
        ("일일보상 아이템ID\n/ 소울칩정보ID",                    14, True,  False, False),  # C48
        ("일일보상 수량",                                          10, True,  False, False),  # C49
        ("보너스 보상 일차\n(bonus_item/day)",                    14, False, False, False),  # C50
        ("보너스 아이템ID",                                        12, False, False, False),  # C51
        ("보너스 수량",                                            10, False, False, False),  # C52
    ]

    hdrs = BASE_NO_REFRESH_HEADERS + monthly_extra + instant_items + daily_reward_hdrs + [
        ("제작완료", 8, False, True, False),  # C53
    ]
    for i, hdr in enumerate(hdrs, start=1):
        write_header(ws, i, hdr[0], hdr[1], hdr[2], hdr[3], hdr[4] if len(hdr) > 4 else False)

    SLOT_LIST = '"PRODUCT_SLOT_TYPE_LARGE,PRODUCT_SLOT_TYPE_LARGE_MONTHLY_ITEM,PRODUCT_SLOT_TYPE_LARGE_MONTHLY_SOULCHIP"'
    PRICE_LIST = '"price_item,payment,free"'
    STORE_LIST = '"STORE_TYPE_ALL,STORE_TYPE_WEB"'
    DESC_LIST  = '"PRODUCT_DESCRIPTION_TYPE_UNSPECIFIED,PRODUCT_DESCRIPTION_TYPE_IMPOSSIBLE"'
    DAILY_TYPE = '"daily_item,ref_soul_chip_info_id"'
    kw_ref = f"'📖 키워드'!$A$2:$A${kw_count + 1}" if kw_count else '""'

    if build_count:
        add_dropdown(ws, 2, f"'📖 키워드'!$E$2:$E${build_count + 1}")
    add_dropdown(ws, 4, SLOT_LIST)
    add_dropdown(ws, 15, DESC_LIST)
    add_dropdown(ws, 20, kw_ref)
    add_dropdown(ws, 22, kw_ref)
    add_dropdown(ws, 27, PRICE_LIST)
    add_dropdown(ws, 32, STORE_LIST)

    # 일일보상 타입 드롭다운: C47 = 33 + 4 + 9 + 1 = 47
    monthly_daily_type_col = len(BASE_NO_REFRESH_HEADERS) + len(monthly_extra) + len(instant_items) + 1
    add_dropdown(ws, monthly_daily_type_col, DAILY_TYPE)

    # 예시행
    ex = [
        "", "", 38, "PRODUCT_SLOT_TYPE_LARGE_MONTHLY_ITEM", "Monthly_Example_Slot_L", 29,
        "UI_StoreGoodsPopup_MonthlyItem", "Monthly_Example_Popup_L", 1700, "", 3, 201, 0, 0,
        "PRODUCT_DESCRIPTION_TYPE_IMPOSSIBLE",
        "월정액 예시 패키지", "", "30일간 일일 보상!", "",
        "$$COMMON_STARTTAMP01", "", "$$INDEFINITE_TIMESTAMP", "", -1, 1, 0,
        "price_item", 1, 1000, "", "", "", "",
        # monthly extra
        30, 60, 9, 10,
        # 즉시지급
        30, "", 1500, "", "", "",  "", "", "",
        # 일일보상
        "daily_item", 30, 100,
        14, 30, 1000,
    ]
    example_fill(ws, 3, len(hdrs))
    for c, v in enumerate(ex, start=1):
        ws.cell(row=3, column=c, value=v).border = thin_border()

    style_data_rows(ws, hdrs, start_row=4)

    instant_base = len(BASE_NO_REFRESH_HEADERS) + len(monthly_extra) + 1
    instant_item_xlookup = [(instant_base + i * 3, instant_base + i * 3 + 1) for i in range(3)]
    populate_xlookup_formulas(ws, {
        "kst": [(20, 21), (22, 23)],
        "item": instant_item_xlookup,
    })

    ws.cell(row=1, column=len(hdrs) + 2,
            value="※ daily_reward 타입: daily_item=아이템 지급 | ref_soul_chip_info_id=소울칩 정보 ID 참조")
    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════
# 입력 탭 5: 프리미엄패스
# ══════════════════════════════════════════════════════
# C1-C33: BASE_NO_REFRESH_HEADERS
# C34: product_id (premium_pass_product.product_id — 보통 product_info key와 동일)
# C35-C43: item_headers(3) (즉시지급)
# C44: 제작완료

def build_premium_pass_sheet(ws, kw_count: int = 0, build_count: int = 0) -> None:
    ws.title = "프리미엄패스"
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 16

    pass_extra = [
        ("product_id\n(보통 key와 동일)", 18, False, False, False),  # C34
    ]
    instant_items = []
    for i in range(3):
        req = (i == 0)
        instant_items += [
            (f"즉시지급 아이템{i} ID",    10, req,  False, False),
            (f"[아이템{i} 이름]",          22, False, False, True),
            (f"즉시지급 아이템{i} 수량",   10, req,  False, False),
        ]

    hdrs = BASE_NO_REFRESH_HEADERS + pass_extra + instant_items + [
        ("제작완료", 8, False, True, False),
    ]
    for i, hdr in enumerate(hdrs, start=1):
        write_header(ws, i, hdr[0], hdr[1], hdr[2], hdr[3], hdr[4] if len(hdr) > 4 else False)

    SLOT_LIST = '"PRODUCT_SLOT_TYPE_LARGE,PRODUCT_SLOT_TYPE_MEDIUM,PRODUCT_SLOT_TYPE_SMALL"'
    PRICE_LIST = '"payment,price_item,free"'
    STORE_LIST = '"STORE_TYPE_ALL,STORE_TYPE_WEB"'
    DESC_LIST  = '"PRODUCT_DESCRIPTION_TYPE_IN7DAYS,PRODUCT_DESCRIPTION_TYPE_UNSPECIFIED"'
    kw_ref = f"'📖 키워드'!$A$2:$A${kw_count + 1}" if kw_count else '""'

    if build_count:
        add_dropdown(ws, 2, f"'📖 키워드'!$E$2:$E${build_count + 1}")
    add_dropdown(ws, 4, SLOT_LIST)
    add_dropdown(ws, 15, DESC_LIST)
    add_dropdown(ws, 20, kw_ref)
    add_dropdown(ws, 22, kw_ref)
    add_dropdown(ws, 27, PRICE_LIST)
    add_dropdown(ws, 32, STORE_LIST)

    # 예시행
    ex = [
        30100002, "", 1, "PRODUCT_SLOT_TYPE_LARGE", "Seasonpass_002_Slot_L", 1,
        "", "", 1, "", 3, "", 0, 0,
        "PRODUCT_DESCRIPTION_TYPE_IN7DAYS",
        "배틀 패스 시즌2", "", "프리미엄 배틀 패스 시즌2", "",
        "$$SEASONS_START_TIMESTAMP_002", "", "$$SEASONS_END_TIMESTAMP_002", "", 1, 1, 0,
        "payment", "dev.cos.pkg.seasonpass2.tier10",
        "com.vng.sea.cos.seasonpass2.tier10",
        "com.vng.vn.cos.seasonpass2.tier10",
        "com.vng.web.cos.seasonpass2.tier10",
        "STORE_TYPE_ALL", 9.99,
        # pass_extra: product_id
        30100002,
        # 즉시지급
        1002, "", 3000, "", "", "",  "", "", "",
    ]
    example_fill(ws, 3, len(hdrs))
    for c, v in enumerate(ex, start=1):
        ws.cell(row=3, column=c, value=v).border = thin_border()

    style_data_rows(ws, hdrs, start_row=4)

    instant_base = len(BASE_NO_REFRESH_HEADERS) + len(pass_extra) + 1
    instant_item_xlookup = [(instant_base + i * 3, instant_base + i * 3 + 1) for i in range(3)]
    populate_xlookup_formulas(ws, {
        "kst": [(20, 21), (22, 23)],
        "item": instant_item_xlookup,
    })

    ws.cell(row=1, column=len(hdrs) + 2,
            value="※ premium_pass_product는 단일 행. product_id = 해당 product_infos key와 동일하게 입력")
    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════
# 입력 탭 6: 스텝상품
# ══════════════════════════════════════════════════════
# C1-C36: BASE_HEADERS (refresh 포함)
# C37+(step-1)*14: step1~3 × (price_id, price_qty, item0_id, [item0_name], item0_qty, ..., item3_id, [item3_name], item3_qty)
# 3스텝 × 14 = 42
# C79: 제작완료

STEP_COUNT = 3
STEP_ITEMS = 4  # 스텝당 아이템 최대 4개

def build_step_sheet(ws, kw_count: int = 0, build_count: int = 0) -> None:
    ws.title = "스텝상품"
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 16

    step_hdrs = []
    for s in range(1, STEP_COUNT + 1):
        step_hdrs += [
            (f"스텝{s} 결제 아이템ID", 10, True,  False, False),
            (f"스텝{s} 결제 수량",     10, True,  False, False),
        ]
        for i in range(STEP_ITEMS):
            required_item = (i == 0)
            step_hdrs += [
                (f"스텝{s} 보상 아이템{i} ID", 10, required_item, False, False),
                (f"[스텝{s} 아이템{i} 이름]",  20, False,         False, True),
                (f"스텝{s} 보상 수량{i}",       10, required_item, False, False),
            ]

    hdrs = BASE_HEADERS + step_hdrs + [
        ("제작완료", 8, False, True, False),
    ]
    for i, hdr in enumerate(hdrs, start=1):
        write_header(ws, i, hdr[0], hdr[1], hdr[2], hdr[3], hdr[4] if len(hdr) > 4 else False)

    add_base_dropdowns(ws, kw_count=kw_count, build_count=build_count)

    # 예시행
    ex = [
        "", "", 34, "PRODUCT_SLOT_TYPE_MEDIUM", "Step_Example_Slot", 1000,
        "UI_StoreGoodsPopup_S", "", "", "", "", "", 0, 0, "",
        "예시 스텝 상품", "", "", "",
        "$$COMMON_STARTTAMP", "", "$$INDEFINITE_TIMESTAMP", "", -1, 1, 0,
        "price_item", 10, 100, "", "", "", "",
        "empty", "", "",
        # step1
        10, 300, 1, "", 10000, "", "", "",
        # step2
        10, 500, 1, "", 20000, "", "", "",
    ]
    example_fill(ws, 3, len(hdrs))
    for c, v in enumerate(ex, start=1):
        ws.cell(row=3, column=c, value=v).border = thin_border()

    style_data_rows(ws, hdrs, start_row=4)

    # XLOOKUP: start/end + 스텝 아이템명
    # per-step cols = 2(price) + STEP_ITEMS*3(id+name+qty)
    step_cols = 2 + STEP_ITEMS * 3
    step_item_xlookup = []
    step_base = len(BASE_HEADERS) + 1  # C37
    for s in range(STEP_COUNT):
        base = step_base + s * step_cols
        for i in range(STEP_ITEMS):
            # item i: id at base+2+i*3, name at base+3+i*3
            item_id_off = 2 + i * 3
            step_item_xlookup.append((base + item_id_off, base + item_id_off + 1))

    populate_xlookup_formulas(ws, {
        "kst": [(20, 21), (22, 23)],
        "item": step_item_xlookup,
    })

    ws.cell(row=1, column=len(hdrs) + 2,
            value="※ 스텝상품: purchase_count_limit = 스텝 수와 동일하게 설정")
    freeze_and_filter(ws, "A4")


# ══════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════

def main() -> None:
    print(f"keywords 로드: {KEYWORDS_PATH}")
    keywords = load_keywords()
    print(f"  → {len(keywords)}개 키워드")

    print(f"items 로드: {ITEMS_PATH}")
    items = load_items()
    print(f"  → {len(items)}개 아이템")

    print(f"sub_categories 로드: {PRODUCT_PATH}")
    sub_cats = load_sub_categories()
    print(f"  → {len(sub_cats)}개 서브카테고리")

    print("badge/tag/popup 로드...")
    badges, tags, popups = load_badge_tag_popup()
    print(f"  → badge {len(badges)}, tag {len(tags)}, popup {len(popups)}")

    print("builds 로드...")
    builds = load_builds()
    print(f"  → {len(builds)}개 빌드 필터")

    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    print("템플릿 탭 생성 중...")

    # 참조 탭
    build_guide_sheet(wb.create_sheet())
    build_keywords_ref_sheet(wb.create_sheet(), keywords, builds)
    build_items_ref_sheet(wb.create_sheet(), items)
    build_subcategory_ref_sheet(wb.create_sheet(), sub_cats)
    build_badge_tag_ref_sheet(wb.create_sheet(), badges, tags, popups)

    kw_count = len(keywords)
    build_count = len(builds)

    # 입력 탭
    build_general_sheet(wb.create_sheet(), kw_count=kw_count, build_count=build_count)
    build_conditional_sheet(wb.create_sheet(), kw_count=kw_count, build_count=build_count)
    build_attendance_sheet(wb.create_sheet(), kw_count=kw_count, build_count=build_count)
    build_monthly_sheet(wb.create_sheet(), kw_count=kw_count, build_count=build_count)
    build_premium_pass_sheet(wb.create_sheet(), kw_count=kw_count, build_count=build_count)
    build_step_sheet(wb.create_sheet(), kw_count=kw_count, build_count=build_count)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"\n[OK] 저장 완료: {OUT_PATH}")
    print(f"   참조 탭 5개 + 입력 탭 6개 = 총 11탭")


if __name__ == "__main__":
    main()
