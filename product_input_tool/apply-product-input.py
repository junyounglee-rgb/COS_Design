"""
apply-product-input.py
product-input-template.xlsx → product.xlsx 삽입

사용법:
  python make_outgame/product/apply-product-input.py all     [template.xlsx]
  python make_outgame/product/apply-product-input.py 일반    [template.xlsx]
  python make_outgame/product/apply-product-input.py 조건부  [template.xlsx]
  python make_outgame/product/apply-product-input.py 출석    [template.xlsx]
  python make_outgame/product/apply-product-input.py 월정액  [template.xlsx]
  python make_outgame/product/apply-product-input.py 패스    [template.xlsx]
  python make_outgame/product/apply-product-input.py 스텝    [template.xlsx]

key 자동생성:
  상품ID 칸이 비어있으면 product_infos 시트 전체 최대 key+1 자동 할당
제작완료:
  적용 후 template 해당 행 "제작완료" 열에 TRUE 기입
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT   = SCRIPT_DIR.parent.parent
EXCEL_DIR   = REPO_ROOT / "excel"
PRODUCT_XLS = EXCEL_DIR / "product.xlsx"
DEFAULT_TMPL = SCRIPT_DIR / "product-input-template.xlsx"

# ── 기본값 ────────────────────────────────────────
DEFAULT_AVAILABLE       = 1
DEFAULT_SHOW_GOODS      = 0
DEFAULT_PURCHASE_LIMIT  = -1
DEFAULT_SHOW_SOLD_OUT   = 0
DEFAULT_ITEM_FOCUS      = "[]{}"
DEFAULT_SHOW_REVIEW     = 0


# ── 유틸 ─────────────────────────────────────────

def cv(cell_or_val) -> str | None:
    """셀 또는 값 → str. 비어있으면 None."""
    v = cell_or_val.value if hasattr(cell_or_val, "value") else cell_or_val
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def ci(cell_or_val, default=None) -> int | None:
    """셀 값 → int. 없으면 default."""
    v = cv(cell_or_val)
    if v is None:
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


def cb(cell_or_val, default=False) -> bool:
    """셀 값 → bool."""
    v = cv(cell_or_val)
    if v is None:
        return default
    return str(v).strip().upper() in ("TRUE", "1", "YES")


def is_done(tmpl_row, done_col: int) -> bool:
    val = tmpl_row[done_col - 1].value
    if val is None:
        return False
    return str(val).strip().upper() in ("TRUE", "1", "완료")


def mark_done(tmpl_row, done_col: int) -> None:
    tmpl_row[done_col - 1].value = True


def set_cell(ws, row: int, col: int, value) -> None:
    if value is not None:
        ws.cell(row=row, column=col, value=value)


def find_last_data_row(ws, min_row: int = 4) -> int:
    last = min_row - 1
    for row in ws.iter_rows(min_row=min_row):
        if any(c.value is not None for c in row):
            last = row[0].row
    return last


def max_key_in_sheet(ws, key_col: int = 1, min_row: int = 4) -> int:
    max_k = 0
    for row in ws.iter_rows(min_row=min_row, max_col=key_col, values_only=True):
        k = row[key_col - 1]
        if k and isinstance(k, (int, float)):
            max_k = max(max_k, int(k))
    return max_k


# ── template 컬럼 상수 (1-indexed, BASE_HEADERS 기준) ──────────────
KEY_COL          = 1
FILTER_COL       = 2
SUB_CAT_COL      = 3   # ref_sub_category_id
SLOT_TYPE_COL    = 4
SLOT_THUMB_COL   = 5
ORDER_COL        = 6
POPUP_RES_COL    = 7   # product_popup_resource_key
POPUP_THUMB_COL  = 8
BADGE_COL        = 9   # badge_info_id
TAG_COL          = 10  # tag_info_id
SLOT_FX_COL      = 11  # slot_fx_info_id
POPUP_FX_COL     = 12  # popup_fx_info_id
SHOW_REVIEW_COL  = 13  # show_review
SHOW_GOODS_COL   = 14  # show_goods
DESC_TYPE_COL    = 15  # description_type
PRODUCT_NAME_COL = 16
SUB_NAME_COL     = 17
DESC_COL         = 18
POPUP_DESC_COL   = 19  # product_popup_description (→ product_infos)
START_TS_COL     = 20
# 21 = [판매시작 KST] xlookup — skip
END_TS_COL       = 22
# 23 = [판매종료 KST] xlookup — skip
PURCHASE_LIMIT_COL   = 24
AVAILABLE_COL        = 25
SHOW_SOLD_OUT_COL    = 26
PRICE_KEY_COL        = 27  # %key price_type
PRICE_P1_COL         = 28
PRICE_P2_COL         = 29
PRICE_P3_COL         = 30
PRICE_P4_COL         = 31
PRICE_P5_COL         = 32
PRICE_P6_COL         = 33
REFRESH_KEY_COL      = 34  # BASE_HEADERS 전용
REFRESH_P1_COL       = 35
REFRESH_P2_COL       = 36

# item_headers(5) 시작 = 37
# 아이템 i 기준: id = 37+i*3, [name xlookup] = 38+i*3, qty = 39+i*3
ITEM_START_COL   = 37
ITEM_SLOT_COLS   = 3  # id, [name], qty per item


def tmpl_item_cols(i: int) -> tuple[int, int]:
    """i번째 아이템의 (id_col, qty_col) 반환 (xlookup name 스킵)."""
    base = ITEM_START_COL + i * ITEM_SLOT_COLS
    return (base, base + 2)  # id, qty


# ── product_infos 컬럼 상수 ───────────────────────
PINFO_KEY_COL          = 1
PINFO_FILTER_COL       = 2
# col3,4,5 = comment cols
PINFO_SUB_CAT_COL      = 6
PINFO_PRODUCT_TYPE_COL = 7
PINFO_SLOT_TYPE_COL    = 8
PINFO_SLOT_THUMB_COL   = 9
PINFO_SHOW_GOODS_COL   = 10
PINFO_SLOT_FX_COL      = 11
PINFO_ORDER_COL        = 12
PINFO_POPUP_RES_COL    = 13
PINFO_POPUP_THUMB_COL  = 14
PINFO_POPUP_FX_COL     = 15
PINFO_POPUP_DECO_COL   = 16  # ref_popup_deco_id (template에 없음, 스킵)
PINFO_BADGE_COL        = 17
PINFO_TAG_COL          = 18
PINFO_POPUP_DESC_COL   = 19
PINFO_DESC_TYPE_COL    = 20
PINFO_SHOW_REVIEW_COL  = 21

# ── general_store_products_shop 컬럼 상수 ─────────
GEN_KEY_COL              = 1
GEN_FILTER_COL           = 2
# col3 = comment
GEN_PRODUCT_NAME_COL     = 4
GEN_SUB_NAME_COL         = 5
GEN_DESC_COL             = 6
GEN_PRICE_KEY_COL        = 7
GEN_PRICE_P1_COL         = 8
GEN_PRICE_P2_COL         = 9
GEN_PRICE_P3_COL         = 10
GEN_PRICE_P4_COL         = 11
GEN_PRICE_P5_COL         = 12
GEN_PRICE_P6_COL         = 13
GEN_START_TS_COL         = 14
GEN_END_TS_COL           = 15
GEN_PURCHASE_LIMIT_COL   = 16
GEN_AVAILABLE_COL        = 17
GEN_SHOW_SOLD_OUT_COL    = 18
GEN_REFRESH_KEY_COL      = 19
GEN_REFRESH_P1_COL       = 20
GEN_REFRESH_P2_COL       = 21
GEN_ITEM_FOCUS_COL       = 22
GEN_ITEM0_ID_COL         = 23
GEN_ITEM0_QTY_COL        = 24
GEN_ITEM1_ID_COL         = 25
GEN_ITEM1_QTY_COL        = 26
GEN_ITEM2_ID_COL         = 27
GEN_ITEM2_QTY_COL        = 28
GEN_ITEM3_ID_COL         = 29
GEN_ITEM3_QTY_COL        = 30

# ── conditional_products 컬럼 상수 ────────────────
COND_KEY_COL              = 1
COND_FILTER_COL           = 2
# col3 = comment
COND_PRODUCT_NAME_COL     = 4
COND_SUB_NAME_COL         = 5
COND_DESC_COL             = 6
COND_UNLOCK_COND_COL      = 7   # ref_unlock_condition_id
COND_PRICE_KEY_COL        = 8
COND_PRICE_P1_COL         = 9
COND_PRICE_P2_COL         = 10
COND_PRICE_P3_COL         = 11
COND_PRICE_P4_COL         = 12
COND_PRICE_P5_COL         = 13  # conditional은 param5까지 (param6 없음)
COND_VALID_TIME_COL       = 14
COND_TRIGGER_COOL_COL     = 15
COND_SUB_DESC_COL         = 16  # sub_description (conditional 전용)
COND_START_TS_COL         = 17
COND_END_TS_COL           = 18
COND_PURCHASE_LIMIT_COL   = 19
COND_AVAILABLE_COL        = 20
COND_SHOW_SOLD_OUT_COL    = 21
COND_REFRESH_KEY_COL      = 22
COND_REFRESH_P1_COL       = 23
COND_REFRESH_P2_COL       = 24
COND_ITEM_FOCUS_COL       = 25
COND_ITEM0_ID_COL         = 26
COND_ITEM0_QTY_COL        = 27
COND_ITEM1_ID_COL         = 28
COND_ITEM1_QTY_COL        = 29
COND_ITEM2_ID_COL         = 30
COND_ITEM2_QTY_COL        = 31
COND_ITEM3_ID_COL         = 32
COND_ITEM3_QTY_COL        = 33

# 조건부 template 전용 extra cols
COND_TMPL_VALID_TIME_COL  = 52
COND_TMPL_TRIGGER_COL     = 53
COND_TMPL_UNLOCK_COND_COL = 54
COND_TMPL_DONE_COL        = 55

# 일반상품 template done col
GEN_TMPL_DONE_COL = 52  # 36 BASE + 15 items(5×3) + 1


# ── product_infos 공통 삽입 ────────────────────────

def insert_product_info(pinfo_ws, dst_row: int, row, product_type: str, key: int) -> None:
    """product_infos 시트에 공통 행 삽입."""
    set_cell(pinfo_ws, dst_row, PINFO_KEY_COL, key)
    set_cell(pinfo_ws, dst_row, PINFO_FILTER_COL,       cv(row[FILTER_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_SUB_CAT_COL,      ci(row[SUB_CAT_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_PRODUCT_TYPE_COL, product_type)
    set_cell(pinfo_ws, dst_row, PINFO_SLOT_TYPE_COL,    cv(row[SLOT_TYPE_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_SLOT_THUMB_COL,   cv(row[SLOT_THUMB_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_SHOW_GOODS_COL,
             ci(row[SHOW_GOODS_COL - 1], default=DEFAULT_SHOW_GOODS))
    set_cell(pinfo_ws, dst_row, PINFO_SLOT_FX_COL,      ci(row[SLOT_FX_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_ORDER_COL,        ci(row[ORDER_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_POPUP_RES_COL,    cv(row[POPUP_RES_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_POPUP_THUMB_COL,  cv(row[POPUP_THUMB_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_POPUP_FX_COL,     ci(row[POPUP_FX_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_BADGE_COL,        ci(row[BADGE_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_TAG_COL,          ci(row[TAG_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_POPUP_DESC_COL,   cv(row[POPUP_DESC_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_DESC_TYPE_COL,    cv(row[DESC_TYPE_COL - 1]))
    set_cell(pinfo_ws, dst_row, PINFO_SHOW_REVIEW_COL,
             ci(row[SHOW_REVIEW_COL - 1], default=DEFAULT_SHOW_REVIEW))


# ── 아이템 슬롯 읽기 유틸 ──────────────────────────

def read_items_from_template(row, max_items: int = 4) -> list[tuple[int | None, int | None]]:
    """template row에서 아이템 (id, qty) 리스트 반환. xlookup name col(+1) 스킵."""
    items = []
    for i in range(max_items):
        id_col, qty_col = tmpl_item_cols(i)
        item_id  = ci(row[id_col - 1])
        item_qty = ci(row[qty_col - 1])
        items.append((item_id, item_qty))
    return items


# ── 일반상품 적용 ─────────────────────────────────

def apply_general(tmpl_ws, pinfo_ws, gen_ws) -> int:
    """일반상품 탭 → product_infos + general_store_products_shop."""
    next_pinfo_row = find_last_data_row(pinfo_ws) + 1
    next_gen_row   = find_last_data_row(gen_ws) + 1
    max_key        = max_key_in_sheet(pinfo_ws)

    added = 0
    for row in tmpl_ws.iter_rows(min_row=4, values_only=False):
        product_name = cv(row[PRODUCT_NAME_COL - 1])
        key_raw      = cv(row[KEY_COL - 1])

        if product_name is None and key_raw is None:
            continue

        if is_done(row, GEN_TMPL_DONE_COL):
            continue

        # key 결정
        if key_raw:
            key = int(float(key_raw))
        else:
            max_key += 1
            key = max_key

        # product_infos 삽입
        insert_product_info(pinfo_ws, next_pinfo_row, row,
                            "PRODUCT_TYPE_GENERAL_PRODUCT", key)

        # general_store_products_shop 삽입
        set_cell(gen_ws, next_gen_row, GEN_KEY_COL,            key)
        set_cell(gen_ws, next_gen_row, GEN_FILTER_COL,         cv(row[FILTER_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PRODUCT_NAME_COL,   product_name)
        set_cell(gen_ws, next_gen_row, GEN_SUB_NAME_COL,       cv(row[SUB_NAME_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_DESC_COL,           cv(row[DESC_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PRICE_KEY_COL,      cv(row[PRICE_KEY_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PRICE_P1_COL,       cv(row[PRICE_P1_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PRICE_P2_COL,       cv(row[PRICE_P2_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PRICE_P3_COL,       cv(row[PRICE_P3_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PRICE_P4_COL,       cv(row[PRICE_P4_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PRICE_P5_COL,       cv(row[PRICE_P5_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PRICE_P6_COL,       cv(row[PRICE_P6_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_START_TS_COL,       cv(row[START_TS_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_END_TS_COL,         cv(row[END_TS_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_PURCHASE_LIMIT_COL,
                 ci(row[PURCHASE_LIMIT_COL - 1], default=DEFAULT_PURCHASE_LIMIT))
        set_cell(gen_ws, next_gen_row, GEN_AVAILABLE_COL,
                 ci(row[AVAILABLE_COL - 1], default=DEFAULT_AVAILABLE))
        set_cell(gen_ws, next_gen_row, GEN_SHOW_SOLD_OUT_COL,
                 ci(row[SHOW_SOLD_OUT_COL - 1], default=DEFAULT_SHOW_SOLD_OUT))
        set_cell(gen_ws, next_gen_row, GEN_REFRESH_KEY_COL,    cv(row[REFRESH_KEY_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_REFRESH_P1_COL,     cv(row[REFRESH_P1_COL - 1]))
        set_cell(gen_ws, next_gen_row, GEN_REFRESH_P2_COL,     cv(row[REFRESH_P2_COL - 1]))

        # item_focus
        gen_ws.cell(row=next_gen_row, column=GEN_ITEM_FOCUS_COL, value=DEFAULT_ITEM_FOCUS)

        # 아이템 (최대 4개)
        item_sheet_cols = [
            (GEN_ITEM0_ID_COL, GEN_ITEM0_QTY_COL),
            (GEN_ITEM1_ID_COL, GEN_ITEM1_QTY_COL),
            (GEN_ITEM2_ID_COL, GEN_ITEM2_QTY_COL),
            (GEN_ITEM3_ID_COL, GEN_ITEM3_QTY_COL),
        ]
        for i, (id_scol, qty_scol) in enumerate(item_sheet_cols):
            id_tcol, qty_tcol = tmpl_item_cols(i)
            item_id  = ci(row[id_tcol - 1])
            item_qty = ci(row[qty_tcol - 1])
            set_cell(gen_ws, next_gen_row, id_scol,  item_id)
            set_cell(gen_ws, next_gen_row, qty_scol, item_qty)

        mark_done(row, GEN_TMPL_DONE_COL)
        next_pinfo_row += 1
        next_gen_row   += 1
        added += 1
        print(f"  [일반] key={key} '{product_name}' 추가")

    return added


# ── 조건부상품 적용 ───────────────────────────────

def apply_conditional(tmpl_ws, pinfo_ws, cond_ws) -> int:
    """조건부상품 탭 → product_infos + conditional_products."""
    next_pinfo_row = find_last_data_row(pinfo_ws) + 1
    next_cond_row  = find_last_data_row(cond_ws) + 1
    max_key        = max_key_in_sheet(pinfo_ws)

    added = 0
    for row in tmpl_ws.iter_rows(min_row=4, values_only=False):
        product_name = cv(row[PRODUCT_NAME_COL - 1])
        key_raw      = cv(row[KEY_COL - 1])

        if product_name is None and key_raw is None:
            continue

        if is_done(row, COND_TMPL_DONE_COL):
            continue

        # key 결정
        if key_raw:
            key = int(float(key_raw))
        else:
            max_key += 1
            key = max_key

        # product_infos 삽입
        insert_product_info(pinfo_ws, next_pinfo_row, row,
                            "PRODUCT_TYPE_CONDITIONAL_PRODUCT", key)

        # conditional_products 삽입
        set_cell(cond_ws, next_cond_row, COND_KEY_COL,            key)
        set_cell(cond_ws, next_cond_row, COND_FILTER_COL,         cv(row[FILTER_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_PRODUCT_NAME_COL,   product_name)
        set_cell(cond_ws, next_cond_row, COND_SUB_NAME_COL,       cv(row[SUB_NAME_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_DESC_COL,           cv(row[DESC_COL - 1]))
        # ref_unlock_condition_id — template C54
        set_cell(cond_ws, next_cond_row, COND_UNLOCK_COND_COL,
                 ci(row[COND_TMPL_UNLOCK_COND_COL - 1]))
        # 결제 방식 (%key, %param1~5 — conditional은 param6 없음)
        set_cell(cond_ws, next_cond_row, COND_PRICE_KEY_COL,  cv(row[PRICE_KEY_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_PRICE_P1_COL,   cv(row[PRICE_P1_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_PRICE_P2_COL,   cv(row[PRICE_P2_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_PRICE_P3_COL,   cv(row[PRICE_P3_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_PRICE_P4_COL,   cv(row[PRICE_P4_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_PRICE_P5_COL,   cv(row[PRICE_P5_COL - 1]))
        # valid_time / trigger_cooltime — template C52, C53
        set_cell(cond_ws, next_cond_row, COND_VALID_TIME_COL,
                 ci(row[COND_TMPL_VALID_TIME_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_TRIGGER_COOL_COL,
                 ci(row[COND_TMPL_TRIGGER_COL - 1]))
        # sub_description — template에 없음 → 빈칸
        set_cell(cond_ws, next_cond_row, COND_START_TS_COL,
                 cv(row[START_TS_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_END_TS_COL,
                 cv(row[END_TS_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_PURCHASE_LIMIT_COL,
                 ci(row[PURCHASE_LIMIT_COL - 1], default=DEFAULT_PURCHASE_LIMIT))
        set_cell(cond_ws, next_cond_row, COND_AVAILABLE_COL,
                 ci(row[AVAILABLE_COL - 1], default=DEFAULT_AVAILABLE))
        set_cell(cond_ws, next_cond_row, COND_SHOW_SOLD_OUT_COL,
                 ci(row[SHOW_SOLD_OUT_COL - 1], default=DEFAULT_SHOW_SOLD_OUT))
        set_cell(cond_ws, next_cond_row, COND_REFRESH_KEY_COL,
                 cv(row[REFRESH_KEY_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_REFRESH_P1_COL,
                 cv(row[REFRESH_P1_COL - 1]))
        set_cell(cond_ws, next_cond_row, COND_REFRESH_P2_COL,
                 cv(row[REFRESH_P2_COL - 1]))

        # item_focus
        cond_ws.cell(row=next_cond_row, column=COND_ITEM_FOCUS_COL,
                     value=DEFAULT_ITEM_FOCUS)

        # 아이템 (최대 4개)
        cond_item_sheet_cols = [
            (COND_ITEM0_ID_COL, COND_ITEM0_QTY_COL),
            (COND_ITEM1_ID_COL, COND_ITEM1_QTY_COL),
            (COND_ITEM2_ID_COL, COND_ITEM2_QTY_COL),
            (COND_ITEM3_ID_COL, COND_ITEM3_QTY_COL),
        ]
        for i, (id_scol, qty_scol) in enumerate(cond_item_sheet_cols):
            id_tcol, qty_tcol = tmpl_item_cols(i)
            item_id  = ci(row[id_tcol - 1])
            item_qty = ci(row[qty_tcol - 1])
            set_cell(cond_ws, next_cond_row, id_scol,  item_id)
            set_cell(cond_ws, next_cond_row, qty_scol, item_qty)

        mark_done(row, COND_TMPL_DONE_COL)
        next_pinfo_row += 1
        next_cond_row  += 1
        added += 1
        print(f"  [조건부] key={key} '{product_name}' 추가")

    return added


# ── 출석패키지 적용 ───────────────────────────────

# attendance_product 컬럼 상수
ATT_KEY_COL              = 1
ATT_REWARD_UNTIL_COL     = 2  # reward_until_days
ATT_MAIL_TMPL_COL        = 3  # mail_template_id
ATT_PRODUCT_NAME_COL     = 4
ATT_DESC_COL             = 5
ATT_SUB_DESC_COL         = 6
ATT_START_TS_COL         = 7
ATT_END_TS_COL           = 8
ATT_PURCHASE_LIMIT_COL   = 9
ATT_AVAILABLE_COL        = 10
ATT_SHOW_SOLD_OUT_COL    = 11
ATT_PRICE_KEY_COL        = 12
ATT_PRICE_P1_COL         = 13
ATT_PRICE_P2_COL         = 14
ATT_PRICE_P3_COL         = 15
ATT_PRICE_P4_COL         = 16
ATT_PRICE_P5_COL         = 17
# attendance_product에는 items/refresh 없음

# attendance_product_rewards 컬럼 상수
ATTR_KEY_COL         = 1  # ^key (attendance_product key)
ATTR_DAY_COL         = 2
ATTR_DAY_FOCUS_COL   = 3
ATTR_ITEM0_ID_COL    = 4
ATTR_ITEM0_QTY_COL   = 5
ATTR_ITEM1_ID_COL    = 6
ATTR_ITEM1_QTY_COL   = 7
ATTR_ITEM2_ID_COL    = 8
ATTR_ITEM2_QTY_COL   = 9

# 출석 template extra cols (BASE_NO_REFRESH = 33)
ATT_TMPL_REWARD_UNTIL_COL = 34  # C34
ATT_TMPL_MAIL_COL         = 35  # C35
ATT_TMPL_DAILY_START_COL  = 36  # C36: 1일_아이템0_ID
ATT_TMPL_DONE_COL         = 216 # C216 (33 + 2 + 180 + 1)
ATTENDANCE_DAYS           = 30
ATT_ITEMS_PER_DAY         = 3


def apply_attendance(tmpl_ws, pinfo_ws, att_ws, attr_ws) -> int:
    """출석패키지 탭 → product_infos + attendance_product + attendance_product_rewards."""
    next_pinfo_row = find_last_data_row(pinfo_ws) + 1
    next_att_row   = find_last_data_row(att_ws) + 1
    next_attr_row  = find_last_data_row(attr_ws) + 1
    max_key        = max_key_in_sheet(pinfo_ws)

    added = 0
    for row in tmpl_ws.iter_rows(min_row=4, values_only=False):
        product_name = cv(row[PRODUCT_NAME_COL - 1])
        key_raw      = cv(row[KEY_COL - 1])

        if product_name is None and key_raw is None:
            continue

        if is_done(row, ATT_TMPL_DONE_COL):
            continue

        # key 결정
        if key_raw:
            key = int(float(key_raw))
        else:
            max_key += 1
            key = max_key

        # product_infos 삽입
        insert_product_info(pinfo_ws, next_pinfo_row, row,
                            "PRODUCT_TYPE_ATTENDANCE_PRODUCT", key)

        # attendance_product 삽입
        set_cell(att_ws, next_att_row, ATT_KEY_COL,           key)
        set_cell(att_ws, next_att_row, ATT_REWARD_UNTIL_COL,
                 ci(row[ATT_TMPL_REWARD_UNTIL_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_MAIL_TMPL_COL,
                 ci(row[ATT_TMPL_MAIL_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_PRODUCT_NAME_COL,  product_name)
        set_cell(att_ws, next_att_row, ATT_DESC_COL,          cv(row[DESC_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_SUB_DESC_COL,      cv(row[POPUP_DESC_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_START_TS_COL,      cv(row[START_TS_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_END_TS_COL,        cv(row[END_TS_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_PURCHASE_LIMIT_COL,
                 ci(row[PURCHASE_LIMIT_COL - 1], default=DEFAULT_PURCHASE_LIMIT))
        set_cell(att_ws, next_att_row, ATT_AVAILABLE_COL,
                 ci(row[AVAILABLE_COL - 1], default=DEFAULT_AVAILABLE))
        set_cell(att_ws, next_att_row, ATT_SHOW_SOLD_OUT_COL,
                 ci(row[SHOW_SOLD_OUT_COL - 1], default=DEFAULT_SHOW_SOLD_OUT))
        set_cell(att_ws, next_att_row, ATT_PRICE_KEY_COL,     cv(row[PRICE_KEY_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_PRICE_P1_COL,      cv(row[PRICE_P1_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_PRICE_P2_COL,      cv(row[PRICE_P2_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_PRICE_P3_COL,      cv(row[PRICE_P3_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_PRICE_P4_COL,      cv(row[PRICE_P4_COL - 1]))
        set_cell(att_ws, next_att_row, ATT_PRICE_P5_COL,      cv(row[PRICE_P5_COL - 1]))

        # attendance_product_rewards — 1일~ATTENDANCE_DAYS일
        # 템플릿: C36~ (id, qty per item) × ATT_ITEMS_PER_DAY × ATTENDANCE_DAYS
        for d in range(1, ATTENDANCE_DAYS + 1):
            # 해당 일차의 아이템 읽기
            day_items = []
            for i in range(ATT_ITEMS_PER_DAY):
                col_offset = ((d - 1) * ATT_ITEMS_PER_DAY + i) * 2
                id_col  = ATT_TMPL_DAILY_START_COL + col_offset
                qty_col = ATT_TMPL_DAILY_START_COL + col_offset + 1
                item_id  = ci(row[id_col - 1])
                item_qty = ci(row[qty_col - 1])
                if item_id is not None:
                    day_items.append((item_id, item_qty))

            # 모든 아이템이 비어있으면 해당 일차 스킵
            if not day_items:
                continue

            set_cell(attr_ws, next_attr_row, ATTR_KEY_COL, key)
            set_cell(attr_ws, next_attr_row, ATTR_DAY_COL, d)
            set_cell(attr_ws, next_attr_row, ATTR_DAY_FOCUS_COL, 0)  # default
            # 아이템 쓰기
            item_cols = [
                (ATTR_ITEM0_ID_COL, ATTR_ITEM0_QTY_COL),
                (ATTR_ITEM1_ID_COL, ATTR_ITEM1_QTY_COL),
                (ATTR_ITEM2_ID_COL, ATTR_ITEM2_QTY_COL),
            ]
            for j, (item_id, item_qty) in enumerate(day_items[:3]):
                set_cell(attr_ws, next_attr_row, item_cols[j][0], item_id)
                set_cell(attr_ws, next_attr_row, item_cols[j][1], item_qty)
            next_attr_row += 1

        mark_done(row, ATT_TMPL_DONE_COL)
        next_pinfo_row += 1
        next_att_row   += 1
        added += 1
        print(f"  [출석] key={key} '{product_name}' 추가")

    return added


# ── 월정액 적용 ───────────────────────────────────

def _get_daily_reward_structure(ws):
    """daily_reward_product 시트 row3 헤더 확인."""
    return [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]


# daily_reward_product 컬럼 상수 (실제 시트)
DR_KEY_COL              = 1
DR_FILTER_COL           = 2
DR_PERIOD_DAYS_COL      = 3
DR_REPURCHASE_COL       = 4  # repurchase_available_days
DR_MAIL_TMPL_COL        = 5
DR_MAIL_BONUS_TMPL_COL  = 6
DR_PRODUCT_NAME_COL     = 7
DR_PRICE_KEY_COL        = 8
DR_PRICE_P1_COL         = 9
DR_PRICE_P2_COL         = 10
DR_PRICE_P3_COL         = 11
DR_PRICE_P4_COL         = 12
DR_PRICE_P5_COL         = 13
DR_DESC_COL             = 14
DR_SUB_DESC_COL         = 15
DR_START_TS_COL         = 16
DR_END_TS_COL           = 17
DR_PURCHASE_LIMIT_COL   = 18
DR_AVAILABLE_COL        = 19
DR_SHOW_SOLD_OUT_COL    = 20
DR_INSTANT_ID_COL       = 21  # 즉시지급 items/0/id
DR_INSTANT_QTY_COL      = 22
DR_BONUS_DAY_COL        = 23  # bonus_item/day
DR_BONUS_ID_COL         = 24  # bonus_item/id
DR_BONUS_QTY_COL        = 25  # bonus_item/qty
DR_DAILY_KEY_COL        = 26  # %key daily_reward
DR_DAILY_P1_COL         = 27
DR_DAILY_P2_COL         = 28

# 월정액 template extra cols (BASE_NO_REFRESH = 33)
# BASE_NO_REFRESH(33) + instant_items(3×3=9) = 42
# 즉시지급: C34~C42 (3 items × 3 cols)
# daily_reward: C43~C45 (%key, %param1, %param2)
# monthly_extra (period_days 등): 아래 참조

def _get_monthly_template_structure(gen_tmpl_ws):
    """월정액 template 컬럼 확인."""
    return {ws.cell(1, c).value: c for c in range(1, gen_tmpl_ws.max_column + 1)
            if gen_tmpl_ws.cell(1, c).value}


# 월정액 template extra col 상수 (gen script 기준 확정값)
# BASE_NO_REFRESH(33) + monthly_extra(4) + instant_items(3×3=9) + daily(3) + bonus(3) + 제작완료
DR_TMPL_PERIOD_DAYS_COL   = 34
DR_TMPL_REPURCHASE_COL    = 35
DR_TMPL_MAIL_COL          = 36
DR_TMPL_MAIL_BONUS_COL    = 37
DR_TMPL_INSTANT_ID_COL    = 38  # 즉시지급 아이템0 ID (xlookup = 39, qty = 40)
DR_TMPL_INSTANT_QTY_COL   = 40  # 즉시지급 수량
DR_TMPL_DAILY_KEY_COL     = 47  # 일일보상 타입 (%key)
DR_TMPL_DAILY_P1_COL      = 48  # 일일보상 아이템ID (%param1)
DR_TMPL_DAILY_P2_COL      = 49  # 일일보상 수량 (%param2)
DR_TMPL_BONUS_DAY_COL     = 50  # 보너스 보상 일차
DR_TMPL_BONUS_ID_COL      = 51  # 보너스 아이템ID
DR_TMPL_BONUS_QTY_COL     = 52  # 보너스 수량
DR_TMPL_DONE_COL          = 53  # 제작완료


def apply_monthly(tmpl_ws, pinfo_ws, dr_ws) -> int:
    """월정액 탭 → product_infos + daily_reward_product."""
    next_pinfo_row = find_last_data_row(pinfo_ws) + 1
    next_dr_row    = find_last_data_row(dr_ws) + 1
    max_key        = max_key_in_sheet(pinfo_ws)

    added = 0
    for row in tmpl_ws.iter_rows(min_row=4, values_only=False):
        product_name = cv(row[PRODUCT_NAME_COL - 1])
        key_raw      = cv(row[KEY_COL - 1])

        if product_name is None and key_raw is None:
            continue

        if is_done(row, DR_TMPL_DONE_COL):
            continue

        # key 결정
        if key_raw:
            key = int(float(key_raw))
        else:
            max_key += 1
            key = max_key

        # product_infos 삽입
        insert_product_info(pinfo_ws, next_pinfo_row, row,
                            "PRODUCT_TYPE_DAILY_REWARD_PRODUCT", key)

        # daily_reward_product 삽입
        set_cell(dr_ws, next_dr_row, DR_KEY_COL,            key)
        set_cell(dr_ws, next_dr_row, DR_FILTER_COL,          cv(row[FILTER_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_PERIOD_DAYS_COL,
                 ci(row[DR_TMPL_PERIOD_DAYS_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_REPURCHASE_COL,
                 ci(row[DR_TMPL_REPURCHASE_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_MAIL_TMPL_COL,
                 ci(row[DR_TMPL_MAIL_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_MAIL_BONUS_TMPL_COL,
                 ci(row[DR_TMPL_MAIL_BONUS_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_PRODUCT_NAME_COL,    product_name)
        set_cell(dr_ws, next_dr_row, DR_PRICE_KEY_COL,       cv(row[PRICE_KEY_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_PRICE_P1_COL,        cv(row[PRICE_P1_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_PRICE_P2_COL,        cv(row[PRICE_P2_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_PRICE_P3_COL,        cv(row[PRICE_P3_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_PRICE_P4_COL,        cv(row[PRICE_P4_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_PRICE_P5_COL,        cv(row[PRICE_P5_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_DESC_COL,            cv(row[DESC_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_SUB_DESC_COL,        cv(row[POPUP_DESC_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_START_TS_COL,        cv(row[START_TS_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_END_TS_COL,          cv(row[END_TS_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_PURCHASE_LIMIT_COL,
                 ci(row[PURCHASE_LIMIT_COL - 1], default=DEFAULT_PURCHASE_LIMIT))
        set_cell(dr_ws, next_dr_row, DR_AVAILABLE_COL,
                 ci(row[AVAILABLE_COL - 1], default=DEFAULT_AVAILABLE))
        set_cell(dr_ws, next_dr_row, DR_SHOW_SOLD_OUT_COL,
                 ci(row[SHOW_SOLD_OUT_COL - 1], default=DEFAULT_SHOW_SOLD_OUT))

        # 즉시지급 아이템0 (id, [xlookup name], qty)
        instant_id  = ci(row[DR_TMPL_INSTANT_ID_COL - 1])
        instant_qty = ci(row[DR_TMPL_INSTANT_QTY_COL - 1])
        set_cell(dr_ws, next_dr_row, DR_INSTANT_ID_COL,  instant_id)
        set_cell(dr_ws, next_dr_row, DR_INSTANT_QTY_COL, instant_qty)

        # 보너스 아이템 (bonus_item)
        set_cell(dr_ws, next_dr_row, DR_BONUS_DAY_COL,
                 ci(row[DR_TMPL_BONUS_DAY_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_BONUS_ID_COL,
                 ci(row[DR_TMPL_BONUS_ID_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_BONUS_QTY_COL,
                 ci(row[DR_TMPL_BONUS_QTY_COL - 1]))

        # 일일 지급 방식 (%key, %param1, %param2)
        set_cell(dr_ws, next_dr_row, DR_DAILY_KEY_COL,
                 cv(row[DR_TMPL_DAILY_KEY_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_DAILY_P1_COL,
                 cv(row[DR_TMPL_DAILY_P1_COL - 1]))
        set_cell(dr_ws, next_dr_row, DR_DAILY_P2_COL,
                 cv(row[DR_TMPL_DAILY_P2_COL - 1]))

        mark_done(row, DR_TMPL_DONE_COL)
        next_pinfo_row += 1
        next_dr_row    += 1
        added += 1
        print(f"  [월정액] key={key} '{product_name}' 추가")

    return added


# ── 프리미엄패스 적용 ──────────────────────────────

# 프리미엄패스 template extra col 상수 (gen script 기준 확정값)
# BASE_NO_REFRESH(33) + product_id(C34) + instant_items(3×3=9: C35~C43) + 제작완료(C44)
PP_TMPL_PRODUCT_ID_COL  = 34
PP_TMPL_INSTANT_ID_COL  = 35  # 즉시지급 아이템0 ID (xlookup = 36, qty = 37)
PP_TMPL_INSTANT_QTY_COL = 37
PP_TMPL_DONE_COL        = 44

# premium_pass_product 컬럼 상수
PP_PRODUCT_ID_COL        = 1   # product_id (단일 행)
PP_PRODUCT_NAME_COL      = 2
PP_DESC_COL              = 3
# col4 = empty?
PP_START_TS_COL          = 5
# col6 = empty?
PP_END_TS_COL            = 7
PP_PURCHASE_LIMIT_COL    = 8
PP_AVAILABLE_COL         = 9
PP_SHOW_SOLD_OUT_COL     = 10
PP_ITEM0_ID_COL          = 11  # items/0/id
PP_ITEM0_QTY_COL         = 12
PP_PRICE_KEY_COL         = 19
PP_PRICE_P1_COL          = 20
PP_PRICE_P2_COL          = 21
PP_PRICE_P3_COL          = 22
PP_PRICE_P4_COL          = 23
PP_PRICE_P5_COL          = 24
PP_PRICE_P6_COL          = 25


def apply_premium_pass(tmpl_ws, pinfo_ws, pp_ws) -> int:
    """프리미엄패스 탭 → product_infos + premium_pass_product."""
    next_pinfo_row = find_last_data_row(pinfo_ws) + 1
    next_pp_row    = find_last_data_row(pp_ws) + 1
    max_key        = max_key_in_sheet(pinfo_ws)

    added = 0
    for row in tmpl_ws.iter_rows(min_row=4, values_only=False):
        product_name = cv(row[PRODUCT_NAME_COL - 1])
        key_raw      = cv(row[KEY_COL - 1])

        if product_name is None and key_raw is None:
            continue

        if is_done(row, PP_TMPL_DONE_COL):
            continue

        # key 결정
        if key_raw:
            key = int(float(key_raw))
        else:
            max_key += 1
            key = max_key

        # product_infos 삽입
        insert_product_info(pinfo_ws, next_pinfo_row, row,
                            "PRODUCT_TYPE_PREMIUM_PASS_PRODUCT", key)

        # premium_pass_product 삽입
        # product_id = template C34 값, 없으면 key로 폴백
        product_id = ci(row[PP_TMPL_PRODUCT_ID_COL - 1]) or key
        set_cell(pp_ws, next_pp_row, PP_PRODUCT_ID_COL,     product_id)
        set_cell(pp_ws, next_pp_row, PP_PRODUCT_NAME_COL,   product_name)
        set_cell(pp_ws, next_pp_row, PP_DESC_COL,           cv(row[DESC_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_START_TS_COL,       cv(row[START_TS_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_END_TS_COL,         cv(row[END_TS_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_PURCHASE_LIMIT_COL,
                 ci(row[PURCHASE_LIMIT_COL - 1], default=DEFAULT_PURCHASE_LIMIT))
        set_cell(pp_ws, next_pp_row, PP_AVAILABLE_COL,
                 ci(row[AVAILABLE_COL - 1], default=DEFAULT_AVAILABLE))
        set_cell(pp_ws, next_pp_row, PP_SHOW_SOLD_OUT_COL,
                 ci(row[SHOW_SOLD_OUT_COL - 1], default=DEFAULT_SHOW_SOLD_OUT))

        # 즉시지급 아이템0 (C35: id, C36: xlookup, C37: qty)
        instant_id  = ci(row[PP_TMPL_INSTANT_ID_COL - 1])
        instant_qty = ci(row[PP_TMPL_INSTANT_QTY_COL - 1])
        set_cell(pp_ws, next_pp_row, PP_ITEM0_ID_COL,  instant_id)
        set_cell(pp_ws, next_pp_row, PP_ITEM0_QTY_COL, instant_qty)

        # 결제 방식 (%key ~ %param6)
        set_cell(pp_ws, next_pp_row, PP_PRICE_KEY_COL, cv(row[PRICE_KEY_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_PRICE_P1_COL,  cv(row[PRICE_P1_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_PRICE_P2_COL,  cv(row[PRICE_P2_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_PRICE_P3_COL,  cv(row[PRICE_P3_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_PRICE_P4_COL,  cv(row[PRICE_P4_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_PRICE_P5_COL,  cv(row[PRICE_P5_COL - 1]))
        set_cell(pp_ws, next_pp_row, PP_PRICE_P6_COL,  cv(row[PRICE_P6_COL - 1]))

        mark_done(row, PP_TMPL_DONE_COL)
        next_pinfo_row += 1
        next_pp_row    += 1
        added += 1
        print(f"  [패스] product_id={product_id} key={key} '{product_name}' 추가")

    return added


# ── 스텝상품 적용 ─────────────────────────────────

# step_products 컬럼 상수 (실제 시트)
STEP_KEY_COL              = 1
STEP_PRODUCT_NAME_COL     = 2
STEP_DESC_COL             = 3
STEP_START_TS_COL         = 4
STEP_END_TS_COL           = 5
STEP_PURCHASE_LIMIT_COL   = 6
STEP_AVAILABLE_COL        = 7
STEP_SHOW_SOLD_OUT_COL    = 8
STEP_REFRESH_KEY_COL      = 9
STEP_REFRESH_P1_COL       = 10
STEP_REFRESH_P2_COL       = 11
STEP_BASE_PRICE_KEY_COL   = 12  # %key product_base/price_type
STEP_BASE_PRICE_P1_COL    = 13
STEP_BASE_PRICE_P2_COL    = 14
STEP_BASE_PRICE_P3_COL    = 15
# steps/0:
STEP0_PRICE_ID_COL  = 16
STEP0_PRICE_QTY_COL = 17
STEP0_ITEM0_ID_COL  = 18
STEP0_ITEM0_QTY_COL = 19
STEP0_ITEM1_ID_COL  = 20
STEP0_ITEM1_QTY_COL = 21
STEP0_ITEM2_ID_COL  = 22
STEP0_ITEM2_QTY_COL = 23
STEP0_ITEM3_ID_COL  = 24
STEP0_ITEM3_QTY_COL = 25
# steps/1:
STEP1_PRICE_ID_COL  = 26
STEP1_PRICE_QTY_COL = 27
STEP1_ITEM0_ID_COL  = 28
STEP1_ITEM0_QTY_COL = 29
STEP1_ITEM1_ID_COL  = 30
STEP1_ITEM1_QTY_COL = 31
STEP1_ITEM2_ID_COL  = 32
STEP1_ITEM2_QTY_COL = 33
STEP1_ITEM3_ID_COL  = 34
STEP1_ITEM3_QTY_COL = 35
# steps/2:
STEP2_PRICE_ID_COL  = 36
STEP2_PRICE_QTY_COL = 37
STEP2_ITEM0_ID_COL  = 38
STEP2_ITEM0_QTY_COL = 39
STEP2_ITEM1_ID_COL  = 40
STEP2_ITEM1_QTY_COL = 41
STEP2_ITEM2_ID_COL  = 42
STEP2_ITEM2_QTY_COL = 43
STEP2_ITEM3_ID_COL  = 44
STEP2_ITEM3_QTY_COL = 45

# 스텝상품 template cols
# BASE_HEADERS(36) + step1 + step2 + step3 + [제작완료]
# 스텝 template: per step = 2(price) + 4*(3cols: id+name+qty) = 14
# step1 시작: C37, step2: C51, step3: C65
STEP_TMPL_STEP_START = 37
STEP_TMPL_STEP_COLS  = 14   # 2 + 4*3
STEP_COUNT_TMPL      = 3
STEP_ITEMS_TMPL      = 4

# 스텝 template에서 스텝 s(0-indexed) 아이템 i의 (id_col, qty_col)
def tmpl_step_item_cols(s: int, i: int) -> tuple[int, int]:
    step_base = STEP_TMPL_STEP_START + s * STEP_TMPL_STEP_COLS
    # +2: price skip, then i*3 (id, [name], qty)
    id_col  = step_base + 2 + i * 3
    qty_col = step_base + 2 + i * 3 + 2
    return (id_col, qty_col)


def tmpl_step_price_cols(s: int) -> tuple[int, int]:
    step_base = STEP_TMPL_STEP_START + s * STEP_TMPL_STEP_COLS
    return (step_base, step_base + 1)


STEP_TMPL_DONE_COL = STEP_TMPL_STEP_START + STEP_COUNT_TMPL * STEP_TMPL_STEP_COLS  # C79


def apply_step(tmpl_ws, pinfo_ws, step_ws) -> int:
    """스텝상품 탭 → product_infos + step_products."""
    next_pinfo_row = find_last_data_row(pinfo_ws) + 1
    next_step_row  = find_last_data_row(step_ws) + 1
    max_key        = max_key_in_sheet(pinfo_ws)

    # step_products 실제 시트 컬럼 구조 확인 (steps/0~2 × items/0~3)
    step_sheet_steps = [
        {
            "price_id":  STEP0_PRICE_ID_COL,  "price_qty": STEP0_PRICE_QTY_COL,
            "items": [(STEP0_ITEM0_ID_COL, STEP0_ITEM0_QTY_COL),
                      (STEP0_ITEM1_ID_COL, STEP0_ITEM1_QTY_COL),
                      (STEP0_ITEM2_ID_COL, STEP0_ITEM2_QTY_COL),
                      (STEP0_ITEM3_ID_COL, STEP0_ITEM3_QTY_COL)],
        },
        {
            "price_id":  STEP1_PRICE_ID_COL,  "price_qty": STEP1_PRICE_QTY_COL,
            "items": [(STEP1_ITEM0_ID_COL, STEP1_ITEM0_QTY_COL),
                      (STEP1_ITEM1_ID_COL, STEP1_ITEM1_QTY_COL),
                      (STEP1_ITEM2_ID_COL, STEP1_ITEM2_QTY_COL),
                      (STEP1_ITEM3_ID_COL, STEP1_ITEM3_QTY_COL)],
        },
        {
            "price_id":  STEP2_PRICE_ID_COL,  "price_qty": STEP2_PRICE_QTY_COL,
            "items": [(STEP2_ITEM0_ID_COL, STEP2_ITEM0_QTY_COL),
                      (STEP2_ITEM1_ID_COL, STEP2_ITEM1_QTY_COL),
                      (STEP2_ITEM2_ID_COL, STEP2_ITEM2_QTY_COL),
                      (STEP2_ITEM3_ID_COL, STEP2_ITEM3_QTY_COL)],
        },
    ]

    added = 0
    for row in tmpl_ws.iter_rows(min_row=4, values_only=False):
        product_name = cv(row[PRODUCT_NAME_COL - 1])
        key_raw      = cv(row[KEY_COL - 1])

        if product_name is None and key_raw is None:
            continue

        if is_done(row, STEP_TMPL_DONE_COL):
            continue

        # key 결정
        if key_raw:
            key = int(float(key_raw))
        else:
            max_key += 1
            key = max_key

        # product_infos 삽입
        insert_product_info(pinfo_ws, next_pinfo_row, row,
                            "PRODUCT_TYPE_STEP_PRODUCT", key)

        # step_products 삽입
        set_cell(step_ws, next_step_row, STEP_KEY_COL,           key)
        set_cell(step_ws, next_step_row, STEP_PRODUCT_NAME_COL,  product_name)
        set_cell(step_ws, next_step_row, STEP_DESC_COL,          cv(row[DESC_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_START_TS_COL,      cv(row[START_TS_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_END_TS_COL,        cv(row[END_TS_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_PURCHASE_LIMIT_COL,
                 ci(row[PURCHASE_LIMIT_COL - 1], default=DEFAULT_PURCHASE_LIMIT))
        set_cell(step_ws, next_step_row, STEP_AVAILABLE_COL,
                 ci(row[AVAILABLE_COL - 1], default=DEFAULT_AVAILABLE))
        set_cell(step_ws, next_step_row, STEP_SHOW_SOLD_OUT_COL,
                 ci(row[SHOW_SOLD_OUT_COL - 1], default=DEFAULT_SHOW_SOLD_OUT))
        set_cell(step_ws, next_step_row, STEP_REFRESH_KEY_COL,   cv(row[REFRESH_KEY_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_REFRESH_P1_COL,    cv(row[REFRESH_P1_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_REFRESH_P2_COL,    cv(row[REFRESH_P2_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_BASE_PRICE_KEY_COL, cv(row[PRICE_KEY_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_BASE_PRICE_P1_COL,  cv(row[PRICE_P1_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_BASE_PRICE_P2_COL,  cv(row[PRICE_P2_COL - 1]))
        set_cell(step_ws, next_step_row, STEP_BASE_PRICE_P3_COL,  cv(row[PRICE_P3_COL - 1]))

        # 스텝별 데이터
        for s, step_info in enumerate(step_sheet_steps):
            price_id_col, price_qty_col = tmpl_step_price_cols(s)
            price_id  = ci(row[price_id_col - 1])
            price_qty = ci(row[price_qty_col - 1])
            set_cell(step_ws, next_step_row, step_info["price_id"],  price_id)
            set_cell(step_ws, next_step_row, step_info["price_qty"], price_qty)

            for i, (sheet_id_col, sheet_qty_col) in enumerate(step_info["items"]):
                tmpl_id_col, tmpl_qty_col = tmpl_step_item_cols(s, i)
                item_id  = ci(row[tmpl_id_col - 1])
                item_qty = ci(row[tmpl_qty_col - 1])
                set_cell(step_ws, next_step_row, sheet_id_col,  item_id)
                set_cell(step_ws, next_step_row, sheet_qty_col, item_qty)

        mark_done(row, STEP_TMPL_DONE_COL)
        next_pinfo_row += 1
        next_step_row  += 1
        added += 1
        print(f"  [스텝] key={key} '{product_name}' 추가")

    return added


# ── 메인 ──────────────────────────────────────────

MODE_ALL    = "all"
MODE_MAP = {
    "일반":   "일반상품",
    "조건부":  "조건부상품",
    "출석":   "출석패키지",
    "월정액":  "월정액",
    "패스":   "프리미엄패스",
    "스텝":   "스텝상품",
}
ALL_MODES = list(MODE_MAP.keys())


def main() -> None:
    args = sys.argv[1:]
    if not args:
        print("사용법: python apply-product-input.py <mode> [template.xlsx]")
        print("  mode: all | 일반 | 조건부 | 출석 | 월정액 | 패스 | 스텝")
        sys.exit(1)

    mode_arg = args[0]
    tmpl_path = Path(args[1]) if len(args) > 1 else DEFAULT_TMPL

    if not tmpl_path.exists():
        print(f"[ERROR] 템플릿 파일 없음: {tmpl_path}")
        sys.exit(1)

    if not PRODUCT_XLS.exists():
        print(f"[ERROR] product.xlsx 없음: {PRODUCT_XLS}")
        sys.exit(1)

    # 처리할 mode 결정
    if mode_arg == MODE_ALL:
        modes = ALL_MODES
    elif mode_arg in MODE_MAP:
        modes = [mode_arg]
    else:
        print(f"[ERROR] 알 수 없는 mode: {mode_arg}")
        print(f"  사용 가능: all | {' | '.join(ALL_MODES)}")
        sys.exit(1)

    print(f"템플릿: {tmpl_path}")
    print(f"대상:   {PRODUCT_XLS}")
    print(f"모드:   {mode_arg}")
    print()

    # 파일 로드
    tmpl_wb = openpyxl.load_workbook(str(tmpl_path))
    prod_wb = openpyxl.load_workbook(str(PRODUCT_XLS))

    pinfo_ws = prod_wb["product_infos"]
    total_added = 0

    for mode in modes:
        tab_name = MODE_MAP[mode]
        if tab_name not in tmpl_wb.sheetnames:
            print(f"  [SKIP] 탭 없음: '{tab_name}'")
            continue

        tmpl_ws = tmpl_wb[tab_name]

        if mode == "일반":
            gen_ws = prod_wb["general_store_products_shop"]
            n = apply_general(tmpl_ws, pinfo_ws, gen_ws)

        elif mode == "조건부":
            cond_ws = prod_wb["conditional_products"]
            n = apply_conditional(tmpl_ws, pinfo_ws, cond_ws)

        elif mode == "출석":
            att_ws  = prod_wb["attendance_product"]
            attr_ws = prod_wb["duct.attendance_product_rewards"]
            n = apply_attendance(tmpl_ws, pinfo_ws, att_ws, attr_ws)

        elif mode == "월정액":
            dr_ws = prod_wb["daily_reward_product"]
            n = apply_monthly(tmpl_ws, pinfo_ws, dr_ws)

        elif mode == "패스":
            pp_ws = prod_wb["premium_pass_product"]
            n = apply_premium_pass(tmpl_ws, pinfo_ws, pp_ws)

        elif mode == "스텝":
            step_ws = prod_wb["step_products"]
            n = apply_step(tmpl_ws, pinfo_ws, step_ws)

        else:
            n = 0

        total_added += n

    if total_added == 0:
        print("추가된 상품 없음 (모두 제작완료 처리되었거나 빈 행).")
        return

    # 저장
    prod_wb.save(str(PRODUCT_XLS))
    tmpl_wb.save(str(tmpl_path))
    print()
    print(f"[OK] 총 {total_added}개 상품 추가")
    print(f"     product.xlsx 저장: {PRODUCT_XLS}")
    print(f"     template 저장 (제작완료 기입): {tmpl_path}")


if __name__ == "__main__":
    main()
