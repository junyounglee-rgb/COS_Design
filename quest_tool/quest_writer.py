"""
quest_writer.py -- quests.xlsx Excel 읽기/쓰기 모듈.
Streamlit에 의존하지 않는 순수 Python 함수로 구현.
캐시는 app.py에서 @st.cache_data로 감쌀 예정.
"""
import os
import re
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_BASE_PATH_PREFIX = "quests/^0/"
_PATH_ROW = 2
_HEADER_ROW = 3

# goal_types.yaml seed — 파일 부존재 시 폴백.
# UI 쪽의 GOAL_TYPES 하드코딩과 동기 유지. `play` 추가 (실데이터 child 8건 사용).
_HARDCODED_GOAL_TYPES: list[dict] = [
    {"key": "daily_login", "label": "출석하기", "params": []},
    {"key": "play", "label": "플레이", "params": []},
    {"key": "play:need_win", "label": "플레이/승리",
     "params": [{"label": "need_win", "options": ["FALSE", "TRUE"]}]},
    {"key": "play_mvp", "label": "MVP", "params": []},
    {"key": "reward_quest:ref_quest_ids", "label": "퀘스트 완료",
     "params": [{"label": "[]{id1,id2,...}", "free_text": True}]},
    # ... 상세 목록은 goal_types.yaml 참조. seed 만 제공.
]


# ---------------------------------------------------------------------------
# 내부 헬퍼
# ---------------------------------------------------------------------------

def _build_header_map(
    ws,
    path_row: int = _PATH_ROW,
    header_row: int = _HEADER_ROW,
    base_prefix: str = _BASE_PATH_PREFIX,
) -> dict[str, int]:
    """Row 2(경로) + Row 3(헤더) 조합으로 {고유키: 컬럼인덱스(1-based)} 반환.
    base_prefix와 일치하는 경로의 컬럼: 헤더 이름 그대로 사용.
    경로가 다른 컬럼: 경로suffix/헤더명 형식으로 고유 키 생성.
    base_prefix가 빈 문자열이면 path_row를 무시하고 헤더명만 사용.
    """
    paths: dict[int, str] = {}
    if base_prefix:
        for cell in ws[path_row]:
            if cell.value is not None:
                paths[cell.column] = str(cell.value)

    result: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        field = str(cell.value)
        col = cell.column
        path = paths.get(col, "")
        # base_prefix 제거 후 suffix 추출
        if base_prefix and path.startswith(base_prefix):
            suffix = path[len(base_prefix):].rstrip("/")
        elif base_prefix:
            suffix = path.rstrip("/")
        else:
            suffix = ""

        if suffix:
            key = f"{suffix}/{field}"
        else:
            key = field
        result[key] = col
    return result


# ---------------------------------------------------------------------------
# 공개 읽기 함수
# ---------------------------------------------------------------------------

def get_header_map(
    xlsx_path: str,
    sheet: str = "quests",
    path_row: int = _PATH_ROW,
    header_row: int = _HEADER_ROW,
) -> dict[str, int]:
    """헤더 행에서 {고유컬럼키: 컬럼인덱스} 반환."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        return _build_header_map(ws, path_row=path_row, header_row=header_row)
    finally:
        wb.close()


def get_existing_keys(xlsx_path: str) -> set[int]:
    """quests 시트 ^key 컬럼의 기존 값 전부 수집."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["quests"]
        header_map = _build_header_map(ws, path_row=_PATH_ROW, header_row=_HEADER_ROW)
        key_col = header_map.get("^key")
        if key_col is None:
            return set()
        existing: set[int] = set()
        for row in ws.iter_rows(min_row=4, values_only=True):
            val = row[key_col - 1]  # 0-based index
            if val is not None:
                try:
                    existing.add(int(val))
                except (ValueError, TypeError):
                    pass
        return existing
    finally:
        wb.close()


def get_existing_keys_by_filter(xlsx_path: str) -> dict[int, str]:
    """quests 시트 {^key: $filter} 매핑 수집.

    R-02 (2026-04-24 Round 2) — suggest_next_parent_key 의 filter-scoped max
    계산에 사용. filter 값이 빈 셀인 행은 빈 문자열("")로 기록.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["quests"]
        header_map = _build_header_map(ws, path_row=_PATH_ROW, header_row=_HEADER_ROW)
        key_col = header_map.get("^key")
        filter_col = header_map.get("$filter")
        if key_col is None:
            return {}
        result: dict[int, str] = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            key_val = row[key_col - 1]
            if key_val is None:
                continue
            try:
                k = int(key_val)
            except (ValueError, TypeError):
                continue
            filter_val = ""
            if filter_col is not None:
                raw = row[filter_col - 1]
                if raw is not None:
                    filter_val = str(raw)
            result[k] = filter_val
        return result
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# ^key 생성 규칙 (quests.xlsx 실 데이터 분석 결과)
# ---------------------------------------------------------------------------
#
# 카테고리 + reset_type 별 키 범위 (실 데이터 기반):
#   GENERAL / NONE(빈값):                     31 ~ 73705  (일반 퀘스트)
#   GENERAL / QUEST_RESET_TYPE_NONE:          30011 ~ 60105
#   GENERAL / QUEST_RESET_TYPE_DAILY:         30001 ~ 51007 (데일리 서브)
#   GENERAL / QUEST_RESET_TYPE_REPEAT:        50008 ~ 51010 (데일리 메인)
#   TOWN:                                     810~830, 5600번대, 500000번대, 10~11자리
#   VAULT_MISSION:                            미사용 (예약)
#
# 발급 전략: 해당 범위의 기존 max + 1 부터 시작, 중복 시 +1 증가 (예측 가능하도록).

KEY_RANGES: dict[tuple[str, str | None], tuple[int, int]] = {
    # (category, reset_type): (lo, hi)  — None은 빈 값/QUEST_RESET_TYPE_NONE 취급
    ("QUEST_CATEGORY_GENERAL", "QUEST_RESET_TYPE_DAILY"): (30001, 59999),
    ("QUEST_CATEGORY_GENERAL", "QUEST_RESET_TYPE_REPEAT"): (50008, 59999),
    ("QUEST_CATEGORY_GENERAL", "QUEST_RESET_TYPE_NONE"): (30011, 99999),
    ("QUEST_CATEGORY_GENERAL", None): (30011, 99999),  # reset_type 빈 값 = NONE 동일
    ("QUEST_CATEGORY_TOWN", None): (500000, 599999),  # 광장: 6자리 500000번대 (간단 케이스)
    ("QUEST_CATEGORY_VAULT_MISSION", None): (80000, 89999),  # 예약
}

DEFAULT_KEY_RANGE: tuple[int, int] = (100000, 999999)


def _get_key_range(category: str | None, reset_type: str | None) -> tuple[int, int]:
    """카테고리/리셋타입에 맞는 키 범위 반환. 매칭 실패 시 DEFAULT_KEY_RANGE."""
    # reset_type 정규화: 빈 문자열 → None
    rt = reset_type if reset_type else None
    # 1) 정확한 매칭
    if (category, rt) in KEY_RANGES:
        return KEY_RANGES[(category, rt)]
    # 2) reset_type=None 폴백 매칭
    if (category, None) in KEY_RANGES:
        return KEY_RANGES[(category, None)]
    # 3) 기본값
    return DEFAULT_KEY_RANGE


def generate_unique_key(
    existing: set[int],
    category: str | None = None,
    reset_type: str | None = None,
    lo: int | None = None,
    hi: int | None = None,
) -> int:
    """카테고리 + reset_type 기준 키 범위에서 유니크 ^key 발급.

    전략:
      - 해당 범위의 `기존 max + 1`부터 시작 (기존 데이터 끝에 자연스럽게 이어지도록).
      - 해당 범위 내 기존 키가 없으면 lo부터 시작.
      - 충돌 시 +1 증가하며 hi까지 탐색.
      - 범위가 꽉 찼으면 ValueError.

    Args:
        existing: 전체 quests.xlsx 기존 ^key 집합.
        category: QUEST_CATEGORY_* (없으면 DEFAULT_KEY_RANGE 사용).
        reset_type: QUEST_RESET_TYPE_* 또는 None.
        lo, hi: 범위 override (테스트/하위호환용). 둘 다 지정 시 KEY_RANGES 무시.
    """
    # lo/hi override 처리 (하위 호환)
    if lo is not None and hi is not None:
        range_lo, range_hi = lo, hi
    else:
        range_lo, range_hi = _get_key_range(category, reset_type)

    # 범위 내 기존 키의 최대값 찾기
    in_range = [k for k in existing if range_lo <= k <= range_hi]
    start = (max(in_range) + 1) if in_range else range_lo

    # start부터 hi까지 순차 탐색
    for candidate in range(start, range_hi + 1):
        if candidate not in existing:
            return candidate

    # 못 찾으면 범위 앞쪽부터 빈 자리 찾기 (예외 케이스)
    for candidate in range(range_lo, start):
        if candidate not in existing:
            return candidate

    raise ValueError(
        f"키 범위 {range_lo}~{range_hi} 가 꽉 찼습니다. "
        f"(category={category}, reset_type={reset_type})"
    )


def load_items(items_path: str) -> list[dict]:
    """items.xlsx -> [{id: int, name: str, category: str, filter: str}] 목록.
    헤더행 3, 컬럼: ^key, name, category, $filter 을 _build_header_map()으로 동적 탐색.
    id가 없는 행 제외. filter 는 누락 시 빈 문자열."""
    wb = load_workbook(items_path, read_only=True, data_only=True)
    try:
        ws = wb["items"] if "items" in wb.sheetnames else wb[wb.sheetnames[0]]
        header_map = _build_header_map(ws, header_row=3, base_prefix="")

        key_col = header_map.get("^key")
        name_col = header_map.get("name")
        cat_col = header_map.get("category")
        filter_col = header_map.get("$filter")

        if key_col is None:
            return []

        items: list[dict] = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            raw_id = row[key_col - 1]
            if raw_id is None:
                continue
            try:
                item_id = int(raw_id)
            except (ValueError, TypeError):
                continue

            name = row[name_col - 1] if name_col else None
            category = row[cat_col - 1] if cat_col else None
            filter_val = row[filter_col - 1] if filter_col else None

            items.append({
                "id": item_id,
                "name": str(name) if name is not None else "",
                "category": str(category) if category is not None else "",
                "filter": str(filter_val) if filter_val is not None else "",
            })
        return items
    finally:
        wb.close()


def load_item_categories(items_path: str) -> list[dict]:
    """items.xlsx `# ItemCategory` 시트 -> [{key, label}] 목록.
    시트 구조 (실 데이터 기준):
      Row 1: header (A='key', B='value') — 한글 라벨과 enum 값 매핑
      Row 2+: A=라벨(한글), B=enum 값 (ITEM_CATEGORY_*)
    반환 형식: [{"key": "ITEM_CATEGORY_GENERAL", "label": "일반"}, ...]
    시트 부존재 시 빈 리스트 반환."""
    wb = load_workbook(items_path, read_only=True, data_only=True)
    try:
        sheet_name = None
        for s in wb.sheetnames:
            # '# ItemCategory' 또는 유사 시트명 대응 (공백/대소문자 관대)
            if s.strip().lower() in ("# itemcategory", "#itemcategory", "itemcategory"):
                sheet_name = s
                break
        if sheet_name is None:
            return []

        ws = wb[sheet_name]
        categories: list[dict] = []
        # Row 1 은 헤더 ('key', 'value'), Row 2 부터 데이터
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            label = row[0]
            key = row[1]
            if key is None:
                continue
            key_str = str(key).strip()
            if not key_str.startswith("ITEM_CATEGORY_"):
                continue
            categories.append({
                "key": key_str,
                "label": str(label).strip() if label is not None else key_str,
            })
        return categories
    finally:
        wb.close()


def load_keywords(keywords_path: str) -> dict[str, dict[str, str]]:
    """keywords.xlsx -> {"build": {id: value}, "timestamp": {id: value}}.
    build 시트: 헤더행 3, A=id, B=value.
    timestamp 시트: 헤더행 3, A=id, B=value.
    시트가 없으면 빈 dict 반환."""
    wb = load_workbook(keywords_path, read_only=True, data_only=True)
    result: dict[str, dict[str, str]] = {}
    try:
        for sheet_name in ["build", "timestamp"]:
            if sheet_name not in wb.sheetnames:
                result[sheet_name] = {}
                continue
            ws = wb[sheet_name]
            mapping: dict[str, str] = {}
            for row in ws.iter_rows(min_row=4, values_only=True):
                id_val = row[0]
                val = row[1]
                if id_val is not None:
                    mapping[str(id_val)] = str(val) if val is not None else ""
            result[sheet_name] = mapping
        return result
    finally:
        wb.close()


def load_quest_templates(quests_path: str) -> dict[str, str]:
    """#quest 시트 -> {B컬럼(quest텍스트): A컬럼(desc텍스트)}.
    헤더행 3, 데이터 행 4~."""
    wb = load_workbook(quests_path, read_only=True, data_only=True)
    try:
        if "#quest" not in wb.sheetnames:
            return {}
        ws = wb["#quest"]
        templates: dict[str, str] = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            # A=desc, B=quest
            desc = row[0]
            quest = row[1]
            if quest is not None and desc is not None:
                templates[str(quest)] = str(desc)
        return templates
    finally:
        wb.close()


def load_dialog_groups(dialog_groups_path: str) -> list[dict]:
    """dialog_groups.xlsx -> [{"id": int, "actor_name": str, "dialog_text": str}].

    finish_town_dialog:dialog_group_id condition 의 param1 selectbox 미리보기용.
    같은 group_id 의 여러 dialog 중 가장 앞 순서(dialog_id 최소) 1건만 추출.

    파일 구조 (2026-04-24 조사):
      - 시트 'dialog_groups.dialogs' 사용
      - 헤더 row 3, 데이터 row 4~
      - B: ^id (group_id), E: dialog_id (순번), F: dialog_text
      - C: #actor_name (화자 표시용 — 주석 컬럼, 미리보기 용도로 사용)

    `actor_name` 헤더는 dialog_actors/0/, /1/, ... 마다 중복되어 base_prefix=""
    모드에서 collisions 발생. 따라서 화자 미리보기는 #actor_name 컬럼(C)을 사용.
    """
    wb = load_workbook(dialog_groups_path, read_only=True, data_only=True)
    try:
        sheet_name = "dialog_groups.dialogs"
        if sheet_name not in wb.sheetnames:
            # 폴백: 시트명이 다를 때 첫 번째 'dialogs' 포함 시트 사용
            candidates = [s for s in wb.sheetnames if "dialog" in s.lower() and s != "Description"]
            if not candidates:
                return []
            sheet_name = candidates[0]
        ws = wb[sheet_name]
        header_map = _build_header_map(ws, header_row=3, base_prefix="")

        id_col = header_map.get("^id")
        actor_col = header_map.get("#actor_name")  # 화자 표시용 주석 컬럼
        text_col = header_map.get("dialog_text")
        dialog_id_col = header_map.get("dialog_id")

        if id_col is None:
            return []

        # group_id -> (min_dialog_id, actor_name, dialog_text)
        first_by_group: dict[int, tuple[int, str, str]] = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            raw_id = row[id_col - 1]
            if raw_id is None:
                continue
            try:
                gid = int(raw_id)
            except (ValueError, TypeError):
                continue
            did_raw = row[dialog_id_col - 1] if dialog_id_col else None
            try:
                did = int(did_raw) if did_raw is not None else 10**9
            except (ValueError, TypeError):
                did = 10**9
            actor = row[actor_col - 1] if actor_col else None
            text = row[text_col - 1] if text_col else None

            prev = first_by_group.get(gid)
            if prev is None or did < prev[0]:
                first_by_group[gid] = (
                    did,
                    str(actor) if actor is not None else "",
                    str(text) if text is not None else "",
                )

        return [
            {"id": gid, "actor_name": a, "dialog_text": t}
            for gid, (_, a, t) in sorted(first_by_group.items())
        ]
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# parent key 제안 (TPL_A/B/C 대상)
# ---------------------------------------------------------------------------

# filter 별 parent key 간격 관례 (실 데이터 기반)
#   - LAUNCH_0 Day1~7:    step 50
#   - HELSINKI_3, ISTANBUL_1, ISTANBUL_3: step 100
#   - DAILY (reset_type 기반): step 10 (30001, 30016, 40001, ...)
FILTER_KEY_STEP: dict[str, int] = {
    "$$LAUNCH_0": 50,
    "$$HELSINKI_3": 100,
    "$$ISTANBUL_1": 100,
    "$$ISTANBUL_3": 100,
}
DEFAULT_PARENT_STEP = 10


def _keys_used_with_filter(
    existing_by_filter: dict[int, str] | None,
    filter_id: str | None,
) -> set[int]:
    """filter 별 사용된 ^key 집합. existing_by_filter 가 None 이면 빈 set."""
    if existing_by_filter is None or not filter_id:
        return set()
    return {k for k, f in existing_by_filter.items() if f == filter_id}


def suggest_next_parent_key(
    existing: set[int],
    filter_id: str | None,
    reset_type: str | None,
    category: str | None = "QUEST_CATEGORY_GENERAL",
    existing_by_filter: dict[int, str] | None = None,
) -> int:
    """filter + reset_type 조합 기준 '기존 max + step' 값을 제안 (filter-scoped).

    R-02 (2026-04-24 Round 2) 수정:
      - 기존: category+reset_type 범위 전체의 max → LAUNCH_0 NONE 제안값이 HELSINKI 구역(73755)
      - 신규: filter 로 사용된 key 만 필터링 후 max+step
      - existing_by_filter 미지원 환경(None)에서는 category/reset_type 범위 max 폴백

    Args:
        existing: quests.xlsx 전체 ^key 집합
        filter_id: "$$LAUNCH_0" 등 (None 이면 폴백)
        reset_type: QUEST_RESET_TYPE_* 중 하나 또는 None
        category: QUEST_CATEGORY_* (_get_key_range 로 범위 결정)
        existing_by_filter: {key: filter} 매핑 (app.py 에서 전달)

    Returns:
        제안 parent key. 관례 없는 조합은 category/reset_type 범위의 max+step.
    """
    # step 결정
    if reset_type == "QUEST_RESET_TYPE_DAILY":
        step = DEFAULT_PARENT_STEP  # 10
    elif filter_id and filter_id in FILTER_KEY_STEP:
        step = FILTER_KEY_STEP[filter_id]
    else:
        step = DEFAULT_PARENT_STEP

    range_lo, range_hi = _get_key_range(category, reset_type)

    # filter-scoped 후보 우선 사용 (filter 가 이미 scope 제한이므로 range 제약 완화)
    filter_keys = _keys_used_with_filter(existing_by_filter, filter_id)

    if filter_keys:
        # filter 로 쓰인 key 들의 max + step — range 제약 없이 filter 영역 내에서 연속 발급
        candidate = max(filter_keys) + step
    else:
        # filter 관례 없음 → category/reset_type 범위 내 max+step (기존 동작)
        in_range = [k for k in existing if range_lo <= k <= range_hi]
        if not in_range:
            return range_lo
        candidate = max(in_range) + step

    # 999999 상한 체크 (_KEY_MAX)
    if candidate > 999999:
        return generate_unique_key(existing, category=category, reset_type=reset_type)
    if candidate in existing:
        return generate_unique_key(existing, category=category, reset_type=reset_type)
    return candidate


def allocate_child_keys(
    existing: set[int],
    parent_key: int,
    n: int,
    category: str | None = None,
    reset_type: str | None = None,
) -> list[int]:
    """parent_key 기준 child ^key N개 발급 (parent+1, +2, ..., +N 순차).

    R-06 (2026-04-24 Round 2) 수정:
      - signature 확장: category/reset_type 으로 fallback 범위 제한
      - parent_key 범위 가드 (1 ≤ parent_key ≤ 999999) — 음수/0 ValueError
      - fallback 시 동일 range 유지 (100000대로 점프 방지)

    Args:
        existing: 전체 ^key 집합
        parent_key: 기준 parent key (양수 정수)
        n: 발급할 child 수 (0 이상)
        category: fallback 범위 결정용 (없으면 parent_key 로부터 +1000 까지 탐색)
        reset_type: fallback 범위 결정용

    Returns:
        길이 N 의 child key 리스트 (모두 existing 와 중복 없음)

    Raises:
        ValueError: parent_key 가 1~999999 범위 밖, 또는 n < 0
    """
    if n < 0:
        raise ValueError(f"n 은 0 이상 정수여야 함: {n}")
    if not isinstance(parent_key, int) or isinstance(parent_key, bool):
        raise ValueError(f"parent_key 는 정수여야 함: {parent_key!r}")
    if parent_key < 1 or parent_key > 999999:
        raise ValueError(f"parent_key 범위(1~999999) 벗어남: {parent_key}")
    if n == 0:
        return []

    taken = set(existing)
    result: list[int] = []
    cursor = parent_key + 1
    offset = 0
    while len(result) < n:
        candidate = cursor + offset
        # 상한 check (동일 category/reset_type 범위 내)
        if candidate > 999999:
            candidate = _fallback_in_range(taken, category, reset_type)
            result.append(candidate)
            taken.add(candidate)
            cursor = candidate + 1
            offset = 0
            continue
        if candidate not in taken:
            result.append(candidate)
            taken.add(candidate)
            cursor = candidate + 1
            offset = 0
        else:
            offset += 1
            if offset > 1000:
                # 1000회 연속 충돌 → 동일 range 내 빈 자리 탐색
                fallback = _fallback_in_range(taken, category, reset_type)
                result.append(fallback)
                taken.add(fallback)
                cursor = fallback + 1
                offset = 0
    return result


def _fallback_in_range(
    taken: set[int],
    category: str | None,
    reset_type: str | None,
) -> int:
    """allocate_child_keys fallback — 동일 category/reset_type 범위 내 빈 자리."""
    return generate_unique_key(taken, category=category, reset_type=reset_type)


# ---------------------------------------------------------------------------
# description 색상 프리셋 (실데이터에서 추출한 빈도 높은 문구)
# ---------------------------------------------------------------------------

DESC_PRESETS: list[str] = [
    "<color=#ffe535>{0}회 전투 승리</color>하기",
    "<color=#ffe535>{0}회 플레이</color>하기",
    "<color=#ffe535>상시 모드에서 {0}회 플레이</color>하기",
    "<color=#ffe535>로테이션 모드에서 {0}회 플레이</color>하기",
    "<color=#ffe535>특별 이벤트에서 {0}회 플레이</color>하기",
    "<color=#ffe535>{0}회 배틀박스 열기</color>",
    "<color=#ffe535>{0}회 코인박스 열기</color>",
    "<color=#ffe535>쿠키 뽑기 {0}회</color>하기",
    "<color=#ffe535>스펠카드 {0}회 사용</color>하기",
    "<color=#ffe535>오븐크라운 {0}개 획득</color>",
    "<color=#ffe535>{0}회 MVP 달성</color>",
]


# ---------------------------------------------------------------------------
# 쓰기 함수
# ---------------------------------------------------------------------------

def _verify_written_key(
    xlsx_path: str,
    key: Any,
    expected_row: int,
    header_map: dict[str, int],
) -> None:
    """저장 후 검증: 해당 행 ^key 값이 일치하는지 확인. 불일치 시 RuntimeError."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["quests"]
        key_col = header_map.get("^key")
        if key_col is None:
            raise RuntimeError("헤더에서 ^key 컬럼을 찾을 수 없습니다.")
        row = ws[expected_row]
        actual = row[key_col - 1].value
        if actual != key:
            raise RuntimeError(
                f"저장 검증 실패: Row {expected_row}의 ^key={actual!r}, 기대값={key!r}"
            )
    finally:
        wb.close()


def _atomic_save(wb, xlsx_path: str) -> None:
    """wb.save() 를 tmp + os.replace 로 감싸 atomic 저장.

    R-07 (2026-04-24 Round 2) 신설:
      - openpyxl wb.save() 는 in-place 덮어쓰기 → 저장 중 예외 시 파일 손상 가능
      - tmp 파일에 먼저 저장 후 os.replace 로 atomic rename
      - 예외 발생 시 원본 파일 보존 (자연 롤백)
    """
    xlsx_path = str(xlsx_path)
    target_dir = os.path.dirname(os.path.abspath(xlsx_path))
    # 동일 디렉토리에 tmp 생성 (다른 드라이브면 os.replace 실패 가능)
    fd, tmp_path = tempfile.mkstemp(
        prefix=".quest_writer_",
        suffix=".xlsx.tmp",
        dir=target_dir,
    )
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, xlsx_path)  # atomic on Windows + POSIX
    except Exception:
        # tmp 정리 (원본 유지)
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
        raise


def _coerce_key_to_int(value: Any, label: str = "^key") -> int | None:
    """^key 값을 int 로 강제 변환. bool 거부, 문자열 변환 실패 시 ValueError.

    R-10 (2026-04-24 Round 2):
      - 기존: try/except 후 pass (str 그대로 기입)
      - 신규: ValueError raise (validator 와 동일 정책)
    """
    if value is None:
        return None
    if isinstance(value, bool):
        raise ValueError(f"{label} 는 bool 타입 불가: {value!r}")
    if isinstance(value, int):
        return value
    try:
        return int(value)
    except (ValueError, TypeError) as e:
        raise ValueError(f"{label} 는 정수여야 함 (실패: {value!r}): {e}")


def append_quest_row(
    xlsx_path: str,
    field_values: dict[str, Any],
    *,
    validate: bool = True,
    items_path: str | None = None,
    keywords_path: str | None = None,
    dialog_groups_path: str | None = None,
) -> int:
    """
    quests 시트 마지막 행 다음에 새 행 추가.
    반환: 추가된 행 번호.

    안전장치:
    1. 파일 열기 직전 ^key 재수집 (세션 캐시 미신뢰)
    2. ^key 타입/중복 확인 (R-10 str → ValueError)
    3. validate=True 시 quest_validator.validate_quest_row() 호출
    4. max_row + 1 위치에만 삽입, _atomic_save() 로 원자 저장
    5. 저장 후 _verify_written_key()

    Args:
        xlsx_path: quests.xlsx 경로
        field_values: 행 데이터
        validate: True 면 quest_validator 호출 (기본). False 로 비활성화 가능.
        items_path/keywords_path/dialog_groups_path: validator FK refs 용. None 이면
          해당 FK 검증 skip.
    """
    # 안전장치 1: 파일 열기 직전 ^key 재수집
    existing_keys = get_existing_keys(xlsx_path)

    new_key = _coerce_key_to_int(field_values.get("^key"))
    if new_key is None:
        raise ValueError("^key 필수")
    if new_key in existing_keys:
        raise ValueError(f"^key {new_key!r} 는 이미 존재합니다.")

    # 안전장치 3: validator 호출
    if validate:
        from quest_validator import build_refs_from_paths, validate_quest_row

        refs = build_refs_from_paths(
            items_path=items_path,
            keywords_path=keywords_path,
            dialog_groups_path=dialog_groups_path,
            quests_path=xlsx_path,
        )
        vrow = dict(field_values)
        vrow["^key"] = new_key
        errors = validate_quest_row(vrow, refs, strict_count_type=False)
        if errors:
            raise ValueError("검증 실패: " + "; ".join(errors))

    # 쓰기 모드로 열기
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb["quests"]
        header_map = _build_header_map(ws, path_row=_PATH_ROW, header_row=_HEADER_ROW)

        target_row = ws.max_row + 1

        for col_name, value in field_values.items():
            col_idx = header_map.get(col_name)
            if col_idx is None:
                continue
            if value is None or value == "":
                continue
            # ^key 는 정수로 기입 (이미 coerce 완료)
            if col_name == "^key":
                value = new_key
            ws.cell(row=target_row, column=col_idx, value=value)

        _atomic_save(wb, xlsx_path)
    finally:
        wb.close()

    # 안전장치 5: 저장 후 검증
    header_map_verify = get_header_map(
        xlsx_path, sheet="quests", path_row=_PATH_ROW, header_row=_HEADER_ROW
    )
    _verify_written_key(xlsx_path, new_key, target_row, header_map_verify)

    return target_row


def append_daily_set(
    xlsx_path: str,
    parent: dict[str, Any],
    children: list[dict[str, Any]],
    *,
    validate: bool = True,
    items_path: str | None = None,
    keywords_path: str | None = None,
    dialog_groups_path: str | None = None,
) -> list[int]:
    """parent 1건 + children N건 일괄 기입 (단일 workbook open/save, atomic).

    자동 처리:
      1. parent `goal_type/type/%param1` = `[]{c1,c2,...}` (children ^key 기준)
      2. parent `goal_count` 이 None/0 이면 len(children) 으로 채움

    안전장치 (R-07, R-08, R-10 반영):
      - _atomic_save 로 tmp+replace 패턴 (save 중 예외 시 원본 파일 보존)
      - validate=True 시 quest_validator.validate_daily_set() 호출
        → parent 불변식 (HIGHEST, reward_quest:ref_quest_ids, goal_count == N) 강제
      - ^key 는 _coerce_key_to_int 로 bool/str → ValueError

    Args:
        xlsx_path: quests.xlsx 경로
        parent: 필드 값 dict. '^key' 필수.
        children: 필드 값 dict 목록. 각 '^key' 필수.
        validate: True 면 quest_validator 호출 (기본).
        items_path/keywords_path/dialog_groups_path: validator FK refs 용.

    Returns:
        [parent_row, child_row_1, ..., child_row_N]

    Raises:
        ValueError: ^key 타입/중복/범위 위반, parent 불변식 위반, FK 위반
    """
    if not children:
        raise ValueError("children 이 비어있습니다. parent 혼자 저장은 append_quest_row 사용.")

    parent_key = _coerce_key_to_int(parent.get("^key"), "parent ^key")
    if parent_key is None:
        raise ValueError("parent 에 ^key 가 없습니다.")

    child_keys: list[int] = []
    for i, ck in enumerate(children):
        raw = ck.get("^key")
        if raw is None:
            raise ValueError(f"children[{i}] 에 ^key 가 없습니다.")
        child_keys.append(_coerce_key_to_int(raw, f"children[{i}] ^key"))

    # 사전 검증 1: parent/child 간 중복
    all_new_keys = [parent_key] + child_keys
    if len(set(all_new_keys)) != len(all_new_keys):
        raise ValueError(f"parent/child 간 ^key 중복: {all_new_keys}")

    # 사전 검증 2: 기존 파일과의 중복
    existing_keys = get_existing_keys(xlsx_path)
    collisions = [k for k in all_new_keys if k in existing_keys]
    if collisions:
        raise ValueError(f"^key 중복 (파일 내 이미 존재): {collisions}")

    # parent goal_type/type/%param1 자동 생성
    parent = dict(parent)
    parent["^key"] = parent_key
    param1_val = "[]{" + ",".join(str(k) for k in child_keys) + "}"
    parent["goal_type/type/%param1"] = param1_val

    # parent goal_count 기본값 = child 개수 (None/0 만 자동 채움)
    # R-08: goal_count != len(children) 는 validator 에서 오류 → 여기서는 None/0 만 보정
    gc_raw = parent.get("goal_count")
    if gc_raw is None or gc_raw == 0 or gc_raw == "":
        parent["goal_count"] = len(children)

    # child ^key 정수화 (기입용)
    children_norm = []
    for c, ck_int in zip(children, child_keys):
        c_copy = dict(c)
        c_copy["^key"] = ck_int
        children_norm.append(c_copy)

    # 안전장치 3: validator 호출
    if validate:
        from quest_validator import build_refs_from_paths, validate_daily_set

        refs = build_refs_from_paths(
            items_path=items_path,
            keywords_path=keywords_path,
            dialog_groups_path=dialog_groups_path,
            quests_path=xlsx_path,
        )
        errors = validate_daily_set(parent, children_norm, refs)
        if errors:
            raise ValueError("검증 실패: " + "; ".join(errors))

    # 쓰기 모드로 열기 (1회만)
    wb = load_workbook(xlsx_path, data_only=True)
    written_rows: list[int] = []
    try:
        ws = wb["quests"]
        header_map = _build_header_map(ws, path_row=_PATH_ROW, header_row=_HEADER_ROW)
        next_row = ws.max_row + 1

        # parent 기입
        for col_name, value in parent.items():
            col_idx = header_map.get(col_name)
            if col_idx is None or value is None or value == "":
                continue
            ws.cell(row=next_row, column=col_idx, value=value)
        written_rows.append(next_row)
        next_row += 1

        # children 기입
        for child in children_norm:
            for col_name, value in child.items():
                col_idx = header_map.get(col_name)
                if col_idx is None or value is None or value == "":
                    continue
                ws.cell(row=next_row, column=col_idx, value=value)
            written_rows.append(next_row)
            next_row += 1

        _atomic_save(wb, xlsx_path)
    finally:
        wb.close()

    # 저장 후 검증
    header_map_verify = get_header_map(
        xlsx_path, sheet="quests", path_row=_PATH_ROW, header_row=_HEADER_ROW
    )
    for key, row_num in zip(all_new_keys, written_rows):
        _verify_written_key(xlsx_path, key, row_num, header_map_verify)

    return written_rows


def parse_quest_texts(pasted_text: str, quest_templates: dict[str, str]) -> list[dict]:
    """붙여넣은 텍스트 -> 배치 행 초안 목록.

    Args:
        pasted_text: 한 줄에 하나씩 quest 텍스트 붙여넣기
        quest_templates: {B컬럼(quest텍스트): A컬럼(desc텍스트)} from load_quest_templates()

    Returns:
        list of {
            description: str (B컬럼값, {0} 그대로 보존),
            town_description: str (A컬럼값, 매칭되면 채움 아니면 빈 문자열),
            matched: bool (템플릿 매칭 성공 여부),
            goal_count: int (기본 1),
            goal_type_key: str (기본 빈 문자열),
            goal_type_param1: str (기본 빈 문자열),
            reward_id: str (기본 빈 문자열),
            qty: int (기본 1),
            delete: bool (기본 False),
        }
    """
    rows = []
    for line in pasted_text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        town_desc = quest_templates.get(line, "")
        rows.append({
            "description": line,
            "town_description": town_desc,
            "matched": bool(town_desc),
            "goal_count": 1,
            "goal_type_key": "",
            "goal_type_param1": "",
            "reward_id": "",
            "qty": 1,
            "delete": False,
        })
    return rows


# ---------------------------------------------------------------------------
# GoalType YAML 로더 (설정 editable)
# ---------------------------------------------------------------------------

def _default_goal_types_path() -> str:
    """quest_tool/goal_types.yaml 기본 경로."""
    return str(Path(__file__).parent / "goal_types.yaml")


def load_goal_types_yaml(yaml_path: str | None = None) -> list[dict]:
    """goal_types.yaml → list[dict] (GOAL_TYPES 포맷).
    파일 부존재/파싱 실패 시 _HARDCODED_GOAL_TYPES 폴백.
    YAML 포맷:
        goal_types:
          - key: ...
            label: ...
            params:
              - label: ...
                options: [...]
                free_text: true|false
                item_picker: true|false
                dialog_picker: true|false
                option_labels: {"100": "상시", ...}
    """
    try:
        import yaml  # type: ignore
    except ImportError:
        return list(_HARDCODED_GOAL_TYPES)

    path = yaml_path or _default_goal_types_path()
    if not os.path.exists(path):
        return list(_HARDCODED_GOAL_TYPES)

    try:
        with open(path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
    except Exception:
        return list(_HARDCODED_GOAL_TYPES)

    if not isinstance(data, dict):
        return list(_HARDCODED_GOAL_TYPES)

    items = data.get("goal_types")
    if not isinstance(items, list) or not items:
        return list(_HARDCODED_GOAL_TYPES)

    result: list[dict] = []
    for entry in items:
        if not isinstance(entry, dict):
            continue
        key = entry.get("key")
        if not key or not isinstance(key, str):
            continue
        label = entry.get("label") or key
        params = entry.get("params") or []
        if not isinstance(params, list):
            params = []
        # params 검증: 각 항목은 dict 이고 최소 label 필드 보유
        clean_params: list[dict] = []
        for p in params:
            if not isinstance(p, dict):
                continue
            if "label" not in p:
                continue
            clean_params.append(p)
        result.append({"key": key, "label": str(label), "params": clean_params})
    return result if result else list(_HARDCODED_GOAL_TYPES)


def save_goal_types_yaml(goal_types: list[dict], yaml_path: str | None = None) -> None:
    """goal_types.yaml atomic 저장.
    입력 포맷은 load_goal_types_yaml() 반환과 동일.
    tmp + os.replace 로 기록."""
    try:
        import yaml  # type: ignore
    except ImportError as e:
        raise RuntimeError("PyYAML 이 설치되어 있지 않습니다") from e

    path = yaml_path or _default_goal_types_path()

    # 검증: 각 entry 는 key/label/params 필수
    clean: list[dict] = []
    for entry in goal_types:
        if not isinstance(entry, dict):
            continue
        key = entry.get("key")
        if not key or not isinstance(key, str):
            continue
        label = entry.get("label") or key
        params = entry.get("params") or []
        clean.append({"key": key, "label": str(label), "params": params})

    payload = {"goal_types": clean}
    data_dir = os.path.dirname(path) or "."
    os.makedirs(data_dir, exist_ok=True)

    fd, tmp_path = tempfile.mkstemp(
        prefix=".goal_types_tmp_", suffix=".yaml", dir=data_dir
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            yaml.safe_dump(
                payload, f, allow_unicode=True, sort_keys=False,
                default_flow_style=False,
            )
        os.replace(tmp_path, path)
    except Exception:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        raise


# ---------------------------------------------------------------------------
# Daily Mission 헬퍼
# ---------------------------------------------------------------------------

_DAY_KEYWORD_RE = re.compile(r"NDAY(\d+)", re.IGNORECASE)


def extract_day_from_keyword(timestamp: str) -> int | None:
    """start_timestamp 문자열에서 N일차 추출.
    예:
      "$$LAUNCH_0_NDAY1" → 1
      "$$NDAY7"          → 7
      "$$INDEFINITE_TIMESTAMP" → None
      ""                 → None
    """
    if not timestamp or not isinstance(timestamp, str):
        return None
    m = _DAY_KEYWORD_RE.search(timestamp)
    if not m:
        return None
    try:
        return int(m.group(1))
    except (ValueError, TypeError):
        return None


def default_parent_desc(start_timestamp: str, existing_parent_count: int = 0) -> str:
    """데일리 미션 parent description 자동 생성.
    실데이터 패턴: `[데일리미션]N일차 전체 퀘스트 완료 보상`
    start_timestamp 에서 NDAY 추출 실패 시 existing_parent_count + 1 폴백.
    """
    day = extract_day_from_keyword(start_timestamp or "")
    if day is None:
        day = max(1, existing_parent_count + 1)
    return f"[데일리미션]{day}일차 전체 퀘스트 완료 보상"


# ---------------------------------------------------------------------------
# nday_mission_events.xlsx 함수들
# ---------------------------------------------------------------------------

def get_existing_event_keys(xlsx_path: str) -> set[int]:
    """nday_mission_events.xlsx events 시트의 기존 ^key 집합 반환.

    events 시트 구조: Row1=설명, Row2=경로, Row3=헤더(^key, description, ...), Row4~=데이터
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["events"]
        header_map = _build_header_map(ws, path_row=2, header_row=3, base_prefix="")
        key_col = header_map.get("^key")
        if key_col is None:
            return set()
        existing: set[int] = set()
        for row in ws.iter_rows(min_row=4, values_only=True):
            val = row[key_col - 1]
            if val is not None:
                try:
                    existing.add(int(val))
                except (ValueError, TypeError):
                    pass
        return existing
    finally:
        wb.close()


def suggest_next_event_key(existing_keys: set[int], group_base: int) -> int:
    """group_base+1 ~ group_base+99 범위에서 사용 안 된 최소 int 반환.

    예: group_base=100, existing_keys={101} → 102 반환
        group_base=300, existing_keys={} → 301 반환

    Raises:
        ValueError: 범위 내 공간 없음 (group_base+1 ~ group_base+99 전부 사용 중)
    """
    lo = group_base + 1
    hi = group_base + 99
    for candidate in range(lo, hi + 1):
        if candidate not in existing_keys:
            return candidate
    raise ValueError(
        f"이벤트 ^key 범위 {lo}~{hi} 가 꽉 찼습니다. (group_base={group_base})"
    )


def load_nday_mission_events(xlsx_path: str) -> list[dict]:
    """nday_mission_events.xlsx events 시트 → [{"key": int, "description": str, ...}].

    사이드바 Load Status 표시용. events 시트만 읽음.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["events"]
        header_map = _build_header_map(ws, path_row=2, header_row=3, base_prefix="")
        key_col = header_map.get("^key")
        desc_col = header_map.get("description")
        if key_col is None:
            return []
        events: list[dict] = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            val = row[key_col - 1]
            if val is None:
                continue
            try:
                k = int(val)
            except (ValueError, TypeError):
                continue
            desc = row[desc_col - 1] if desc_col else None
            events.append({
                "key": k,
                "description": str(desc) if desc is not None else "",
            })
        return events
    finally:
        wb.close()


def append_nday_mission_event(
    xlsx_path: str,
    event_dict: dict,
    event_day_dict: dict,
) -> tuple[int, int]:
    """events 시트 + events.day 시트에 각 1행씩 추가. 단일 workbook open → _atomic_save.

    Args:
        xlsx_path: nday_mission_events.xlsx 경로
        event_dict: {"^key": int, "description": str, "start_timestamp": str,
                     "end_timestamp": str, "mission_active_days": int}
        event_day_dict: {"^key": int, "day": 1, "description": str,
                         "quest_ids": "[]{...}", "finish_quest_id": int}

    Returns:
        (events_row, events_day_row) — 추가된 행 번호 tuple

    Raises:
        ValueError: events.^key 중복 시
    """
    event_key = _coerce_key_to_int(event_dict.get("^key"), "events ^key")
    if event_key is None:
        raise ValueError("events ^key 필수")

    # 중복 체크
    existing = get_existing_event_keys(xlsx_path)
    if event_key in existing:
        raise ValueError(f"events ^key {event_key!r} 는 이미 존재합니다.")

    wb = load_workbook(xlsx_path, data_only=True)
    try:
        # events 시트 기입
        ws_ev = wb["events"]
        hm_ev = _build_header_map(ws_ev, path_row=2, header_row=3, base_prefix="")
        ev_row = ws_ev.max_row + 1
        for col_name, value in event_dict.items():
            col_idx = hm_ev.get(col_name)
            if col_idx is None or value is None or value == "":
                continue
            ws_ev.cell(row=ev_row, column=col_idx, value=value)

        # events.day 시트 기입
        ws_day = wb["events.day"]
        hm_day = _build_header_map(ws_day, path_row=2, header_row=3, base_prefix="")
        day_row = ws_day.max_row + 1
        for col_name, value in event_day_dict.items():
            col_idx = hm_day.get(col_name)
            if col_idx is None or value is None or value == "":
                continue
            ws_day.cell(row=day_row, column=col_idx, value=value)

        _atomic_save(wb, xlsx_path)
    finally:
        wb.close()

    return ev_row, day_row
