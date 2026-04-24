"""Capture daily mission baseline for STEP D7.5 reproduction test.

실 quests.xlsx 에서 ^key 40001 / 40011 의 parent + child 전체를 필드 단위로 캡처하여
`tests/fixtures/daily_mission_baseline.json` 으로 저장한다.

Usage:
    python scripts/capture_daily_mission_baseline.py

Target rows:
    parent 40001 + children 40002~40006
    parent 40011 + children 40012~40016

각 행은 `{^key, fields: {header: value, ...}}` 형태로 저장되며, 값은 셀 타입 그대로
(int/str/float/None) 보존한다. JSON 직렬화를 위해 None 은 null 로 변환된다.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

# 실 데이터 경로 — 필요 시 조정
DEFAULT_QUESTS_PATH = r"D:\COS_Project\cos-data\excel\quests.xlsx"

# 캡처 대상 parent ^key
TARGET_PARENTS = [40001, 40011]

# baseline 저장 경로
REPO_ROOT = Path(__file__).resolve().parent.parent
BASELINE_PATH = REPO_ROOT / "tests" / "fixtures" / "daily_mission_baseline.json"


def _build_header_map(ws) -> dict[int, str]:
    """Row 2 (path) + Row 3 (header) 복합키 → 컬럼 index(1-based) 매핑."""
    header_map: dict[int, str] = {}
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        row2 = ws.cell(row=2, column=col).value
        row3 = ws.cell(row=3, column=col).value
        # 복합키: row2 가 있으면 row2 우선, 없으면 row3
        key = str(row2).strip() if row2 else (str(row3).strip() if row3 else "")
        if key:
            header_map[col] = key
    return header_map


def _capture_row(ws, row_idx: int, header_map: dict[int, str]) -> dict:
    """한 행을 필드 단위로 캡처."""
    fields: dict[str, object] = {}
    for col, header in header_map.items():
        val = ws.cell(row=row_idx, column=col).value
        fields[header] = val
    return {"row": row_idx, "fields": fields}


def _find_parent_and_children(ws, header_map: dict[int, str], parent_key: int) -> list[dict]:
    """parent ^key 로 row 찾고, goal_type/%param1 의 child ^key 들로 child 행 추출."""
    # ^key 컬럼 찾기
    key_col = None
    for col, h in header_map.items():
        if h == "^key":
            key_col = col
            break
    if key_col is None:
        raise RuntimeError("^key 컬럼을 찾을 수 없음")

    # goal_type/%param1 컬럼 찾기
    param1_col = None
    for col, h in header_map.items():
        if h in ("goal_type/type/%param1", "goal_type/%param1"):
            param1_col = col
            break

    # parent 행 스캔
    parent_row_idx = None
    for r in range(4, ws.max_row + 1):
        val = ws.cell(row=r, column=key_col).value
        if val == parent_key:
            parent_row_idx = r
            break
    if parent_row_idx is None:
        raise RuntimeError(f"parent ^key={parent_key} 를 찾을 수 없음")

    # child ^key 파싱 — parent goal_type/%param1 = "[]{40002,40003,...}"
    child_keys: list[int] = []
    if param1_col is not None:
        param1_val = ws.cell(row=parent_row_idx, column=param1_col).value
        if isinstance(param1_val, str) and param1_val.startswith("[]{") and param1_val.endswith("}"):
            inner = param1_val[3:-1]
            for tok in inner.split(","):
                tok = tok.strip()
                if tok.isdigit():
                    child_keys.append(int(tok))

    # parent 행 캡처
    captured = [_capture_row(ws, parent_row_idx, header_map)]

    # child 행 스캔
    child_rows: dict[int, int] = {}
    for r in range(4, ws.max_row + 1):
        val = ws.cell(row=r, column=key_col).value
        if isinstance(val, int) and val in child_keys:
            child_rows[val] = r

    # child_keys 순서대로 캡처
    for ck in child_keys:
        if ck in child_rows:
            captured.append(_capture_row(ws, child_rows[ck], header_map))
        else:
            print(f"  ⚠ child ^key={ck} 를 찾을 수 없음 (건너뜀)", file=sys.stderr)

    return captured


def main() -> int:
    quests_path = DEFAULT_QUESTS_PATH
    if len(sys.argv) > 1:
        quests_path = sys.argv[1]

    if not Path(quests_path).exists():
        print(f"ERROR: quests.xlsx 파일 없음: {quests_path}", file=sys.stderr)
        return 1

    print(f"Loading: {quests_path}")
    wb = load_workbook(quests_path, data_only=True)
    if "quests" not in wb.sheetnames:
        print(f"ERROR: 'quests' 시트 없음 (시트 목록: {wb.sheetnames})", file=sys.stderr)
        return 1

    ws = wb["quests"]
    header_map = _build_header_map(ws)
    print(f"  headers: {len(header_map)} columns")

    baseline: dict = {"source": str(quests_path), "parents": []}
    for pk in TARGET_PARENTS:
        print(f"\nCapturing parent ^key={pk}...")
        rows = _find_parent_and_children(ws, header_map, pk)
        print(f"  captured {len(rows)} rows (parent + {len(rows) - 1} children)")
        baseline["parents"].append({"parent_key": pk, "rows": rows})

    # JSON 저장 — None 은 null, int/str 은 원본 타입 유지
    BASELINE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(BASELINE_PATH, "w", encoding="utf-8") as f:
        json.dump(baseline, f, ensure_ascii=False, indent=2, default=str)

    total_rows = sum(len(p["rows"]) for p in baseline["parents"])
    print(f"\n✓ Saved baseline to {BASELINE_PATH}")
    print(f"  total rows: {total_rows}")
    print(f"  total fields per row: {len(header_map)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
