"""
create_fixtures.py -- 테스트용 픽스처 Excel 파일 생성.
실 데이터에서 지정된 행만 발췌하여 tests/fixtures/ 에 저장.
"""
from pathlib import Path

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

SRC_DIR = Path("D:/COS_Project/cos-data/excel")
DST_DIR = Path(__file__).parent / "fixtures"
DST_DIR.mkdir(parents=True, exist_ok=True)


def copy_rows(src_ws, dst_ws, row_numbers: list[int]) -> None:
    """지정 행 번호 목록을 순서대로 dst_ws에 append."""
    for rn in row_numbers:
        row_data = []
        for cell in src_ws[rn]:
            row_data.append(cell.value)
        dst_ws.append(row_data)


def create_quests_fixture() -> None:
    src_path = SRC_DIR / "quests.xlsx"
    dst_path = DST_DIR / "quests_test.xlsx"

    wb_src = load_workbook(src_path, read_only=True, data_only=True)
    wb_dst = Workbook()
    # 기본 시트 제거
    wb_dst.remove(wb_dst.active)

    # --- quests 시트 ---
    ws_src = wb_src["quests"]
    ws_dst = wb_dst.create_sheet("quests")
    # Row 1~3 (메타/헤더) + Row 4, 5 (GENERAL) + Row 511, 512 (TOWN)
    copy_rows(ws_src, ws_dst, [1, 2, 3, 4, 5, 511, 512])

    # --- #quest 시트 ---
    ws_src2 = wb_src["#quest"]
    ws_dst2 = wb_dst.create_sheet("#quest")
    # Row 1~3 (메타/헤더) + Row 3~7 (5개 템플릿: 실제 데이터 Row 3~7)
    # 스펙: Row 1~3 헤더 + Row 4~8 (5개 템플릿) → 총 8행
    copy_rows(ws_src2, ws_dst2, [1, 2, 3, 4, 5, 6, 7, 8])

    wb_src.close()
    wb_dst.save(dst_path)
    print(f"[OK] quests_test.xlsx -> {dst_path} (quests: 7행, #quest: 8행)")


def create_items_fixture() -> None:
    src_path = SRC_DIR / "items.xlsx"
    dst_path = DST_DIR / "items_test.xlsx"

    wb_src = load_workbook(src_path, read_only=True, data_only=True)
    # 첫 번째 시트 사용
    sheet_name = wb_src.sheetnames[0]
    ws_src = wb_src[sheet_name]

    wb_dst = Workbook()
    wb_dst.remove(wb_dst.active)
    ws_dst = wb_dst.create_sheet(sheet_name)

    # 헤더 3행 + 지정 데이터 행
    data_rows = [4, 7, 76, 77, 112, 124, 331, 332, 447, 518, 519, 520,
                 1042, 1043, 1044, 1442, 1443, 1444, 1837]
    copy_rows(ws_src, ws_dst, [1, 2, 3] + data_rows)

    wb_src.close()
    wb_dst.save(dst_path)
    total = 3 + len(data_rows)
    print(f"[OK] items_test.xlsx -> {dst_path} ({total}행, 시트: {sheet_name})")


def create_keywords_fixture() -> None:
    src_path = SRC_DIR / "keywords.xlsx"
    dst_path = DST_DIR / "keywords_test.xlsx"

    wb_src = load_workbook(src_path, read_only=True, data_only=True)
    wb_dst = Workbook()
    wb_dst.remove(wb_dst.active)

    for sheet_name in ["build", "timestamp"]:
        if sheet_name not in wb_src.sheetnames:
            print(f"[WARN] {sheet_name} 시트 없음 — 스킵")
            continue
        ws_src = wb_src[sheet_name]
        ws_dst = wb_dst.create_sheet(sheet_name)
        # 전체 복사
        count = 0
        for row in ws_src.iter_rows(values_only=True):
            ws_dst.append(list(row))
            count += 1
        print(f"[OK] keywords_test.xlsx [{sheet_name}] -> {count}행 복사")

    wb_src.close()
    wb_dst.save(dst_path)
    print(f"[OK] keywords_test.xlsx -> {dst_path}")


def create_dialog_groups_fixture() -> None:
    """dialog_groups.xlsx 에서 ~10 group 발췌. 각 group 당 처음 몇 행만 포함."""
    src_path = SRC_DIR / "dialog_groups.xlsx"
    dst_path = DST_DIR / "dialog_groups_test.xlsx"

    if not src_path.exists():
        print(f"[WARN] dialog_groups.xlsx 원본 없음: {src_path} — 스킵")
        return

    wb_src = load_workbook(src_path, read_only=True, data_only=True)
    wb_dst = Workbook()
    wb_dst.remove(wb_dst.active)

    sheet_name = "dialog_groups.dialogs"
    if sheet_name not in wb_src.sheetnames:
        print(f"[WARN] {sheet_name} 시트 없음 — 첫 번째 dialog 시트 사용")
        candidates = [s for s in wb_src.sheetnames if "dialog" in s.lower() and s != "Description"]
        sheet_name = candidates[0] if candidates else wb_src.sheetnames[0]

    ws_src = wb_src[sheet_name]
    ws_dst = wb_dst.create_sheet(sheet_name)

    # Row 1~3 (메타/헤더) + 각 group 의 시작 몇 행
    # col B = ^id (group_id), col E = dialog_id
    data_rows: list[int] = []
    seen_groups: dict[int, int] = {}  # group_id -> count
    for rn in range(4, min(ws_src.max_row, 200) + 1):
        row = ws_src[rn]
        gid = row[1].value  # col B
        if gid is None:
            continue
        try:
            gid_int = int(gid)
        except (ValueError, TypeError):
            continue
        count = seen_groups.get(gid_int, 0)
        if len(seen_groups) >= 8 and gid_int not in seen_groups:
            continue  # 최대 8 group 까지만
        if count < 3:  # 그룹당 최대 3행
            data_rows.append(rn)
            seen_groups[gid_int] = count + 1
        if len(seen_groups) >= 8 and all(v >= 3 for v in seen_groups.values()):
            break

    copy_rows(ws_src, ws_dst, [1, 2, 3] + data_rows)
    wb_src.close()
    wb_dst.save(dst_path)
    print(f"[OK] dialog_groups_test.xlsx -> {dst_path} ({3 + len(data_rows)}행, group {len(seen_groups)}건)")


if __name__ == "__main__":
    print("=== 픽스처 생성 시작 ===")
    create_quests_fixture()
    create_items_fixture()
    create_keywords_fixture()
    create_dialog_groups_fixture()
    print("=== 픽스처 생성 완료 ===")
    # 생성 결과 파일 크기 출력
    for name in [
        "quests_test.xlsx",
        "items_test.xlsx",
        "keywords_test.xlsx",
        "dialog_groups_test.xlsx",
    ]:
        p = DST_DIR / name
        if p.exists():
            print(f"  {name}: {p.stat().st_size:,} bytes")
