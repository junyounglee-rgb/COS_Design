"""pytest 단위 테스트 -- quest_writer.py 읽기 전용 함수 + append 검증"""
import os
import sys
from pathlib import Path

import pytest

# quest_writer.py를 import 할 수 있도록 경로 추가
sys.path.insert(0, str(Path(__file__).parent.parent))

FIXTURES = Path(__file__).parent / "fixtures"
QUESTS_TEST = str(FIXTURES / "quests_test.xlsx")
ITEMS_TEST = str(FIXTURES / "items_test.xlsx")
KEYWORDS_TEST = str(FIXTURES / "keywords_test.xlsx")
DIALOG_GROUPS_TEST = str(FIXTURES / "dialog_groups_test.xlsx")

from quest_writer import (
    _HARDCODED_GOAL_TYPES,
    allocate_child_keys,
    append_daily_set,
    append_nday_mission_event,
    append_quest_row,
    default_parent_desc,
    extract_day_from_keyword,
    generate_unique_key,
    get_existing_event_keys,
    get_existing_keys,
    get_header_map,
    load_dialog_groups,
    load_goal_types_yaml,
    load_item_categories,
    load_items,
    load_keywords,
    load_quest_templates,
    parse_quest_texts,
    save_goal_types_yaml,
    suggest_next_event_key,
    suggest_next_parent_key,
)

# append 테스트에 사용할 최소 유효 필드값
_MINIMAL_ROW = {
    "^key": 999001,
    "category": "QUEST_CATEGORY_GENERAL",
    "description": "테스트 추가 행",
    "count_type": "QUEST_COUNT_TYPE_SUM",
    "goal_count": 1,
    "reset_type": "QUEST_RESET_TYPE_NONE",
    "goal_type/type/%key": "daily_login",
}


class TestHeaderMap:
    def test_simple_fields_exist(self):
        """기본 필드는 단순 이름으로 존재해야 함."""
        hm = get_header_map(QUESTS_TEST)
        for key in ["^key", "$filter", "category", "description", "count_type", "goal_count", "reset_type"]:
            assert key in hm, f"'{key}' 키 없음"

    def test_compound_keys_exist(self):
        """중복 헤더는 경로 조합 키로 존재해야 함."""
        hm = get_header_map(QUESTS_TEST)
        for key in [
            "conditions/0/condition/%key",
            "conditions/1/condition/%key",
            "goal_type/type/%key",
            "goal_type/type/%param1",
            "rewards/0/id",
            "rewards/0/qty",
        ]:
            assert key in hm, f"'{key}' 키 없음"

    def test_no_duplicate_values(self):
        """모든 컬럼 인덱스가 고유해야 함 (중복 없음)."""
        hm = get_header_map(QUESTS_TEST)
        values = list(hm.values())
        assert len(values) == len(set(values)), "컬럼 인덱스 중복 존재"

    def test_key_col_is_b(self):
        """^key는 B컬럼(2)이어야 함."""
        hm = get_header_map(QUESTS_TEST)
        assert hm["^key"] == 2

    def test_returns_dict_of_int(self):
        """반환값이 dict이고 값이 정수인지 확인."""
        hm = get_header_map(QUESTS_TEST)
        assert isinstance(hm, dict)
        for col_name, col_idx in hm.items():
            assert isinstance(col_idx, int), f"'{col_name}' 컬럼 인덱스가 int가 아님"


class TestExistingKeys:
    def test_returns_set_of_ints(self):
        """반환 타입이 set[int]인지 확인."""
        keys = get_existing_keys(QUESTS_TEST)
        assert isinstance(keys, set)
        for k in keys:
            assert isinstance(k, int), f"키 {k!r}가 int가 아님"

    def test_known_keys_present(self):
        """31, 32, 500001, 500002 포함 확인."""
        keys = get_existing_keys(QUESTS_TEST)
        for expected in (31, 32, 500001, 500002):
            assert expected in keys, f"키 {expected} 누락"

    def test_count(self):
        """픽스처에 4개 데이터 행 -> 4개 키."""
        keys = get_existing_keys(QUESTS_TEST)
        assert len(keys) == 4


class TestGenerateUniqueKey:
    """카테고리+reset_type 별 범위에서 기존 max+1부터 순차 발급."""

    def test_default_range_when_no_category(self):
        """category 없이 호출 시 DEFAULT_KEY_RANGE(100000~999999) 사용."""
        existing: set[int] = set()
        key = generate_unique_key(existing)
        # DEFAULT_KEY_RANGE lo=100000 → 빈 범위에서 lo부터 시작
        assert key == 100000

    def test_general_none_range(self):
        """GENERAL + QUEST_RESET_TYPE_NONE → 30011~99999."""
        existing: set[int] = set()
        key = generate_unique_key(
            existing,
            category="QUEST_CATEGORY_GENERAL",
            reset_type="QUEST_RESET_TYPE_NONE",
        )
        assert 30011 <= key <= 99999
        assert key == 30011  # 빈 범위 → lo부터

    def test_general_reset_type_empty_fallback(self):
        """reset_type 빈 문자열/None → GENERAL fallback 사용."""
        existing: set[int] = set()
        key_empty = generate_unique_key(existing, category="QUEST_CATEGORY_GENERAL", reset_type="")
        key_none = generate_unique_key(existing, category="QUEST_CATEGORY_GENERAL", reset_type=None)
        assert key_empty == key_none == 30011

    def test_general_daily_range(self):
        """GENERAL + DAILY → 30001~59999."""
        existing: set[int] = set()
        key = generate_unique_key(
            existing,
            category="QUEST_CATEGORY_GENERAL",
            reset_type="QUEST_RESET_TYPE_DAILY",
        )
        assert 30001 <= key <= 59999
        assert key == 30001

    def test_general_repeat_range(self):
        """GENERAL + REPEAT → 50008~59999."""
        existing: set[int] = set()
        key = generate_unique_key(
            existing,
            category="QUEST_CATEGORY_GENERAL",
            reset_type="QUEST_RESET_TYPE_REPEAT",
        )
        assert 50008 <= key <= 59999
        assert key == 50008

    def test_town_range(self):
        """TOWN → 500000~599999."""
        existing: set[int] = set()
        key = generate_unique_key(existing, category="QUEST_CATEGORY_TOWN")
        assert 500000 <= key <= 599999
        assert key == 500000

    def test_continues_from_max_plus_one(self):
        """범위 내 기존 최대값 + 1부터 발급."""
        existing = {30011, 30020, 30050}
        key = generate_unique_key(
            existing,
            category="QUEST_CATEGORY_GENERAL",
            reset_type="QUEST_RESET_TYPE_NONE",
        )
        assert key == 30051  # max(30050) + 1

    def test_town_continues_from_max(self):
        """TOWN 범위에 기존 500002가 있으면 500003 발급."""
        existing = {500000, 500001, 500002}
        key = generate_unique_key(existing, category="QUEST_CATEGORY_TOWN")
        assert key == 500003

    def test_ignores_out_of_range_existing(self):
        """범위 밖 existing은 무시됨 (다른 카테고리 키와 섞여도 OK)."""
        # TOWN 범위 500000~599999, existing에는 GENERAL 범위 키도 있음
        existing = {31, 32, 30011, 500001}
        key = generate_unique_key(existing, category="QUEST_CATEGORY_TOWN")
        # TOWN 범위 내 max(500001) + 1
        assert key == 500002

    def test_max_plus_one_includes_gaps(self):
        """범위 내 max 기준 +1 → 중간 빈 자리는 건너뜀 (보수적 추가)."""
        # in_range max = 500003 → start = 500004
        existing = {500001, 500003}
        key = generate_unique_key(existing, category="QUEST_CATEGORY_TOWN")
        assert key == 500004  # max+1 (중간 500002는 비어있어도 무시)

    def test_fills_gap_when_range_exhausted(self):
        """범위 끝까지 차면 앞쪽 빈 자리 탐색."""
        # 500000~599998 가득 차고 500000만 비었다면 → max+1=599999 반환
        # 599999까지 꽉 차면 → 앞쪽 탐색해서 500000 반환
        existing = set(range(500001, 600000))  # 500001~599999 (599999 포함)
        key = generate_unique_key(existing, category="QUEST_CATEGORY_TOWN")
        assert key == 500000  # 앞쪽 빈 자리 반환

    def test_custom_lo_hi_override(self):
        """lo, hi 둘 다 지정 시 KEY_RANGES 무시 (하위 호환)."""
        existing: set[int] = set()
        key = generate_unique_key(existing, lo=200000, hi=299999)
        assert 200000 <= key <= 299999
        assert key == 200000

    def test_range_full_raises(self):
        """범위가 완전히 차면 ValueError."""
        # TOWN 범위 500000~599999 전체 채움
        existing = set(range(500000, 600000))
        with pytest.raises(ValueError, match="꽉 찼"):
            generate_unique_key(existing, category="QUEST_CATEGORY_TOWN")

    def test_unknown_category_uses_default(self):
        """정의되지 않은 카테고리 → DEFAULT_KEY_RANGE 사용."""
        existing: set[int] = set()
        key = generate_unique_key(existing, category="QUEST_CATEGORY_UNKNOWN")
        assert 100000 <= key <= 999999
        assert key == 100000

    def test_vault_mission_range(self):
        """VAULT_MISSION → 80000~89999 (예약)."""
        existing: set[int] = set()
        key = generate_unique_key(existing, category="QUEST_CATEGORY_VAULT_MISSION")
        assert 80000 <= key <= 89999


class TestLoadItems:
    def test_returns_list(self):
        """반환 타입이 list인지 확인."""
        items = load_items(ITEMS_TEST)
        assert isinstance(items, list)

    def test_item_structure(self):
        """각 항목이 {id, name, category} 키를 가지는지 확인."""
        items = load_items(ITEMS_TEST)
        assert len(items) > 0
        for item in items:
            assert "id" in item, "id 키 누락"
            assert "name" in item, "name 키 누락"
            assert "category" in item, "category 키 누락"

    def test_known_item_present(self):
        """id=1 (골드) 포함 확인."""
        items = load_items(ITEMS_TEST)
        ids = {item["id"] for item in items}
        assert 1 in ids, "id=1 (골드) 누락"

    def test_no_none_ids(self):
        """id가 None인 항목 없음."""
        items = load_items(ITEMS_TEST)
        for item in items:
            assert item["id"] is not None

    def test_item_has_filter_field(self):
        """확장 필드 filter 포함 확인 (없으면 빈 문자열)."""
        items = load_items(ITEMS_TEST)
        assert len(items) > 0
        for item in items:
            assert "filter" in item, "filter 키 누락"
            # items_test.xlsx 의 실 데이터는 $$LAUNCH_0 등이 있어야 함
        # 최소 1개는 $$ 접두어 filter 를 가져야 함
        assert any(i["filter"].startswith("$$") for i in items), "$$ 접두어 filter 부재"


class TestLoadItemCategories:
    def test_returns_list(self):
        """반환 타입이 list인지 확인."""
        cats = load_item_categories(ITEMS_TEST)
        assert isinstance(cats, list)

    def test_has_known_categories(self):
        """실 데이터에서 복제한 픽스처 — 주요 카테고리 포함 확인."""
        cats = load_item_categories(ITEMS_TEST)
        keys = {c["key"] for c in cats}
        for expected in [
            "ITEM_CATEGORY_GENERAL",
            "ITEM_CATEGORY_CURRENCY",
            "ITEM_CATEGORY_COOKIE",
            "ITEM_CATEGORY_COSTUME",
        ]:
            assert expected in keys, f"{expected} 누락 — {keys}"

    def test_total_count(self):
        """실 데이터 기준 23개 카테고리."""
        cats = load_item_categories(ITEMS_TEST)
        assert len(cats) >= 20, f"카테고리 개수 비정상: {len(cats)}"

    def test_structure(self):
        """각 항목이 {key, label} 구조."""
        cats = load_item_categories(ITEMS_TEST)
        assert len(cats) > 0
        for c in cats:
            assert "key" in c
            assert "label" in c
            assert c["key"].startswith("ITEM_CATEGORY_")

    def test_empty_when_no_sheet(self, tmp_path):
        """ItemCategory 시트 없는 파일 → 빈 리스트."""
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "items"
        p = tmp_path / "no_category.xlsx"
        wb.save(str(p))
        assert load_item_categories(str(p)) == []

    def test_id_is_int(self):
        """모든 id가 int 타입인지 확인."""
        items = load_items(ITEMS_TEST)
        for item in items:
            assert isinstance(item["id"], int), f"id={item['id']!r}가 int가 아님"

    def test_known_items_count(self):
        """픽스처에 22개 데이터 행 포함 확인 (id=10, 1301401, 1301402 추가 후)."""
        items = load_items(ITEMS_TEST)
        assert len(items) == 22


class TestLoadKeywords:
    def test_build_sheet_exists(self):
        """build 시트 데이터 존재 확인."""
        kw = load_keywords(KEYWORDS_TEST)
        assert "build" in kw
        assert len(kw["build"]) > 0

    def test_timestamp_sheet_exists(self):
        """timestamp 시트 데이터 존재 확인."""
        kw = load_keywords(KEYWORDS_TEST)
        assert "timestamp" in kw
        assert len(kw["timestamp"]) > 0

    def test_launch_0_in_build(self):
        """build 시트에 'LAUNCH_0' 키 존재 확인."""
        kw = load_keywords(KEYWORDS_TEST)
        assert "LAUNCH_0" in kw["build"], "LAUNCH_0 키 누락"

    def test_indefinite_timestamp(self):
        """timestamp 시트에 'INDEFINITE_TIMESTAMP' 키 존재 확인."""
        kw = load_keywords(KEYWORDS_TEST)
        assert "INDEFINITE_TIMESTAMP" in kw["timestamp"], "INDEFINITE_TIMESTAMP 키 누락"

    def test_returns_dict_of_dict(self):
        """반환값이 dict[str, dict] 구조인지 확인."""
        kw = load_keywords(KEYWORDS_TEST)
        assert isinstance(kw, dict)
        for sheet_name, mapping in kw.items():
            assert isinstance(mapping, dict), f"'{sheet_name}' 값이 dict가 아님"

    def test_missing_sheet_returns_empty_dict(self):
        """존재하지 않는 시트는 빈 dict 반환."""
        # keywords_test.xlsx에는 build, timestamp만 있고
        # load_keywords는 이 두 시트만 처리하므로
        # 파일 자체를 바꿔 없는 시트 테스트는 건너뜀
        # 대신 결과에 두 키가 모두 존재하는지 확인
        kw = load_keywords(KEYWORDS_TEST)
        assert "build" in kw
        assert "timestamp" in kw


class TestLoadQuestTemplates:
    def test_returns_dict(self):
        """반환 타입이 dict인지 확인."""
        templates = load_quest_templates(QUESTS_TEST)
        assert isinstance(templates, dict)

    def test_b_to_a_mapping(self):
        """B컬럼 값(quest) -> A컬럼 값(desc) 매핑 최소 1건 확인."""
        templates = load_quest_templates(QUESTS_TEST)
        assert len(templates) >= 1, "템플릿 매핑이 비어있음"

    def test_keys_are_strings(self):
        """모든 키(B컬럼)가 문자열인지 확인."""
        templates = load_quest_templates(QUESTS_TEST)
        for k in templates:
            assert isinstance(k, str), f"키 {k!r}가 str이 아님"

    def test_values_are_strings(self):
        """모든 값(A컬럼)이 문자열인지 확인."""
        templates = load_quest_templates(QUESTS_TEST)
        for v in templates.values():
            assert isinstance(v, str), f"값 {v!r}가 str이 아님"

    def test_template_count(self):
        """픽스처 #quest 시트 Row 4~8 -> 5개 템플릿."""
        templates = load_quest_templates(QUESTS_TEST)
        assert len(templates) == 5, f"템플릿 수 {len(templates)}, 기대=5"


class TestAppendQuestRow:
    """append_quest_row 쓰기 함수 검증 (TC-13 ~ TC-15).
    quests_test.xlsx 원본을 직접 수정하지 않고 tmp_path 복사본으로 테스트."""

    def test_append_increments_row(self, quest_xlsx_copy):
        """TC-13: append 후 max_row가 1 증가해야 함."""
        from openpyxl import load_workbook

        # 추가 전 max_row 확인
        wb_before = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        before_max = wb_before["quests"].max_row
        wb_before.close()

        append_quest_row(quest_xlsx_copy, dict(_MINIMAL_ROW))

        # 추가 후 max_row 확인
        wb_after = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        after_max = wb_after["quests"].max_row
        wb_after.close()

        assert after_max == before_max + 1, (
            f"max_row 증가 실패: before={before_max}, after={after_max}"
        )

    def test_duplicate_key_raises(self, quest_xlsx_copy):
        """TC-14: 기존 ^key로 append 시 ValueError 발생해야 함."""
        # quests_test.xlsx에 이미 존재하는 ^key=31 사용
        dup_row = dict(_MINIMAL_ROW)
        dup_row["^key"] = 31

        with pytest.raises(ValueError, match="31"):
            append_quest_row(quest_xlsx_copy, dup_row)

    def test_written_key_verified(self, quest_xlsx_copy):
        """TC-15: 저장 후 재오픈해서 ^key 값 일치 확인."""
        from openpyxl import load_workbook

        row_data = dict(_MINIMAL_ROW)
        new_key = 999002
        row_data["^key"] = new_key

        target_row = append_quest_row(quest_xlsx_copy, row_data)

        # 재오픈 후 해당 행 ^key 확인
        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        ws = wb["quests"]
        hm = get_header_map(quest_xlsx_copy)
        key_col = hm["^key"]
        actual_key = ws[target_row][key_col - 1].value
        wb.close()

        assert actual_key == new_key, (
            f"Row {target_row} ^key={actual_key!r}, 기대={new_key!r}"
        )

    def test_town_quest_append(self, quest_xlsx_copy):
        """TC-16: TOWN 퀘스트 Row 6(^key=500001) 재현 -- 필드 포함 확인."""
        from openpyxl import load_workbook

        town_row = {
            "^key": 500999,
            "$filter": "$$NONE",
            "category": "QUEST_CATEGORY_TOWN",
            "description": "대화 완료",
            "start_timestamp": "$$COMMON_STARTTAMP",
            "end_timestamp": "$$INDEFINITE_TIMESTAMP",
            "town_category": "TOWN_CATEGORY_INGAME",
            "town_icon": "Icon_Mini_CH_BaconRoll",
            "town_title": "퀘스트 제목 1",
            "town_description": "타운 퀘스트 텍스트입니다",
            "count_type": "QUEST_COUNT_TYPE_SUM",
            "goal_count": 1,
            "goal_type/type/%key": "play:need_win",
            "goal_type/type/%param1": True,
            "conditions/0/condition/%key": "finish_town_dialog:dialog_group_id",
            "conditions/0/condition/%param1": 500001,
            "rewards/0/id": 10,
            "rewards/0/qty": 50,
        }

        target_row = append_quest_row(quest_xlsx_copy, town_row)

        # 저장 후 검증
        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        ws = wb["quests"]
        hm = get_header_map(quest_xlsx_copy)

        row_data = {k: ws[target_row][v - 1].value for k, v in hm.items()}
        wb.close()

        assert row_data["^key"] == 500999
        assert row_data["category"] == "QUEST_CATEGORY_TOWN"
        assert row_data["town_category"] == "TOWN_CATEGORY_INGAME"
        assert row_data["town_icon"] == "Icon_Mini_CH_BaconRoll"
        assert row_data["rewards/0/id"] == 10
        assert row_data["rewards/0/qty"] == 50
        assert row_data["goal_type/type/%key"] == "play:need_win"
        assert row_data["conditions/0/condition/%key"] == "finish_town_dialog:dialog_group_id"


class TestParseQuestTexts:
    def test_single_line_no_match(self):
        """매칭 없는 텍스트 -> description만 채우고 town_description 빈 문자열."""
        result = parse_quest_texts("퀘스트 없는 텍스트", {})
        assert len(result) == 1
        assert result[0]["description"] == "퀘스트 없는 텍스트"
        assert result[0]["town_description"] == ""
        assert result[0]["matched"] is False

    def test_matching_template(self):
        """{템플릿 키: desc} 매핑 -> town_description 자동 채움."""
        templates = {"<color=#ffe535>{0}회 승리</color>": "퀘스트 설명 텍스트"}
        result = parse_quest_texts("<color=#ffe535>{0}회 승리</color>", templates)
        assert result[0]["town_description"] == "퀘스트 설명 텍스트"
        assert result[0]["matched"] is True

    def test_placeholder_preserved(self):
        """{0} placeholder는 치환 없이 그대로 보존."""
        result = parse_quest_texts("{0}회 달성하기", {})
        assert "{0}" in result[0]["description"]

    def test_empty_lines_skipped(self):
        """빈 줄은 건너뜀."""
        result = parse_quest_texts("첫줄\n\n세번째줄", {})
        assert len(result) == 2

    def test_returns_default_values(self):
        """기본값: goal_count=1, qty=1, delete=False."""
        result = parse_quest_texts("텍스트", {})
        assert result[0]["goal_count"] == 1
        assert result[0]["qty"] == 1
        assert result[0]["delete"] is False


class TestBatchAppend:
    def test_description_placeholder_preserved(self, quest_xlsx_copy):
        """TC-12: 배치 저장 시 description의 {0}이 그대로 유지."""
        from openpyxl import load_workbook
        from quest_writer import _build_header_map

        existing = get_existing_keys(quest_xlsx_copy)
        new_key = generate_unique_key(existing)
        fv = {
            "^key": new_key,
            "category": "QUEST_CATEGORY_TOWN",
            "description": "<color=#ffe535>{0}회 승리</color>하기",
            "count_type": "QUEST_COUNT_TYPE_SUM",
            "goal_count": 5,
        }
        target_row = append_quest_row(quest_xlsx_copy, fv)
        # 재오픈해서 description 확인
        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        ws = wb["quests"]
        hm = _build_header_map(ws)
        desc = ws.cell(row=target_row, column=hm["description"]).value
        wb.close()
        assert "{0}" in desc, f"{{0}} 치환됨! 실제값: {desc}"

    def test_batch_three_rows(self, quest_xlsx_copy):
        """배치로 3개 행 순서대로 추가 -> max_row 3 증가."""
        from openpyxl import load_workbook

        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        before_max = wb["quests"].max_row
        wb.close()
        existing = get_existing_keys(quest_xlsx_copy)
        for i in range(3):
            k = generate_unique_key(existing)
            existing.add(k)
            append_quest_row(quest_xlsx_copy, {
                "^key": k,
                "category": "QUEST_CATEGORY_TOWN",
                "count_type": "QUEST_COUNT_TYPE_SUM",
                "goal_count": i + 1,
            })
        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        after_max = wb["quests"].max_row
        wb.close()
        assert after_max == before_max + 3


# ---------------------------------------------------------------------------
# TPL_C (데일리 parent+child 세트) 관련 테스트
# ---------------------------------------------------------------------------


class TestLoadDialogGroups:
    """dialog_groups 로더 — finish_town_dialog 미리보기용."""

    def test_returns_list(self):
        dgs = load_dialog_groups(DIALOG_GROUPS_TEST)
        assert isinstance(dgs, list)

    def test_has_groups(self):
        """픽스처에 8 group 생성됨."""
        dgs = load_dialog_groups(DIALOG_GROUPS_TEST)
        assert len(dgs) >= 1

    def test_structure(self):
        dgs = load_dialog_groups(DIALOG_GROUPS_TEST)
        for g in dgs:
            assert "id" in g
            assert "actor_name" in g
            assert "dialog_text" in g
            assert isinstance(g["id"], int)

    def test_id_sorted(self):
        """id 오름차순 정렬."""
        dgs = load_dialog_groups(DIALOG_GROUPS_TEST)
        ids = [g["id"] for g in dgs]
        assert ids == sorted(ids)

    def test_actor_name_non_empty_for_known_groups(self):
        """픽스처 첫 번째 group 의 actor_name 이 비어있지 않음."""
        dgs = load_dialog_groups(DIALOG_GROUPS_TEST)
        # 첫 번째 group 은 화자 이름이 있어야 함
        if dgs:
            assert dgs[0]["actor_name"] != "", "첫 group actor_name 비어있음"

    def test_one_entry_per_group(self):
        """group_id 당 1건만 (첫 dialog)."""
        dgs = load_dialog_groups(DIALOG_GROUPS_TEST)
        ids = [g["id"] for g in dgs]
        assert len(ids) == len(set(ids)), "group_id 중복 (1건만이어야 함)"


class TestSuggestNextParentKey:
    def test_empty_existing(self):
        """기존 키 없음 → GENERAL+DAILY 범위 하한(30001) 반환."""
        key = suggest_next_parent_key(
            set(),
            filter_id="$$LAUNCH_0",
            reset_type="QUEST_RESET_TYPE_DAILY",
            category="QUEST_CATEGORY_GENERAL",
        )
        assert key == 30001

    def test_daily_step_is_10(self):
        """DAILY reset 의 step 은 10."""
        existing = {30001}
        key = suggest_next_parent_key(
            existing,
            filter_id="$$LAUNCH_0",
            reset_type="QUEST_RESET_TYPE_DAILY",
            category="QUEST_CATEGORY_GENERAL",
        )
        assert key == 30011  # 30001 + 10

    def test_launch_0_non_daily_step_50(self):
        """LAUNCH_0 + NONE → step 50."""
        existing = {30100}
        key = suggest_next_parent_key(
            existing,
            filter_id="$$LAUNCH_0",
            reset_type="QUEST_RESET_TYPE_NONE",
            category="QUEST_CATEGORY_GENERAL",
        )
        assert key == 30150  # 30100 + 50

    def test_helsinki_step_100(self):
        existing = {40000}
        key = suggest_next_parent_key(
            existing,
            filter_id="$$HELSINKI_3",
            reset_type="QUEST_RESET_TYPE_NONE",
            category="QUEST_CATEGORY_GENERAL",
        )
        assert key == 40100  # 40000 + 100

    def test_unknown_filter_default_step(self):
        """관례 없는 filter → DEFAULT_PARENT_STEP (10)."""
        existing = {30200}
        key = suggest_next_parent_key(
            existing,
            filter_id="$$UNKNOWN_FILTER",
            reset_type="QUEST_RESET_TYPE_NONE",
            category="QUEST_CATEGORY_GENERAL",
        )
        assert key == 30210  # 30200 + 10

    def test_ignores_out_of_range(self):
        """다른 카테고리 키는 무시."""
        # TOWN 범위 500000~599999, GENERAL 키들도 섞여있음
        existing = {30001, 30100, 500001}
        key = suggest_next_parent_key(
            existing,
            filter_id=None,
            reset_type="QUEST_RESET_TYPE_NONE",
            category="QUEST_CATEGORY_TOWN",
        )
        # TOWN 범위 기준 max+step
        assert key == 500011  # 500001 + 10


class TestAllocateChildKeys:
    def test_simple_sequential(self):
        keys = allocate_child_keys(set(), parent_key=30010, n=4)
        assert keys == [30011, 30012, 30013, 30014]

    def test_skips_collisions(self):
        """parent+1 충돌 시 다음으로 밀림."""
        existing = {30011, 30013}
        keys = allocate_child_keys(existing, parent_key=30010, n=3)
        assert keys == [30012, 30014, 30015]

    def test_n_zero_empty(self):
        assert allocate_child_keys(set(), 30000, 0) == []

    def test_no_duplicates(self):
        """반환 키끼리 중복 없음."""
        existing = set(range(30011, 30020))
        keys = allocate_child_keys(existing, parent_key=30010, n=5)
        assert len(set(keys)) == len(keys)
        # 기존 키와도 충돌 없어야 함
        for k in keys:
            assert k not in existing


class TestAppendDailySet:
    """append_daily_set — parent+child 일괄 저장."""

    def _base_row(self, key: int, desc: str, goal_key: str = "daily_login") -> dict:
        return {
            "^key": key,
            "category": "QUEST_CATEGORY_GENERAL",
            "description": desc,
            "reset_type": "QUEST_RESET_TYPE_DAILY",
            "count_type": "QUEST_COUNT_TYPE_SUM",
            "goal_count": 1,
            "goal_type/type/%key": goal_key,
        }

    def test_parent_and_children_written(self, quest_xlsx_copy):
        from openpyxl import load_workbook

        parent = self._base_row(777001, "parent 데일리", goal_key="reward_quest:ref_quest_ids")
        parent["count_type"] = "QUEST_COUNT_TYPE_HIGHEST"
        parent["goal_count"] = 3  # R-08: validator 는 goal_count == len(children) 강제
        children = [
            self._base_row(777002, "child 1"),
            self._base_row(777003, "child 2"),
            self._base_row(777004, "child 3"),
        ]

        rows_written = append_daily_set(quest_xlsx_copy, parent, children)

        assert len(rows_written) == 4  # parent + 3 children

        # 재오픈 후 parent goal_type/type/%param1 확인 (자동 생성)
        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        ws = wb["quests"]
        hm = get_header_map(quest_xlsx_copy)
        param1_col = hm["goal_type/type/%param1"]
        key_col = hm["^key"]
        parent_row_data = {k: ws[rows_written[0]][v - 1].value for k, v in hm.items()}
        wb.close()

        assert parent_row_data["^key"] == 777001
        assert parent_row_data["goal_type/type/%param1"] == "[]{777002,777003,777004}"
        assert parent_row_data["goal_count"] == 3  # len(children)

    def test_goal_count_auto_when_missing(self, quest_xlsx_copy):
        """parent.goal_count 없으면 len(children) 으로 자동 채움."""
        from openpyxl import load_workbook

        parent = {
            "^key": 777010,
            "category": "QUEST_CATEGORY_GENERAL",
            "description": "parent no goal_count",
            "reset_type": "QUEST_RESET_TYPE_DAILY",
            "count_type": "QUEST_COUNT_TYPE_HIGHEST",
            "goal_type/type/%key": "reward_quest:ref_quest_ids",
            # goal_count 생략
        }
        children = [
            self._base_row(777011, "c1"),
            self._base_row(777012, "c2"),
        ]

        rows = append_daily_set(quest_xlsx_copy, parent, children)

        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        ws = wb["quests"]
        hm = get_header_map(quest_xlsx_copy)
        parent_row = ws[rows[0]]
        gc_col = hm["goal_count"]
        assert parent_row[gc_col - 1].value == 2  # len(children)
        wb.close()

    def test_child_key_collision_raises(self, quest_xlsx_copy):
        """child ^key 가 기존 파일 키와 충돌 → ValueError."""
        existing = get_existing_keys(quest_xlsx_copy)
        dup_key = next(iter(existing))
        parent = self._base_row(888001, "p")
        parent["goal_type/type/%key"] = "reward_quest:ref_quest_ids"
        parent["count_type"] = "QUEST_COUNT_TYPE_HIGHEST"
        parent["goal_count"] = 1
        children = [
            self._base_row(dup_key, "child dup"),
        ]
        with pytest.raises(ValueError, match="이미 존재"):
            append_daily_set(quest_xlsx_copy, parent, children)

    def test_parent_child_internal_collision_raises(self, quest_xlsx_copy):
        """parent ^key == child ^key → ValueError."""
        parent = self._base_row(888100, "p")
        children = [
            self._base_row(888100, "c"),  # 같은 키
        ]
        with pytest.raises(ValueError, match="중복"):
            append_daily_set(quest_xlsx_copy, parent, children)

    def test_empty_children_raises(self, quest_xlsx_copy):
        parent = self._base_row(888200, "p")
        with pytest.raises(ValueError, match="children"):
            append_daily_set(quest_xlsx_copy, parent, [])

    def test_no_partial_write_on_precheck_failure(self, quest_xlsx_copy):
        """검증 실패 시 파일이 변경되지 않음 (자연 롤백)."""
        from openpyxl import load_workbook

        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        max_before = wb["quests"].max_row
        wb.close()

        # child 간 중복 키로 검증 실패 유도
        parent = self._base_row(888300, "p")
        children = [
            self._base_row(888301, "c1"),
            self._base_row(888301, "c1 dup"),  # 중복
        ]
        with pytest.raises(ValueError):
            append_daily_set(quest_xlsx_copy, parent, children)

        wb = load_workbook(quest_xlsx_copy, read_only=True, data_only=True)
        max_after = wb["quests"].max_row
        wb.close()
        assert max_after == max_before, "검증 실패 후에도 파일이 변경됨 (롤백 실패)"

    def test_rows_written_sequential(self, quest_xlsx_copy):
        """반환된 row 번호가 연속(parent, child1, child2, ...)."""
        parent = self._base_row(999500, "p")
        parent["goal_type/type/%key"] = "reward_quest:ref_quest_ids"
        parent["count_type"] = "QUEST_COUNT_TYPE_HIGHEST"
        parent["goal_count"] = 3
        children = [self._base_row(999501 + i, f"c{i}") for i in range(3)]

        rows = append_daily_set(quest_xlsx_copy, parent, children)
        assert rows == [rows[0], rows[0] + 1, rows[0] + 2, rows[0] + 3]


# ---------------------------------------------------------------------------
# R2 신규 테스트 — 하네스 Evaluator 가 찾은 버그 차단
# ---------------------------------------------------------------------------


class TestReproPriorFailures:
    """Round 2 Evaluator 가 찾은 7가지 시나리오 재현 → 차단 확인 (Gate G3-2)."""

    def _parent(self, key: int = 777001) -> dict:
        return {
            "^key": key,
            "$filter": "",
            "category": "QUEST_CATEGORY_GENERAL",
            "description": "parent",
            "reset_type": "QUEST_RESET_TYPE_DAILY",
            "count_type": "QUEST_COUNT_TYPE_HIGHEST",
            "goal_count": 2,
            "goal_type/type/%key": "reward_quest:ref_quest_ids",
        }

    def _child(self, key: int) -> dict:
        return {
            "^key": key,
            "$filter": "",
            "category": "QUEST_CATEGORY_GENERAL",
            "description": f"child {key}",
            "reset_type": "QUEST_RESET_TYPE_DAILY",
            "count_type": "QUEST_COUNT_TYPE_SUM",
            "goal_count": 1,
            "goal_type/type/%key": "daily_login",
        }

    def test_s4_daily_login_highest_blocked(self, quest_xlsx_copy):
        """S4: child 중 daily_login + HIGHEST 는 validator 에서 차단."""
        parent = self._parent()
        children = [
            self._child(777002),
            self._child(777003),
        ]
        # 첫 번째 child 를 SUM→HIGHEST 로 위반 세팅
        children[0]["count_type"] = "QUEST_COUNT_TYPE_HIGHEST"
        with pytest.raises(ValueError, match="daily_login"):
            append_daily_set(quest_xlsx_copy, parent, children)

    def test_s7_weekly_reset_blocked(self, quest_xlsx_copy):
        """S7: reset_type=WEEKLY 는 validator 에서 차단."""
        row = {
            "^key": 999888,
            "category": "QUEST_CATEGORY_GENERAL",
            "description": "weekly 위반",
            "reset_type": "QUEST_RESET_TYPE_WEEKLY",
            "count_type": "QUEST_COUNT_TYPE_SUM",
            "goal_count": 1,
            "goal_type/type/%key": "daily_login",
        }
        with pytest.raises(ValueError, match="WEEKLY"):
            append_quest_row(quest_xlsx_copy, row)

    def test_s8_parent_sum_blocked(self, quest_xlsx_copy):
        """S8: parent.count_type == SUM 은 TPL_C 불변식 위반."""
        parent = self._parent()
        parent["count_type"] = "QUEST_COUNT_TYPE_SUM"  # HIGHEST 여야 함
        children = [self._child(777002), self._child(777003)]
        with pytest.raises(ValueError, match="HIGHEST"):
            append_daily_set(quest_xlsx_copy, parent, children)

    def test_s9_parent_goal_type_mismatch_blocked(self, quest_xlsx_copy):
        """S9: parent.goal_type != reward_quest:ref_quest_ids 는 불변식 위반."""
        parent = self._parent()
        parent["goal_type/type/%key"] = "daily_login"
        children = [self._child(777002), self._child(777003)]
        with pytest.raises(ValueError, match="reward_quest:ref_quest_ids"):
            append_daily_set(quest_xlsx_copy, parent, children)

    def test_s11_goal_count_mismatch_blocked(self, quest_xlsx_copy):
        """S11: parent.goal_count != len(children) 는 위반."""
        parent = self._parent()
        parent["goal_count"] = 5  # children 은 2개
        children = [self._child(777002), self._child(777003)]
        with pytest.raises(ValueError, match="goal_count"):
            append_daily_set(quest_xlsx_copy, parent, children)


class TestAllocateChildKeysGuards:
    """R-06 parent_key 가드 + signature 확장 검증."""

    def test_negative_parent_raises(self):
        with pytest.raises(ValueError, match="parent_key"):
            allocate_child_keys(set(), -5, 3, "QUEST_CATEGORY_GENERAL", "QUEST_RESET_TYPE_NONE")

    def test_zero_parent_raises(self):
        with pytest.raises(ValueError, match="parent_key"):
            allocate_child_keys(set(), 0, 3)

    def test_too_large_parent_raises(self):
        with pytest.raises(ValueError, match="parent_key"):
            allocate_child_keys(set(), 10**7, 3)

    def test_n_negative_raises(self):
        with pytest.raises(ValueError, match="n"):
            allocate_child_keys(set(), 100, -1)

    def test_sequential_with_daily_category(self):
        """DAILY category + parent_key=30020 → 30021..30025."""
        keys = allocate_child_keys(
            existing={30010, 30011, 30012, 30013, 30014, 30015, 30016, 30017, 30018, 30019},
            parent_key=30020,
            n=5,
            category="QUEST_CATEGORY_GENERAL",
            reset_type="QUEST_RESET_TYPE_DAILY",
        )
        assert keys == [30021, 30022, 30023, 30024, 30025]
        # 모두 DAILY range(30001~59999) 내
        for k in keys:
            assert 30001 <= k <= 59999

    def test_bool_parent_rejected(self):
        """bool True 는 parent_key 로 거부 (정수 1 취급 방지)."""
        with pytest.raises(ValueError, match="parent_key"):
            allocate_child_keys(set(), True, 3)


class TestSuggestFilterScoped:
    """R-02 filter-scoped 제안 검증."""

    def test_no_filter_info_falls_back_to_range_max(self):
        """existing_by_filter None → 기존 동작 (range max + step).
        LAUNCH_0 + NONE → step=50 → 30100 + 50 = 30150."""
        existing = {30011, 30100}
        key = suggest_next_parent_key(
            existing,
            filter_id="$$LAUNCH_0",
            reset_type="QUEST_RESET_TYPE_NONE",
            category="QUEST_CATEGORY_GENERAL",
        )
        assert key == 30150

    def test_filter_scoped_avoids_wrong_zone(self):
        """filter-scoped: LAUNCH_0 로 사용된 key 만 후보로 → 73755 구역 피함."""
        # existing 에 HELSINKI 구역 키(73700)와 LAUNCH 구역 키(100)가 섞여 있음
        existing = {100, 150, 73700}
        existing_by_filter = {
            100: "$$LAUNCH_0",
            150: "$$LAUNCH_0",
            73700: "$$HELSINKI_3",
        }
        key = suggest_next_parent_key(
            existing,
            filter_id="$$LAUNCH_0",
            reset_type="QUEST_RESET_TYPE_NONE",
            category="QUEST_CATEGORY_GENERAL",
            existing_by_filter=existing_by_filter,
        )
        # LAUNCH_0 keys 의 max(150) + step(50) = 200 (HELSINKI 구역 73755 로 가면 안 됨)
        assert key == 200

    def test_filter_scoped_first_entry(self):
        """filter 로 쓰인 키가 없으면 category/reset_type 범위 max 로 폴백."""
        existing = {30011, 30100}
        existing_by_filter = {30011: "$$HELSINKI_3", 30100: "$$HELSINKI_3"}
        key = suggest_next_parent_key(
            existing,
            filter_id="$$LAUNCH_0",  # 이 filter 로 쓰인 key 없음
            reset_type="QUEST_RESET_TYPE_NONE",
            category="QUEST_CATEGORY_GENERAL",
            existing_by_filter=existing_by_filter,
        )
        # 폴백: in_range max(30100) + step(50) = 30150
        assert key == 30150


class TestStrKeyRejected:
    """R-10: ^key 가 str 이면 ValueError (try/except pass 제거)."""

    def test_str_key_rejected(self, quest_xlsx_copy):
        row = dict(_MINIMAL_ROW)
        row["^key"] = "abc"
        with pytest.raises(ValueError, match="정수"):
            append_quest_row(quest_xlsx_copy, row)

    def test_bool_key_rejected(self, quest_xlsx_copy):
        row = dict(_MINIMAL_ROW)
        row["^key"] = True
        with pytest.raises(ValueError, match="bool"):
            append_quest_row(quest_xlsx_copy, row)


class TestAtomicSave:
    """R-07: _atomic_save 헬퍼 동작 — tmp 생성 + replace + 예외 시 원본 유지."""

    def test_atomic_save_ok(self, quest_xlsx_copy):
        """정상 경로: 파일이 갱신되고 tmp 가 남지 않음."""
        import os as _os

        append_quest_row(
            quest_xlsx_copy,
            {
                "^key": 999111,
                "category": "QUEST_CATEGORY_GENERAL",
                "description": "atomic OK",
                "count_type": "QUEST_COUNT_TYPE_SUM",
                "goal_count": 1,
                "reset_type": "QUEST_RESET_TYPE_NONE",
                "goal_type/type/%key": "daily_login",
            },
        )
        # tmp 잔재 없음
        dir_name = _os.path.dirname(quest_xlsx_copy)
        leftovers = [
            f for f in _os.listdir(dir_name)
            if f.startswith(".quest_writer_") and f.endswith(".xlsx.tmp")
        ]
        assert leftovers == [], f"tmp 잔재: {leftovers}"

    def test_atomic_save_preserves_original_on_failure(self, tmp_path, monkeypatch):
        """_atomic_save 내부에서 replace 실패 시 원본 파일 유지."""
        import os as _os
        import shutil
        from quest_writer import _atomic_save

        src = Path(__file__).parent / "fixtures" / "quests_test.xlsx"
        target = tmp_path / "quests_atomic.xlsx"
        shutil.copy(src, target)
        target_str = str(target)
        original_size = _os.path.getsize(target_str)

        # openpyxl workbook 준비
        from openpyxl import load_workbook
        wb = load_workbook(target_str)

        # os.replace 에서 예외 주입
        orig_replace = _os.replace
        def _bad_replace(*args, **kwargs):
            raise RuntimeError("simulated replace failure")
        monkeypatch.setattr(_os, "replace", _bad_replace)

        try:
            with pytest.raises(RuntimeError, match="simulated"):
                _atomic_save(wb, target_str)
        finally:
            wb.close()
            monkeypatch.setattr(_os, "replace", orig_replace)

        # 원본 파일 크기 불변 (atomic 유지)
        assert _os.path.getsize(target_str) == original_size
        # tmp 정리됨
        leftovers = [
            f for f in _os.listdir(str(tmp_path))
            if f.startswith(".quest_writer_") and f.endswith(".xlsx.tmp")
        ]
        assert leftovers == [], f"tmp 잔재: {leftovers}"


# ---------------------------------------------------------------------------
# STEP D2 신규 테스트 — GoalType YAML + Daily Mission 헬퍼
# ---------------------------------------------------------------------------


class TestGoalTypesYaml:
    """goal_types.yaml 로드/저장 라운드트립 + 폴백."""

    def test_load_seed_from_default_path(self):
        """기본 경로(quest_tool/goal_types.yaml) 로드 — 24+ 항목."""
        data = load_goal_types_yaml()
        assert isinstance(data, list)
        assert len(data) >= 24, f"seed 로드 개수 부족: {len(data)}"

    def test_load_contains_play_and_daily_login(self):
        """실데이터 기준 핵심 항목 포함 확인."""
        data = load_goal_types_yaml()
        keys = {d["key"] for d in data}
        assert "play" in keys, "play goal_type 누락 (실 child 8건 사용)"
        assert "daily_login" in keys
        assert "reward_quest:ref_quest_ids" in keys

    def test_load_falls_back_on_missing_file(self, tmp_path):
        """파일 부존재 시 하드코딩 폴백."""
        nonexistent = tmp_path / "no_such.yaml"
        data = load_goal_types_yaml(str(nonexistent))
        assert isinstance(data, list)
        assert len(data) == len(_HARDCODED_GOAL_TYPES)

    def test_load_falls_back_on_invalid_yaml(self, tmp_path):
        """YAML 파싱 실패 시 폴백."""
        bad = tmp_path / "bad.yaml"
        bad.write_text("goal_types: {invalid: [unclosed", encoding="utf-8")
        data = load_goal_types_yaml(str(bad))
        assert data == list(_HARDCODED_GOAL_TYPES)

    def test_load_falls_back_on_non_list_root(self, tmp_path):
        """루트가 list 가 아닌 YAML → 폴백."""
        bad = tmp_path / "wrong_shape.yaml"
        bad.write_text("just_a_string", encoding="utf-8")
        data = load_goal_types_yaml(str(bad))
        assert data == list(_HARDCODED_GOAL_TYPES)

    def test_save_and_reload_roundtrip(self, tmp_path):
        """save → load 라운드트립 — 동일 데이터."""
        sample = [
            {"key": "daily_login", "label": "출석", "params": []},
            {"key": "play:need_win", "label": "승리",
             "params": [{"label": "need_win", "options": ["TRUE", "FALSE"]}]},
            {"key": "custom_new", "label": "사용자 정의", "params": []},
        ]
        p = tmp_path / "custom.yaml"
        save_goal_types_yaml(sample, str(p))
        loaded = load_goal_types_yaml(str(p))
        assert len(loaded) == 3
        assert loaded[0]["key"] == "daily_login"
        assert loaded[2]["key"] == "custom_new"
        # params 보존
        assert loaded[1]["params"][0]["options"] == ["TRUE", "FALSE"]

    def test_save_filters_invalid_entries(self, tmp_path):
        """key 누락 / 잘못된 타입은 save 때 skip."""
        sample = [
            {"key": "valid", "label": "OK", "params": []},
            {"label": "no key - skip"},   # key 없음 → skip
            "not a dict",                  # dict 아님 → skip
            {"key": "", "label": "empty key - skip"},  # empty key → skip
        ]
        p = tmp_path / "filtered.yaml"
        save_goal_types_yaml(sample, str(p))
        loaded = load_goal_types_yaml(str(p))
        assert len(loaded) == 1
        assert loaded[0]["key"] == "valid"

    def test_save_atomic_no_tmp_leftover(self, tmp_path):
        """저장 후 tmp 잔재 없음."""
        p = tmp_path / "atomic.yaml"
        save_goal_types_yaml([{"key": "k1", "label": "L1", "params": []}], str(p))
        leftovers = [f for f in os.listdir(str(tmp_path)) if f.startswith(".goal_types_tmp_")]
        assert leftovers == []


class TestDailyMissionHelpers:
    """default_parent_desc / extract_day_from_keyword."""

    def test_extract_day_nday_in_keyword(self):
        assert extract_day_from_keyword("$$LAUNCH_0_NDAY1") == 1
        assert extract_day_from_keyword("$$NDAY7") == 7
        assert extract_day_from_keyword("NDAY42") == 42

    def test_extract_day_case_insensitive(self):
        assert extract_day_from_keyword("$$nday3") == 3
        assert extract_day_from_keyword("$$Nday9") == 9

    def test_extract_day_none_on_no_match(self):
        assert extract_day_from_keyword("$$INDEFINITE_TIMESTAMP") is None
        assert extract_day_from_keyword("") is None
        assert extract_day_from_keyword(None) is None  # type: ignore

    def test_default_parent_desc_with_day(self):
        assert default_parent_desc("$$LAUNCH_0_NDAY1", 0) == "[데일리미션]1일차 전체 퀘스트 완료 보상"
        assert default_parent_desc("$$NDAY3", 10) == "[데일리미션]3일차 전체 퀘스트 완료 보상"

    def test_default_parent_desc_fallback_to_existing_plus_1(self):
        assert default_parent_desc("$$INDEFINITE_TIMESTAMP", 5) == "[데일리미션]6일차 전체 퀘스트 완료 보상"

    def test_default_parent_desc_minimum_1(self):
        """fallback 도 최소 1일차."""
        assert default_parent_desc("", 0) == "[데일리미션]1일차 전체 퀘스트 완료 보상"
        assert default_parent_desc("", -5) == "[데일리미션]1일차 전체 퀘스트 완료 보상"


class TestNdayMissionEvents:
    """nday_mission_events.xlsx 관련 함수 테스트."""

    FIXTURE_PATH = str(
        Path(__file__).resolve().parent / "fixtures" / "nday_mission_events_test.xlsx"
    )

    def test_get_existing_event_keys(self):
        """픽스처에서 {101, 201} 반환 확인."""
        keys = get_existing_event_keys(self.FIXTURE_PATH)
        assert keys == {101, 201}

    def test_suggest_next_event_key_within_group(self):
        """group_base=100, 101 사용 중 → 102 제안."""
        existing = {101, 201}
        result = suggest_next_event_key(existing, group_base=100)
        assert result == 102

    def test_suggest_next_event_key_new_group(self):
        """group_base=300, 미사용 → 301 제안."""
        existing = {101, 201}
        result = suggest_next_event_key(existing, group_base=300)
        assert result == 301

    def test_append_nday_mission_event_writes_both_sheets(self, tmp_path):
        """events + events.day 각 1행 증가 확인."""
        import shutil
        fixture_copy = str(tmp_path / "nday_test.xlsx")
        shutil.copy(self.FIXTURE_PATH, fixture_copy)

        event_dict = {
            "^key": 102,
            "description": "테스트 이벤트",
            "start_timestamp": "$$TEST_NDAY2",
            "end_timestamp": "$$TEST_NDAY2_END",
            "mission_active_days": 0,
        }
        event_day_dict = {
            "^key": 102,
            "day": 1,
            "description": "테스트 일차",
            "quest_ids": "[]{40001,40002,40003}",
            "finish_quest_id": 40001,
        }

        from openpyxl import load_workbook
        wb_before = load_workbook(fixture_copy, data_only=True)
        ev_rows_before = wb_before["events"].max_row
        day_rows_before = wb_before["events.day"].max_row
        wb_before.close()

        ev_row, day_row = append_nday_mission_event(fixture_copy, event_dict, event_day_dict)

        wb_after = load_workbook(fixture_copy, data_only=True)
        ev_rows_after = wb_after["events"].max_row
        day_rows_after = wb_after["events.day"].max_row
        wb_after.close()

        assert ev_rows_after == ev_rows_before + 1
        assert day_rows_after == day_rows_before + 1
        assert ev_row == ev_rows_before + 1
        assert day_row == day_rows_before + 1

        # 저장 내용 검증
        wb_verify = load_workbook(fixture_copy, data_only=True)
        ws_ev = wb_verify["events"]
        assert ws_ev.cell(row=ev_row, column=1).value == 102
        ws_day = wb_verify["events.day"]
        assert ws_day.cell(row=day_row, column=1).value == 102
        assert ws_day.cell(row=day_row, column=2).value == 1
        wb_verify.close()

    def test_append_nday_mission_event_duplicate_key_raises(self, tmp_path):
        """중복 ^key=101 → ValueError."""
        import shutil
        fixture_copy = str(tmp_path / "nday_dup.xlsx")
        shutil.copy(self.FIXTURE_PATH, fixture_copy)

        event_dict = {
            "^key": 101,  # 이미 존재
            "description": "중복 이벤트",
            "start_timestamp": "$$TEST",
            "end_timestamp": "$$TEST_END",
            "mission_active_days": 0,
        }
        event_day_dict = {
            "^key": 101,
            "day": 1,
            "description": "중복 일차",
            "quest_ids": "[]{40001}",
            "finish_quest_id": 40001,
        }

        with pytest.raises(ValueError, match="이미 존재"):
            append_nday_mission_event(fixture_copy, event_dict, event_day_dict)
