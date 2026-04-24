"""STEP N-day 재현 테스트 — "가정의 달! 스페셜 미션" 7일치.

사용자 예시:
- parent description: "[데일리미션]1일차 전체 퀘스트 완료 보상" ~ "[데일리미션]7일차 전체 퀘스트 완료 보상"
- events.description: "가정의 달! 스페셜 미션" (7일 모두 동일)
- events.day.description: "데일리 미션 1일차" ~ "데일리 미션 7일차"

이 테스트는 render_daily_mission_form 의 "저장" 버튼 플로우를 함수 호출 수준에서 재현한다:
1) for d in 0..6: append_daily_set(quests_path, parent_row_d, child_rows_d)
2) for d in 0..6: append_nday_mission_event(nday_path, event_dict_d, event_day_dict_d)
3) 양쪽 xlsx 의 셀 값, 키 연결(quest_ids <-> child ^keys, finish_quest_id <-> parent ^key)를 검증.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from quest_writer import (
    _build_header_map as qw_build_header_map,
    allocate_child_keys,
    append_daily_set,
    append_nday_mission_event,
    get_existing_event_keys,
    get_existing_keys,
    get_existing_keys_by_filter,
    get_header_map,
    suggest_next_parent_key,
)

FIXTURES = Path(__file__).resolve().parent / "fixtures"
QUESTS_SRC = FIXTURES / "quests_test.xlsx"
NDAY_SRC = FIXTURES / "nday_mission_events_test.xlsx"


# --- 테스트용 입력 데이터 ("가정의 달! 스페셜 미션" 7일치) ---
EVENT_NAME = "가정의 달! 스페셜 미션"
N_DAYS = 7
N_CHILDREN = 5
DAILY_FILTER = "$$LAUNCH_0"
DAILY_START = "$$NDAY5"
DAILY_END = "$$NDAY5_END"
MISSION_ACTIVE_DAYS = 0


def _build_day_data() -> list[dict]:
    """UI 의 dm_days 구조를 재현."""
    days: list[dict] = []
    for d in range(N_DAYS):
        day_num = d + 1
        children = []
        for i in range(N_CHILDREN):
            children.append({
                "description": f"{day_num}일차 미션 {i + 1}",
                "goal_type_key": "play:need_win",
                "goal_type_param1": "TRUE",
                "goal_count": 3,
                "reward_category": "",
                "reward_id": 0,
                "reward_qty": 1,
            })
        days.append({
            "parent_description": f"[데일리미션]{day_num}일차 전체 퀘스트 완료 보상",
            "event_description": EVENT_NAME,
            "event_day_description": f"데일리 미션 {day_num}일차",
            "children": children,
        })
    return days


def _perform_save(quests_path: str, nday_path: str) -> dict:
    """render_daily_mission_form 의 저장 버튼 플로우 재현.

    Returns:
        {"parent_keys": [...], "event_keys": [...], "child_matrix": [[...], ...],
         "group_base": int, "saved": [...]}
    """
    days = _build_day_data()

    # --- 첫 parent ^key 제안 ---
    existing_keys = get_existing_keys(quests_path)
    existing_by_filter = get_existing_keys_by_filter(quests_path)
    first_parent = suggest_next_parent_key(
        existing_keys,
        filter_id=DAILY_FILTER,
        reset_type="QUEST_RESET_TYPE_DAILY",
        category="QUEST_CATEGORY_GENERAL",
        existing_by_filter=existing_by_filter,
    )
    parent_keys = [first_parent + 10 * i for i in range(N_DAYS)]

    # --- 이벤트 그룹 자동 계산 ---
    existing_event_keys = get_existing_event_keys(nday_path)
    used_groups = {(ek // 100) * 100 for ek in existing_event_keys if ek >= 100}
    group_base = (max(used_groups) + 100) if used_groups else 100
    event_keys = [group_base + 1 + i for i in range(N_DAYS)]

    # --- child keys 할당 (누적) ---
    child_matrix: list[list[int]] = []
    existing_accum = set(existing_keys)
    for d in range(N_DAYS):
        existing_accum.add(parent_keys[d])
        cks = allocate_child_keys(existing_accum, parent_keys[d], N_CHILDREN)
        existing_accum.update(cks)
        child_matrix.append(cks)

    # --- 저장 루프 ---
    saved = []
    for d in range(N_DAYS):
        day = days[d]
        pk = parent_keys[d]
        ek = event_keys[d]
        cks = child_matrix[d]

        parent_row = {
            "^key": pk,
            "$filter": DAILY_FILTER,
            "category": "QUEST_CATEGORY_GENERAL",
            "description": day["parent_description"],
            "start_timestamp": DAILY_START,
            "end_timestamp": DAILY_END,
            "reset_type": "QUEST_RESET_TYPE_DAILY",
            "count_type": "QUEST_COUNT_TYPE_HIGHEST",
            "goal_count": N_CHILDREN,
            "goal_type/type/%key": "reward_quest:ref_quest_ids",
        }
        child_rows = []
        for i in range(N_CHILDREN):
            c = day["children"][i]
            ck = cks[i]
            cd = {
                "^key": ck,
                "$filter": DAILY_FILTER,
                "category": "QUEST_CATEGORY_GENERAL",
                "description": c["description"],
                "start_timestamp": DAILY_START,
                "end_timestamp": DAILY_END,
                "reset_type": "QUEST_RESET_TYPE_DAILY",
                "count_type": "QUEST_COUNT_TYPE_SUM",
                "goal_count": c["goal_count"],
                "goal_type/type/%key": c["goal_type_key"],
                "goal_type/type/%param1": c["goal_type_param1"],
            }
            child_rows.append(cd)

        append_daily_set(quests_path, parent_row, child_rows)

        event_dict = {
            "^key": ek,
            "description": day["event_description"],
            "start_timestamp": DAILY_START,
            "end_timestamp": DAILY_END,
            "mission_active_days": MISSION_ACTIVE_DAYS,
        }
        event_day_dict = {
            "^key": ek,
            "day": 1,
            "description": day["event_day_description"],
            "quest_ids": "[]{" + ",".join(str(k) for k in cks) + "}",
            "finish_quest_id": pk,
        }
        append_nday_mission_event(nday_path, event_dict, event_day_dict)
        saved.append({"day": d + 1, "pk": pk, "ek": ek, "cks": cks})

    return {
        "parent_keys": parent_keys,
        "event_keys": event_keys,
        "child_matrix": child_matrix,
        "group_base": group_base,
        "saved": saved,
        "days": days,
    }


def _build_hm_quests(ws) -> dict[str, int]:
    return qw_build_header_map(ws, path_row=2, header_row=3, base_prefix="quests/^0/")


def _build_hm_events(ws) -> dict[str, int]:
    """events / events.day 시트는 append_nday_mission_event 와 동일하게 base_prefix='' 사용.
    필드명만 키로 써서 ^key, description, quest_ids, finish_quest_id 를 직접 조회."""
    return qw_build_header_map(ws, path_row=2, header_row=3, base_prefix="")


# =============================================================================
# 테스트 클래스
# =============================================================================

class TestNdayReproduce7Days:
    """가정의 달! 스페셜 미션 7일치 재현 검증."""

    @pytest.fixture
    def work(self, tmp_path):
        """quests/nday_mission_events 테스트 픽스처를 tmp_path 로 복사."""
        q_work = tmp_path / "quests_work.xlsx"
        n_work = tmp_path / "nday_work.xlsx"
        shutil.copy(QUESTS_SRC, q_work)
        shutil.copy(NDAY_SRC, n_work)
        return {"quests_path": str(q_work), "nday_path": str(n_work)}

    def test_save_flow_produces_expected_key_layout(self, work):
        """저장 후 parent_keys / event_keys / child_matrix 구조 검증."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        # parent_keys: 첫 값에서 +10 씩 N_DAYS개
        assert len(result["parent_keys"]) == N_DAYS
        assert all(
            result["parent_keys"][i + 1] - result["parent_keys"][i] == 10
            for i in range(N_DAYS - 1)
        )

        # event_keys: group_base+1 ~ group_base+N_DAYS
        assert result["event_keys"] == [
            result["group_base"] + 1 + i for i in range(N_DAYS)
        ]
        # group_base 는 기존 {101, 201} 존재하므로 300
        assert result["group_base"] == 300
        assert result["event_keys"] == [301, 302, 303, 304, 305, 306, 307]

        # child_matrix: N_DAYS × N_CHILDREN, 전역 유일
        assert len(result["child_matrix"]) == N_DAYS
        all_ck = [ck for row in result["child_matrix"] for ck in row]
        assert len(all_ck) == N_DAYS * N_CHILDREN
        assert len(set(all_ck)) == len(all_ck), "child keys 전역 유일 실패"
        # parent 와 child 중복 없음
        assert set(all_ck).isdisjoint(set(result["parent_keys"]))

    def test_quests_xlsx_has_all_parent_child_rows(self, work):
        """quests.xlsx 에 parent 7개 + child 35개 기입 확인."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["quests_path"], data_only=True)
        ws = wb["quests"]
        hm = _build_hm_quests(ws)
        wb.close()

        key_col = hm["^key"]

        wb = load_workbook(work["quests_path"], data_only=True)
        ws = wb["quests"]
        all_keys: list[int] = []
        for r in range(4, ws.max_row + 1):
            v = ws.cell(row=r, column=key_col).value
            if isinstance(v, int):
                all_keys.append(v)
        wb.close()

        for pk in result["parent_keys"]:
            assert pk in all_keys, f"parent ^key={pk} 없음"
        for row in result["child_matrix"]:
            for ck in row:
                assert ck in all_keys, f"child ^key={ck} 없음"

    def test_quests_parent_descriptions_are_N_day_pattern(self, work):
        """parent description: [데일리미션]N일차 전체 퀘스트 완료 보상."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["quests_path"], data_only=True)
        ws = wb["quests"]
        hm = _build_hm_quests(ws)

        key_col = hm["^key"]
        desc_col = hm["description"]

        for d, pk in enumerate(result["parent_keys"]):
            for r in range(4, ws.max_row + 1):
                if ws.cell(row=r, column=key_col).value == pk:
                    desc = ws.cell(row=r, column=desc_col).value
                    expected = f"[데일리미션]{d + 1}일차 전체 퀘스트 완료 보상"
                    assert desc == expected, f"parent {pk} desc: {desc!r} != {expected!r}"
                    break
            else:
                pytest.fail(f"parent ^key={pk} 행 못 찾음")
        wb.close()

    def test_quests_parent_goal_param1_matches_children(self, work):
        """parent 의 goal_type/%param1 = []{child1,child2,...}."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["quests_path"], data_only=True)
        ws = wb["quests"]
        hm = _build_hm_quests(ws)

        key_col = hm["^key"]
        # quests.xlsx 에는 여러 %param1 컬럼이 있음 (conditions/, goal_type/, reward_condition/, ...).
        # parent goal_type/type/%param1 을 정확히 지정.
        assert "goal_type/type/%param1" in hm, f"goal_type/type/%param1 컬럼 없음. hm keys: {list(hm.keys())}"
        param1_col = hm["goal_type/type/%param1"]

        for d, pk in enumerate(result["parent_keys"]):
            expected_p1 = "[]{" + ",".join(str(k) for k in result["child_matrix"][d]) + "}"
            for r in range(4, ws.max_row + 1):
                if ws.cell(row=r, column=key_col).value == pk:
                    actual = ws.cell(row=r, column=param1_col).value
                    assert actual == expected_p1, (
                        f"parent {pk} %param1: {actual!r} != {expected_p1!r}"
                    )
                    break
            else:
                pytest.fail(f"parent ^key={pk} 행 못 찾음")
        wb.close()

    def test_nday_events_sheet_has_7_rows(self, work):
        """nday_mission_events.xlsx events 시트에 7행 추가 확인."""
        wb_before = load_workbook(NDAY_SRC, data_only=True)
        ev_before = wb_before["events"].max_row
        day_before = wb_before["events.day"].max_row
        wb_before.close()

        _perform_save(work["quests_path"], work["nday_path"])

        wb_after = load_workbook(work["nday_path"], data_only=True)
        ev_after = wb_after["events"].max_row
        day_after = wb_after["events.day"].max_row
        wb_after.close()

        assert ev_after - ev_before == N_DAYS
        assert day_after - day_before == N_DAYS

    def test_nday_events_all_have_same_event_description(self, work):
        """events.description = '가정의 달! 스페셜 미션' 7건 일치."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["nday_path"], data_only=True)
        ws = wb["events"]
        hm = _build_hm_events(ws)
        key_col = hm["^key"]
        desc_col = hm["description"]

        for ek in result["event_keys"]:
            for r in range(4, ws.max_row + 1):
                if ws.cell(row=r, column=key_col).value == ek:
                    desc = ws.cell(row=r, column=desc_col).value
                    assert desc == EVENT_NAME, (
                        f"event {ek} desc: {desc!r} != {EVENT_NAME!r}"
                    )
                    break
            else:
                pytest.fail(f"event ^key={ek} 행 못 찾음")
        wb.close()

    def test_nday_events_day_descriptions_are_N_day_pattern(self, work):
        """events.day.description = '데일리 미션 N일차' 패턴."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["nday_path"], data_only=True)
        ws = wb["events.day"]
        hm = _build_hm_events(ws)
        key_col = hm["^key"]
        desc_col = hm["description"]

        for d, ek in enumerate(result["event_keys"]):
            for r in range(4, ws.max_row + 1):
                if ws.cell(row=r, column=key_col).value == ek:
                    desc = ws.cell(row=r, column=desc_col).value
                    expected = f"데일리 미션 {d + 1}일차"
                    assert desc == expected, (
                        f"events.day {ek} desc: {desc!r} != {expected!r}"
                    )
                    break
            else:
                pytest.fail(f"events.day ^key={ek} 행 못 찾음")
        wb.close()

    def test_cross_reference_quest_ids_match_child_keys(self, work):
        """events.day.quest_ids 가 해당 day 의 child ^keys 를 정확히 포함."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["nday_path"], data_only=True)
        ws = wb["events.day"]
        hm = _build_hm_events(ws)
        key_col = hm["^key"]
        qids_col = hm["quest_ids"]

        for d, ek in enumerate(result["event_keys"]):
            expected = "[]{" + ",".join(str(k) for k in result["child_matrix"][d]) + "}"
            for r in range(4, ws.max_row + 1):
                if ws.cell(row=r, column=key_col).value == ek:
                    actual = ws.cell(row=r, column=qids_col).value
                    assert actual == expected, (
                        f"events.day {ek} quest_ids: {actual!r} != {expected!r}"
                    )
                    break
            else:
                pytest.fail(f"events.day ^key={ek} 행 못 찾음")
        wb.close()

    def test_cross_reference_finish_quest_id_matches_parent(self, work):
        """events.day.finish_quest_id 가 해당 day 의 parent ^key 와 일치."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["nday_path"], data_only=True)
        ws = wb["events.day"]
        hm = _build_hm_events(ws)
        key_col = hm["^key"]
        fqid_col = hm["finish_quest_id"]

        for d, ek in enumerate(result["event_keys"]):
            for r in range(4, ws.max_row + 1):
                if ws.cell(row=r, column=key_col).value == ek:
                    actual = ws.cell(row=r, column=fqid_col).value
                    assert actual == result["parent_keys"][d], (
                        f"events.day {ek} finish_quest_id: "
                        f"{actual!r} != {result['parent_keys'][d]!r}"
                    )
                    break
            else:
                pytest.fail(f"events.day ^key={ek} 행 못 찾음")
        wb.close()

    def test_timestamps_are_identical_between_quests_and_events(self, work):
        """events.start/end_timestamp == quests.start/end_timestamp (실 데이터 패턴)."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        # quests
        wb_q = load_workbook(work["quests_path"], data_only=True)
        ws_q = wb_q["quests"]
        hm_q = _build_hm_quests(ws_q)
        key_col_q = hm_q["^key"]
        st_col_q = hm_q["start_timestamp"]
        en_col_q = hm_q["end_timestamp"]

        # events
        wb_e = load_workbook(work["nday_path"], data_only=True)
        ws_e = wb_e["events"]
        hm_e = _build_hm_events(ws_e)
        key_col_e = hm_e["^key"]
        st_col_e = hm_e["start_timestamp"]
        en_col_e = hm_e["end_timestamp"]

        for d in range(N_DAYS):
            pk = result["parent_keys"][d]
            ek = result["event_keys"][d]

            q_start = None; q_end = None
            for r in range(4, ws_q.max_row + 1):
                if ws_q.cell(row=r, column=key_col_q).value == pk:
                    q_start = ws_q.cell(row=r, column=st_col_q).value
                    q_end = ws_q.cell(row=r, column=en_col_q).value
                    break
            e_start = None; e_end = None
            for r in range(4, ws_e.max_row + 1):
                if ws_e.cell(row=r, column=key_col_e).value == ek:
                    e_start = ws_e.cell(row=r, column=st_col_e).value
                    e_end = ws_e.cell(row=r, column=en_col_e).value
                    break

            assert q_start == e_start, (
                f"day {d+1}: quests.start_timestamp={q_start!r} "
                f"!= events.start_timestamp={e_start!r}"
            )
            assert q_end == e_end, (
                f"day {d+1}: quests.end_timestamp={q_end!r} "
                f"!= events.end_timestamp={e_end!r}"
            )

        wb_q.close()
        wb_e.close()

    def test_event_key_equals_event_day_key(self, work):
        """각 이벤트의 events.^key == events.day.^key (사용자 확정 사항)."""
        result = _perform_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["nday_path"], data_only=True)
        ws_ev = wb["events"]
        ws_day = wb["events.day"]
        hm_ev = _build_hm_events(ws_ev)
        hm_day = _build_hm_events(ws_day)

        ev_keys: set[int] = set()
        for r in range(4, ws_ev.max_row + 1):
            v = ws_ev.cell(row=r, column=hm_ev["^key"]).value
            if isinstance(v, int):
                ev_keys.add(v)

        day_keys: set[int] = set()
        for r in range(4, ws_day.max_row + 1):
            v = ws_day.cell(row=r, column=hm_day["^key"]).value
            if isinstance(v, int):
                day_keys.add(v)

        for ek in result["event_keys"]:
            assert ek in ev_keys, f"events.{ek} 없음"
            assert ek in day_keys, f"events.day.{ek} 없음"

        wb.close()

    def test_event_group_auto_advances_on_next_save(self, work):
        """1차 저장(300대) 후, 2차 저장 시 자동으로 400대 할당 확인."""
        result1 = _perform_save(work["quests_path"], work["nday_path"])
        assert result1["group_base"] == 300

        # 두 번째 저장 — 같은 파일에 이어서
        result2 = _perform_save(work["quests_path"], work["nday_path"])
        assert result2["group_base"] == 400, (
            f"2차 group_base 는 400 이어야 함: {result2['group_base']}"
        )
        assert result2["event_keys"] == [401, 402, 403, 404, 405, 406, 407]


# =============================================================================
# goal_type 별 param1 매핑 검증 (사용자 확정 규칙)
# =============================================================================
# | goal_type %key          | %param1 값                                   |
# |-------------------------|----------------------------------------------|
# | play                    | 없음 (컬럼 비움)                              |
# | play:need_win           | FALSE 또는 TRUE                              |
# | play_report:key         | PLAY_REPORT_KEY_* 중 하나                    |
# | get_ovencrown           | 없음                                          |
# | open_battle_box:box_type| BOX_TYPE_BATTLE / BOX_TYPE_COIN / ...        |
# | pull_gacha:gacha_type   | GACHA_TYPE_COOKIE_GACHA / ...                |


GOAL_TYPE_PARAM_CASES = [
    # (goal_type_key, goal_type_param1_input, expected_xlsx_value)
    ("play", None, None),
    ("play:need_win", "TRUE", "TRUE"),
    ("play:need_win", "FALSE", "FALSE"),
    ("play_report:key", "PLAY_REPORT_KEY_ULTIMATE_DAMAGE", "PLAY_REPORT_KEY_ULTIMATE_DAMAGE"),
    ("play_report:key", "PLAY_REPORT_KEY_KILL_COUNT", "PLAY_REPORT_KEY_KILL_COUNT"),
    ("get_ovencrown", None, None),
    ("open_battle_box:box_type", "BOX_TYPE_BATTLE", "BOX_TYPE_BATTLE"),
    ("pull_gacha:gacha_type", "GACHA_TYPE_COOKIE_GACHA", "GACHA_TYPE_COOKIE_GACHA"),
]


class TestGoalTypeParamMapping:
    """child 의 goal_type/%key → goal_type/%param1 규칙 검증.

    UI 의 render_param_fields 가 각 goal_type 에 맞는 param1 을 생성하고,
    append_daily_set 이 xlsx 에 정확히 기록하는지 확인.
    """

    @pytest.fixture
    def work(self, tmp_path):
        q_work = tmp_path / "quests_work.xlsx"
        n_work = tmp_path / "nday_work.xlsx"
        shutil.copy(QUESTS_SRC, q_work)
        shutil.copy(NDAY_SRC, n_work)
        return {"quests_path": str(q_work), "nday_path": str(n_work)}

    def test_play_report_key_has_ultimate_damage_option(self):
        """PLAY_REPORT_KEY_ULTIMATE_DAMAGE 옵션이 goal_types.yaml 에 추가됨."""
        from quest_writer import load_goal_types_yaml
        gtypes = load_goal_types_yaml()
        pr = next((g for g in gtypes if g["key"] == "play_report:key"), None)
        assert pr is not None, "play_report:key goal_type 없음"
        params = pr.get("params", [])
        assert params, "play_report:key params 비어있음"
        opts = params[0].get("options", [])
        assert "PLAY_REPORT_KEY_ULTIMATE_DAMAGE" in opts, (
            f"PLAY_REPORT_KEY_ULTIMATE_DAMAGE 옵션 없음. 현재: {opts}"
        )

    @pytest.mark.parametrize("gkey,p1_input,p1_expected", GOAL_TYPE_PARAM_CASES)
    def test_param1_written_correctly(self, work, gkey, p1_input, p1_expected):
        """각 goal_type 케이스별로 param1 이 xlsx 에 올바르게 기록되는지."""
        # 1일치, 단일 child (테스트 빠른 실행을 위해)
        existing_keys = get_existing_keys(work["quests_path"])
        existing_by_filter = get_existing_keys_by_filter(work["quests_path"])
        first_parent = suggest_next_parent_key(
            existing_keys,
            filter_id=DAILY_FILTER,
            reset_type="QUEST_RESET_TYPE_DAILY",
            category="QUEST_CATEGORY_GENERAL",
            existing_by_filter=existing_by_filter,
        )
        cks = allocate_child_keys(existing_keys | {first_parent}, first_parent, 1)
        ck = cks[0]

        parent_row = {
            "^key": first_parent,
            "$filter": DAILY_FILTER,
            "category": "QUEST_CATEGORY_GENERAL",
            "description": "[데일리미션]1일차 전체 퀘스트 완료 보상",
            "start_timestamp": DAILY_START,
            "end_timestamp": DAILY_END,
            "reset_type": "QUEST_RESET_TYPE_DAILY",
            "count_type": "QUEST_COUNT_TYPE_HIGHEST",
            "goal_count": 1,
            "goal_type/type/%key": "reward_quest:ref_quest_ids",
        }
        child_dict: dict = {
            "^key": ck,
            "$filter": DAILY_FILTER,
            "category": "QUEST_CATEGORY_GENERAL",
            "description": f"테스트 child ({gkey})",
            "start_timestamp": DAILY_START,
            "end_timestamp": DAILY_END,
            "reset_type": "QUEST_RESET_TYPE_DAILY",
            "count_type": "QUEST_COUNT_TYPE_SUM",
            "goal_count": 1,
            "goal_type/type/%key": gkey,
        }
        # app.py 와 동일한 규칙: param1 이 truthy 일 때만 셋팅
        if p1_input:
            child_dict["goal_type/type/%param1"] = p1_input

        append_daily_set(work["quests_path"], parent_row, [child_dict])

        # 검증
        wb = load_workbook(work["quests_path"], data_only=True)
        ws = wb["quests"]
        hm = _build_hm_quests(ws)
        key_col = hm["^key"]
        gkey_col = hm["goal_type/type/%key"]
        gp1_col = hm["goal_type/type/%param1"]

        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=key_col).value == ck:
                actual_gkey = ws.cell(row=r, column=gkey_col).value
                actual_p1 = ws.cell(row=r, column=gp1_col).value
                assert actual_gkey == gkey, (
                    f"child goal_type/%key: {actual_gkey!r} != {gkey!r}"
                )
                if p1_expected is None:
                    assert actual_p1 in (None, ""), (
                        f"child goal_type/%param1: 기대 비어있음, 실제={actual_p1!r}"
                    )
                else:
                    assert actual_p1 == p1_expected, (
                        f"child goal_type/%param1: {actual_p1!r} != {p1_expected!r}"
                    )
                break
        else:
            pytest.fail(f"child ^key={ck} 행 못 찾음")
        wb.close()


class TestNdayReproduceMixedGoalTypes:
    """가정의 달 7일치 + 하루 child 5개를 각기 다른 goal_type 으로 구성해서 저장.

    사용자 요구: 다양한 goal_type 이 한 번에 쓰여도 param1 규칙이 모두 지켜져야 함.
    """

    MIXED_CHILDREN = [
        # (goal_type_key, goal_type_param1) — 하루 5개 child
        ("play", None),
        ("play:need_win", "TRUE"),
        ("play_report:key", "PLAY_REPORT_KEY_ULTIMATE_DAMAGE"),
        ("open_battle_box:box_type", "BOX_TYPE_BATTLE"),
        ("pull_gacha:gacha_type", "GACHA_TYPE_COOKIE_GACHA"),
    ]

    @pytest.fixture
    def work(self, tmp_path):
        q_work = tmp_path / "quests_work.xlsx"
        n_work = tmp_path / "nday_work.xlsx"
        shutil.copy(QUESTS_SRC, q_work)
        shutil.copy(NDAY_SRC, n_work)
        return {"quests_path": str(q_work), "nday_path": str(n_work)}

    def _build_mixed_day_data(self) -> list[dict]:
        days: list[dict] = []
        for d in range(N_DAYS):
            day_num = d + 1
            children = []
            for i, (gkey, p1) in enumerate(self.MIXED_CHILDREN):
                children.append({
                    "description": f"{day_num}일차 미션 {i + 1} ({gkey})",
                    "goal_type_key": gkey,
                    "goal_type_param1": p1,
                    "goal_count": 3,
                })
            days.append({
                "parent_description": f"[데일리미션]{day_num}일차 전체 퀘스트 완료 보상",
                "event_description": EVENT_NAME,
                "event_day_description": f"데일리 미션 {day_num}일차",
                "children": children,
            })
        return days

    def _perform_mixed_save(self, quests_path: str, nday_path: str) -> dict:
        days = self._build_mixed_day_data()
        existing_keys = get_existing_keys(quests_path)
        existing_by_filter = get_existing_keys_by_filter(quests_path)
        first_parent = suggest_next_parent_key(
            existing_keys,
            filter_id=DAILY_FILTER,
            reset_type="QUEST_RESET_TYPE_DAILY",
            category="QUEST_CATEGORY_GENERAL",
            existing_by_filter=existing_by_filter,
        )
        parent_keys = [first_parent + 10 * i for i in range(N_DAYS)]

        existing_event_keys = get_existing_event_keys(nday_path)
        used_groups = {(ek // 100) * 100 for ek in existing_event_keys if ek >= 100}
        group_base = (max(used_groups) + 100) if used_groups else 100
        event_keys = [group_base + 1 + i for i in range(N_DAYS)]

        child_matrix: list[list[int]] = []
        existing_accum = set(existing_keys)
        n_children = len(self.MIXED_CHILDREN)
        for d in range(N_DAYS):
            existing_accum.add(parent_keys[d])
            cks = allocate_child_keys(existing_accum, parent_keys[d], n_children)
            existing_accum.update(cks)
            child_matrix.append(cks)

        for d in range(N_DAYS):
            day = days[d]
            pk = parent_keys[d]
            ek = event_keys[d]
            cks = child_matrix[d]

            parent_row = {
                "^key": pk,
                "$filter": DAILY_FILTER,
                "category": "QUEST_CATEGORY_GENERAL",
                "description": day["parent_description"],
                "start_timestamp": DAILY_START,
                "end_timestamp": DAILY_END,
                "reset_type": "QUEST_RESET_TYPE_DAILY",
                "count_type": "QUEST_COUNT_TYPE_HIGHEST",
                "goal_count": n_children,
                "goal_type/type/%key": "reward_quest:ref_quest_ids",
            }
            child_rows = []
            for i in range(n_children):
                c = day["children"][i]
                ck = cks[i]
                cd: dict = {
                    "^key": ck,
                    "$filter": DAILY_FILTER,
                    "category": "QUEST_CATEGORY_GENERAL",
                    "description": c["description"],
                    "start_timestamp": DAILY_START,
                    "end_timestamp": DAILY_END,
                    "reset_type": "QUEST_RESET_TYPE_DAILY",
                    "count_type": "QUEST_COUNT_TYPE_SUM",
                    "goal_count": c["goal_count"],
                    "goal_type/type/%key": c["goal_type_key"],
                }
                # app.py 규칙: truthy 한 p1 만 셋팅
                if c.get("goal_type_param1"):
                    cd["goal_type/type/%param1"] = c["goal_type_param1"]
                child_rows.append(cd)

            append_daily_set(quests_path, parent_row, child_rows)
            append_nday_mission_event(
                nday_path,
                {
                    "^key": ek,
                    "description": day["event_description"],
                    "start_timestamp": DAILY_START,
                    "end_timestamp": DAILY_END,
                    "mission_active_days": MISSION_ACTIVE_DAYS,
                },
                {
                    "^key": ek,
                    "day": 1,
                    "description": day["event_day_description"],
                    "quest_ids": "[]{" + ",".join(str(k) for k in cks) + "}",
                    "finish_quest_id": pk,
                },
            )

        return {
            "parent_keys": parent_keys,
            "event_keys": event_keys,
            "child_matrix": child_matrix,
            "group_base": group_base,
        }

    def test_mixed_goal_types_all_params_written_correctly(self, work):
        """7일치 × 5 child = 35 child 각각이 자기 goal_type 에 맞는 param1 을 가짐."""
        result = self._perform_mixed_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["quests_path"], data_only=True)
        ws = wb["quests"]
        hm = _build_hm_quests(ws)
        key_col = hm["^key"]
        gkey_col = hm["goal_type/type/%key"]
        gp1_col = hm["goal_type/type/%param1"]

        # 모든 row 를 한 번에 dict 로 만들어두고 검사
        rows_by_key: dict[int, dict] = {}
        for r in range(4, ws.max_row + 1):
            v = ws.cell(row=r, column=key_col).value
            if isinstance(v, int):
                rows_by_key[v] = {
                    "goal_type_key": ws.cell(row=r, column=gkey_col).value,
                    "goal_type_param1": ws.cell(row=r, column=gp1_col).value,
                }
        wb.close()

        for d in range(N_DAYS):
            for i, (gkey_expected, p1_expected) in enumerate(self.MIXED_CHILDREN):
                ck = result["child_matrix"][d][i]
                assert ck in rows_by_key, f"day{d + 1} child{i + 1} ^key={ck} 행 없음"
                row = rows_by_key[ck]
                assert row["goal_type_key"] == gkey_expected, (
                    f"day{d + 1} child{i + 1}: goal_type/%key={row['goal_type_key']!r} "
                    f"!= {gkey_expected!r}"
                )
                if p1_expected is None:
                    assert row["goal_type_param1"] in (None, ""), (
                        f"day{d + 1} child{i + 1}: param1 비어있어야 함, 실제={row['goal_type_param1']!r}"
                    )
                else:
                    assert row["goal_type_param1"] == p1_expected, (
                        f"day{d + 1} child{i + 1}: param1={row['goal_type_param1']!r} "
                        f"!= {p1_expected!r}"
                    )

    def test_mixed_goal_types_nday_still_correct(self, work):
        """mixed goal_type 저장이어도 nday_mission_events 는 규칙대로 기록됨."""
        result = self._perform_mixed_save(work["quests_path"], work["nday_path"])

        wb = load_workbook(work["nday_path"], data_only=True)
        ws_ev = wb["events"]
        ws_day = wb["events.day"]
        hm_ev = _build_hm_events(ws_ev)
        hm_day = _build_hm_events(ws_day)

        # events 시트에 event_keys 모두 존재 + description 동일
        ev_by_key: dict[int, str] = {}
        for r in range(4, ws_ev.max_row + 1):
            k = ws_ev.cell(row=r, column=hm_ev["^key"]).value
            desc = ws_ev.cell(row=r, column=hm_ev["description"]).value
            if isinstance(k, int):
                ev_by_key[k] = desc

        for ek in result["event_keys"]:
            assert ek in ev_by_key, f"events ^key={ek} 없음"
            assert ev_by_key[ek] == EVENT_NAME

        # events.day 시트에 각 event 의 quest_ids 가 child_matrix 와 일치
        day_by_key: dict[int, dict] = {}
        for r in range(4, ws_day.max_row + 1):
            k = ws_day.cell(row=r, column=hm_day["^key"]).value
            if isinstance(k, int):
                day_by_key[k] = {
                    "quest_ids": ws_day.cell(row=r, column=hm_day["quest_ids"]).value,
                    "finish_quest_id": ws_day.cell(row=r, column=hm_day["finish_quest_id"]).value,
                }
        wb.close()

        for d, ek in enumerate(result["event_keys"]):
            assert ek in day_by_key, f"events.day ^key={ek} 없음"
            expected_qids = "[]{" + ",".join(str(k) for k in result["child_matrix"][d]) + "}"
            assert day_by_key[ek]["quest_ids"] == expected_qids
            assert day_by_key[ek]["finish_quest_id"] == result["parent_keys"][d]
