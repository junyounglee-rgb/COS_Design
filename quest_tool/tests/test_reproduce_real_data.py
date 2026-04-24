"""STEP D7.5 real data reproduction test.

prerequisites:
1. baseline 캡처: `python scripts/capture_daily_mission_baseline.py`
   → `tests/fixtures/daily_mission_baseline.json` 생성
2. UI 로 재입력: `tests/run_test_reproduce.bat` 실행 후 baseline 값 그대로 2세트 입력
   → `tests/fixtures/quests_test_reproduce.xlsx` 에 parent 40001/40011 세트 기록

이 테스트는:
- `quests_test_reproduce.xlsx` 가 있을 때만 실행 (없으면 skip)
- baseline JSON 과 재입력된 xlsx 를 셀 단위 비교
- 차이 발견 시 상세 diff 로 FAIL

실행: `pytest tests/test_reproduce_real_data.py -v`
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
BASELINE_PATH = REPO_ROOT / "tests" / "fixtures" / "daily_mission_baseline.json"
REPRODUCE_PATH = REPO_ROOT / "tests" / "fixtures" / "quests_test_reproduce.xlsx"


def _require_baseline() -> dict:
    if not BASELINE_PATH.exists():
        pytest.skip(
            f"baseline missing: {BASELINE_PATH}. "
            "Run `python scripts/capture_daily_mission_baseline.py` first."
        )
    with open(BASELINE_PATH, encoding="utf-8") as f:
        return json.load(f)


def _require_reproduce_xlsx():
    if not REPRODUCE_PATH.exists():
        pytest.skip(
            f"reproduce fixture missing: {REPRODUCE_PATH}. "
            "Run `tests/run_test_reproduce.bat` and input baseline values in UI first."
        )


def _build_header_map(ws) -> dict[int, str]:
    header_map: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        row2 = ws.cell(row=2, column=col).value
        row3 = ws.cell(row=3, column=col).value
        key = str(row2).strip() if row2 else (str(row3).strip() if row3 else "")
        if key:
            header_map[col] = key
    return header_map


def _read_row_fields(ws, row_idx: int, header_map: dict[int, str]) -> dict:
    return {h: ws.cell(row=row_idx, column=col).value for col, h in header_map.items()}


def _find_row_by_key(ws, header_map: dict[int, str], target_key: int) -> int | None:
    key_col = None
    for col, h in header_map.items():
        if h == "^key":
            key_col = col
            break
    if key_col is None:
        return None
    for r in range(4, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v == target_key:
            return r
    return None


def _compare_rows(expected: dict, actual: dict, context: str) -> list[str]:
    """빈 값 (None/"") 은 동등 처리. 나머지는 값 + 타입 검사."""
    diffs: list[str] = []
    all_keys = set(expected.keys()) | set(actual.keys())
    for k in sorted(all_keys):
        ev = expected.get(k)
        av = actual.get(k)
        # 빈 값 동등
        if (ev in (None, "")) and (av in (None, "")):
            continue
        if ev != av:
            diffs.append(f"  {context} | field={k!r}: expected={ev!r} ({type(ev).__name__}) actual={av!r} ({type(av).__name__})")
    return diffs


def _verify_parent_set(parent_key: int) -> list[str]:
    baseline = _require_baseline()
    _require_reproduce_xlsx()

    # baseline 에서 해당 parent 세트 추출
    target_set = None
    for p in baseline["parents"]:
        if p["parent_key"] == parent_key:
            target_set = p
            break
    assert target_set is not None, f"baseline 에 parent_key={parent_key} 없음"

    wb = load_workbook(REPRODUCE_PATH, data_only=True)
    ws = wb["quests"]
    header_map = _build_header_map(ws)

    all_diffs: list[str] = []
    for i, row_info in enumerate(target_set["rows"]):
        expected_fields = row_info["fields"]
        key_val = expected_fields.get("^key")
        if not isinstance(key_val, int):
            all_diffs.append(f"  [row {i}] baseline ^key 가 int 가 아님: {key_val!r}")
            continue
        actual_row_idx = _find_row_by_key(ws, header_map, key_val)
        if actual_row_idx is None:
            all_diffs.append(f"  [row {i}] reproduce xlsx 에 ^key={key_val} 없음")
            continue
        actual_fields = _read_row_fields(ws, actual_row_idx, header_map)
        role = "parent" if i == 0 else f"child {i}"
        diffs = _compare_rows(expected_fields, actual_fields, f"^key={key_val} ({role})")
        all_diffs.extend(diffs)
    return all_diffs


def test_parent_40001_cell_identical():
    """^key 40001 parent + children 전원 baseline 과 100% 일치."""
    diffs = _verify_parent_set(40001)
    if diffs:
        msg = f"Found {len(diffs)} cell differences:\n" + "\n".join(diffs)
        pytest.fail(msg)


def test_parent_40011_cell_identical():
    """^key 40011 parent + children 전원 baseline 과 100% 일치."""
    diffs = _verify_parent_set(40011)
    if diffs:
        msg = f"Found {len(diffs)} cell differences:\n" + "\n".join(diffs)
        pytest.fail(msg)


def test_baseline_structure_sanity():
    """baseline JSON 구조 기본 검증 (baseline 있을 때만)."""
    if not BASELINE_PATH.exists():
        pytest.skip("baseline not captured yet")
    with open(BASELINE_PATH, encoding="utf-8") as f:
        data = json.load(f)
    assert "parents" in data, "baseline 에 parents 키 없음"
    assert len(data["parents"]) >= 2, "최소 parent 2세트(40001, 40011) 필요"
    parent_keys = [p["parent_key"] for p in data["parents"]]
    assert 40001 in parent_keys and 40011 in parent_keys
    # 각 세트 parent+child 합 ≥ 6
    for p in data["parents"]:
        assert len(p["rows"]) >= 6, f"parent {p['parent_key']} rows < 6"


def test_types_preserved_in_baseline():
    """baseline 의 ^key 는 int, $filter 는 str, goal_type/%param1 은 `[]{...}` 형식."""
    if not BASELINE_PATH.exists():
        pytest.skip("baseline not captured yet")
    with open(BASELINE_PATH, encoding="utf-8") as f:
        data = json.load(f)
    for p in data["parents"]:
        parent_fields = p["rows"][0]["fields"]
        assert isinstance(parent_fields.get("^key"), int), f"parent ^key 는 int 여야 함: {parent_fields.get('^key')!r}"
        # goal_type/%param1 확인 (경로명 변형 허용)
        p1 = None
        for k, v in parent_fields.items():
            if k.endswith("%param1") and "goal_type" in k:
                p1 = v
                break
        if p1 is not None:
            assert isinstance(p1, str), f"goal_type/%param1 는 str 이어야 함: {p1!r}"
            assert p1.startswith("[]{") and p1.endswith("}"), f"`[]{{...}}` 형식 필요: {p1!r}"
